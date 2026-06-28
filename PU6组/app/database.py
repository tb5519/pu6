import csv
import hashlib
import io
import json
import os
import re
import uuid
from calendar import monthrange
from contextlib import contextmanager
from datetime import datetime, timedelta
from threading import Lock

from flask import Blueprint, current_app, g, jsonify, request
from openpyxl import load_workbook

from app.auth import login_required
from app.classes import (
    calculate_monthly_completion,
    classify_student_habit,
    current_title_week_number,
    get_student_weeks,
    parse_completion,
)
from app.daily import field_value, parse_count
from app.teachers import TEACHERS, infer_teacher_id_from_class_name, normalize_teacher_id, teacher_id_for_username, teacher_label


database_bp = Blueprint("database", __name__, url_prefix="/api/database")
COMPLETION_SNAPSHOTS_LOCK = Lock()
DATABASE_SETTINGS_LOCK = Lock()
REMINDER_ACTIONS_LOCK = Lock()
MONTHLY_ARCHIVES_LOCK = Lock()
REMINDER_PLANS_LOCK = Lock()
LEARNING_TARGET_RATES = {0.26, 0.28, 0.3}
DEFAULT_LEARNING_TARGET_RATE = 0.26
COMPLETION_UPLOAD_TEACHER_ID = "wenyun_joanna"
LEARNING_EXCLUDED_TEACHER_IDS = {"wenyun_joanna"}
COMPLETION_UPLOAD_HEADER_SCAN_LIMIT = 20
GMV_UNIT_PRICE = 2980
GMV_SECTIONS = {
    "renewal": {"label": "续费GMV", "field": "renewal_orders"},
    "referral": {"label": "转介绍GMV", "field": "referral_conversions"},
}
COMPLETION_PERFORMANCE_TIERS = [
    {"level": 1, "label": "一档", "reward": 440, "factor": 1},
    {"level": 2, "label": "二档", "reward": 400, "factor": 0.96},
    {"level": 3, "label": "三档", "reward": 360, "factor": 0.92},
    {"level": 4, "label": "四档", "reward": 330, "factor": 0.88},
    {"level": 5, "label": "五档", "reward": 250, "factor": 0.8},
    {"level": 6, "label": "六档", "reward": 200, "factor": 0.75},
]
COMPLETION_PERFORMANCE_BASE_TARGETS = {
    1: 97,
    2: 97.5,
    3: 98,
    4: 98,
    5: 97,
    6: 98.5,
    7: 98,
    8: 96.5,
    9: 96,
    10: 95.5,
    11: 95,
    12: 94.5,
    13: 94,
    14: 93.5,
    15: 93,
    16: 92.5,
    17: 91.5,
    18: 90.5,
    19: 89.5,
    20: 88.5,
    21: 87,
    22: 86,
    23: 84,
    24: 84,
    25: 83,
    26: 82.5,
    27: 82,
    28: 81.5,
    29: 80,
    30: 79.5,
    31: 79,
    32: 78,
    33: 76.5,
    34: 75,
    35: 72.5,
    36: 71.5,
    37: 70,
    38: 69,
    39: 67,
    40: 66,
}
REMINDER_ASSESSMENT_WEEK_LIMIT = 41
REMINDER_HIGH_GAP_THRESHOLD = 10
REMINDER_DAILY_MIN_CLASSES = 2
REMINDER_PLAN_RULE_VERSION = "home-class-source-v2"
REMINDER_WEEK_FLOW = [
    {"key": "monday", "label": "周一", "new_count": REMINDER_DAILY_MIN_CLASSES, "recover_from": None, "focus_count": 0},
    {"key": "tuesday", "label": "周二", "new_count": REMINDER_DAILY_MIN_CLASSES, "recover_from": None, "focus_count": 0},
    {"key": "wednesday", "label": "周三", "new_count": REMINDER_DAILY_MIN_CLASSES, "recover_from": "monday", "focus_count": 0},
    {"key": "thursday", "label": "周四", "new_count": REMINDER_DAILY_MIN_CLASSES, "recover_from": "tuesday", "focus_count": 0},
    {"key": "friday", "label": "周五", "new_count": REMINDER_DAILY_MIN_CLASSES, "recover_from": None, "focus_count": 0},
]
REMINDER_FLOW_BY_KEY = {item["key"]: item for item in REMINDER_WEEK_FLOW}
REMINDER_RECOVERY_TARGET_BY_ORIGIN = {
    item["recover_from"]: item["key"]
    for item in REMINDER_WEEK_FLOW
    if item.get("recover_from")
}

COMPLETION_CATEGORIES = ["完课超赞", "异常断课", "断续上课", "长期不上课", "周末欠缺", "偶尔断课", "暂无数据"]
def reminder_local_class_keys(item):
    keys = set()
    for value in (item.get("name", ""), item.get("note", "")):
        keys.update(reminder_class_match_keys(value))
    return keys


def reminder_week_is_assessed(week_number):
    if week_number is None:
        return True
    try:
        week = int(week_number)
    except (TypeError, ValueError):
        return True
    return week in COMPLETION_PERFORMANCE_BASE_TARGETS


DATABASE_METRICS = {
    "learning": {"field": "learning_status", "label": "学情"},
    "renewal": {"field": "renewal_orders", "label": "续费单量"},
}
REFERRAL_FIELDS = {
    "leads": "referral_leads",
    "conversions": "referral_conversions",
}
COMPLETION_UPLOAD_COLUMNS = {
    "class_name": {"班级名称", "班级名", "班级", "课程班级", "班级名称班主任", "classname", "class"},
    "teacher": {"班主任", "老师", "教师", "带班老师", "班主任老师", "teacher"},
    "student_count": {"在班学员数", "在班人数", "学员数", "学生数", "班级人数", "带班人数", "studentcount", "students"},
    "completion_rate": {
        "完成度",
        "完课率",
        "完成率",
        "本月完成度",
        "本月完课率",
        "整体完成度",
        "平均完课",
        "completion",
    },
    "last_month_completion": {
        "上个月完课率",
        "上月完课率",
        "上个月完成度",
        "上月完成度",
        "上月完课",
        "上月完成率",
        "lastmonthcompletion",
    },
}


def classes_file():
    return current_app.config["CLASSES_FILE"]


def users_file():
    return current_app.config["USERS_FILE"]


def completion_assignments_file():
    return current_app.config["COMPLETION_ASSIGNMENTS_FILE"]


def completion_snapshots_file():
    return current_app.config["COMPLETION_SNAPSHOTS_FILE"]


def completion_reminder_actions_file():
    return current_app.config["COMPLETION_REMINDER_ACTIONS_FILE"]


def completion_reminder_plans_file():
    return current_app.config["COMPLETION_REMINDER_PLANS_FILE"]


def daily_report_file():
    return current_app.config["DAILY_REPORT_FILE"]


def database_settings_file():
    return current_app.config["DATABASE_SETTINGS_FILE"]


def monthly_archives_file():
    return current_app.config["MONTHLY_ARCHIVES_FILE"]


def load_json(path, default):
    if not path.exists():
        return default
    with path.open("r", encoding="utf-8") as file:
        return json.load(file)


@contextmanager
def json_file_lock(path):
    lock_path = path.with_suffix(f"{path.suffix}.lock")
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    with lock_path.open("a+", encoding="utf-8") as lock_file:
        if os.name == "nt":
            import msvcrt

            lock_file.seek(0)
            if not lock_file.read(1):
                lock_file.write("0")
                lock_file.flush()
            lock_file.seek(0)
            msvcrt.locking(lock_file.fileno(), msvcrt.LK_LOCK, 1)
            try:
                yield
            finally:
                lock_file.seek(0)
                msvcrt.locking(lock_file.fileno(), msvcrt.LK_UNLCK, 1)
        else:
            import fcntl

            fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX)
            try:
                yield
            finally:
                fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)


def write_json_atomic(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(f"{path.name}.{uuid.uuid4().hex}.tmp")
    with temp_path.open("w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=2)
        file.flush()
        os.fsync(file.fileno())
    temp_path.replace(path)


def save_json(path, data):
    with json_file_lock(path):
        write_json_atomic(path, data)


def update_json(path, default, updater):
    with json_file_lock(path):
        store = load_json(path, default)
        result = updater(store)
        write_json_atomic(path, store)
    return result


def load_completion_assignments():
    store = load_json(completion_assignments_file(), {})
    return [
        item
        for item in store.get("classes", [])
        if str(item.get("name", "")).strip()
    ]


def load_completion_snapshots():
    store = load_json(completion_snapshots_file(), {})
    store.setdefault("snapshots", {})
    store.setdefault("last_month", {})
    return store


def save_completion_snapshots(store):
    save_json(completion_snapshots_file(), store)


def load_reminder_actions():
    store = load_json(completion_reminder_actions_file(), {})
    store.setdefault("records", [])
    return store


def save_reminder_actions(store):
    save_json(completion_reminder_actions_file(), store)


def load_reminder_plans():
    store = load_json(completion_reminder_plans_file(), {})
    store.setdefault("plans", {})
    return store


def save_reminder_plans(store):
    save_json(completion_reminder_plans_file(), store)


def load_database_settings():
    settings = load_json(database_settings_file(), {})
    learning = settings.setdefault("learning", {})
    learning.setdefault("classes", {})
    learning.setdefault("teachers", {})
    settings.setdefault("gmv", {})
    return settings


def save_database_settings(settings):
    save_json(database_settings_file(), settings)


def load_monthly_archives():
    store = load_json(monthly_archives_file(), {})
    store.setdefault("archives", {})
    return store


def save_monthly_archives(store):
    save_json(monthly_archives_file(), store)


def normalize_month(value):
    raw_value = str(value or "").strip()
    if not raw_value:
        return datetime.now().strftime("%Y-%m")
    try:
        return datetime.strptime(raw_value, "%Y-%m").strftime("%Y-%m")
    except ValueError:
        raise ValueError("统计月份格式不正确。") from None


def normalize_date(value):
    raw_value = str(value or "").strip()
    if not raw_value:
        return datetime.now().strftime("%Y-%m-%d")
    try:
        return datetime.strptime(raw_value, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        raise ValueError("统计日期格式不正确。") from None


def month_end_date(month_key):
    year, month = [int(part) for part in month_key.split("-")]
    return f"{month_key}-{monthrange(year, month)[1]:02d}"


def month_week_index(date_text):
    day = int(str(date_text)[-2:])
    if day <= 7:
        return 0
    if day <= 14:
        return 1
    if day <= 21:
        return 2
    return 3


def previous_month_key(month_key):
    year, month = [int(part) for part in month_key.split("-")]
    if month == 1:
        return f"{year - 1}-12"
    return f"{year}-{month - 1:02d}"


def next_month_key(month_key):
    year, month = [int(part) for part in month_key.split("-")]
    if month == 12:
        return f"{year + 1}-01"
    return f"{year}-{month + 1:02d}"


def next_month_start_date(month_key):
    return f"{next_month_key(month_key)}-01"


def current_user_teacher_id():
    user = getattr(g, "user", None)
    if not user:
        return ""
    return (
        normalize_teacher_id(user.get("teacher_id"))
        or normalize_teacher_id(user.get("username"))
        or teacher_id_for_username(user.get("username"))
    )


def can_upload_completion_data():
    return current_user_teacher_id() == COMPLETION_UPLOAD_TEACHER_ID


def can_manage_gmv():
    return current_user_teacher_id() == COMPLETION_UPLOAD_TEACHER_ID


def blank_category_counts():
    return {category: 0 for category in COMPLETION_CATEGORIES}


def average(values):
    valid_values = [value for value in values if value is not None]
    if not valid_values:
        return None
    return round(sum(valid_values) / len(valid_values), 2)


def teacher_id_for_class_owner(owner):
    owner_text = str(owner or "").strip()
    if not owner_text:
        return ""
    users = load_json(users_file(), {})
    user = users.get(owner_text)
    if not user:
        normalized_owner = owner_text.lower()
        user = next(
            (
                item
                for username, item in users.items()
                if str(username or "").strip().lower() == normalized_owner
            ),
            None,
        )
    if not user:
        return teacher_id_for_username(owner_text)
    return (
        normalize_teacher_id(user.get("teacher_id"))
        or normalize_teacher_id(user.get("username"))
        or teacher_id_for_username(user.get("username"))
    )


def class_teacher_id(item):
    return (
        normalize_teacher_id(item.get("teacher_id"))
        or teacher_id_for_class_owner(item.get("owner"))
        or infer_teacher_id_from_class_name(item.get("name"))
    )


def parse_float(value, default=0):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    if number < 0:
        return default
    return number


def parse_gmv_amount(value):
    if value is None or value == "":
        return None
    return rounded_metric(parse_float(value, 0))


def normalize_target_rate(value):
    rate = parse_float(value, DEFAULT_LEARNING_TARGET_RATE)
    return rate if rate in LEARNING_TARGET_RATES else DEFAULT_LEARNING_TARGET_RATE


def rounded_metric(value):
    return round(float(value or 0), 2)


def normalize_completion_header(value):
    text = str(value or "").replace("\u3000", " ").strip().lower()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[：:（）()\[\]【】{}<>《》\"“”'‘’、,，.。/\\|_\-]+", "", text)
    return text.replace("必填", "").replace("选填", "")


def pick_upload_column(headers, candidates):
    normalized = [normalize_completion_header(header) for header in headers]
    candidate_values = [normalize_completion_header(candidate) for candidate in candidates]
    lookup = {header: index for index, header in enumerate(normalized) if header}
    for candidate in candidate_values:
        index = lookup.get(candidate)
        if index is not None:
            return index
    for index, header in enumerate(normalized):
        if not header:
            continue
        for candidate in candidate_values:
            if candidate and candidate in header:
                return index
    return None


def row_value(row, index):
    if index is None or index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else str(value).strip()


def non_empty_upload_cells(row):
    return [str(value).strip() for value in row if str(value or "").strip()]


def readable_upload_headers(rows):
    candidates = []
    for row in rows[:COMPLETION_UPLOAD_HEADER_SCAN_LIMIT]:
        cells = non_empty_upload_cells(row)
        if cells:
            candidates.append("、".join(cells[:10]))
    return "；".join(candidates[:5])


def find_completion_upload_header(rows):
    best = None
    for row_index, row in enumerate(rows[:COMPLETION_UPLOAD_HEADER_SCAN_LIMIT]):
        headers = [str(value or "").strip() for value in row]
        if not any(headers):
            continue
        indexes = {
            key: pick_upload_column(headers, candidates)
            for key, candidates in COMPLETION_UPLOAD_COLUMNS.items()
        }
        score = 0
        score += 8 if indexes["class_name"] is not None else 0
        score += 6 if indexes["teacher"] is not None else 0
        score += 6 if indexes["student_count"] is not None else 0
        score += 8 if indexes["completion_rate"] is not None else 0
        score += 3 if indexes["last_month_completion"] is not None else 0
        if best is None or score > best["score"]:
            best = {"score": score, "row_index": row_index, "indexes": indexes}
    return best


def completion_upload_missing_error(column_name, examples, rows):
    preview = readable_upload_headers(rows)
    message = f"表格需要包含“{column_name}”列，常见列名：{examples}。"
    if preview:
        message += f" 已读取到的前几行内容：{preview}"
    return message


def parse_student_count(value):
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return max(0, int(value))
    text = str(value).replace(",", "").strip()
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return 0
    return max(0, int(float(match.group(0))))


def completion_class_id(class_name):
    normalized = normalize_completion_header(class_name)
    digest = hashlib.sha1(normalized.encode("utf-8")).hexdigest()[:16]
    return f"completion-{digest}"


def teacher_id_from_upload(teacher_text, class_name):
    return (
        normalize_teacher_id(teacher_text)
        or teacher_id_for_username(teacher_text)
        or infer_teacher_id_from_class_name(teacher_text)
        or infer_teacher_id_from_class_name(class_name)
    )


def rows_to_completion_snapshot(rows):
    if not rows:
        return []
    header = find_completion_upload_header(rows)
    indexes = header["indexes"] if header else {}
    header_row_index = header["row_index"] if header else 0

    if indexes.get("class_name") is None:
        raise ValueError(completion_upload_missing_error("班级名称", "班级名称、班级、班级名", rows))
    if indexes.get("student_count") is None:
        raise ValueError(completion_upload_missing_error("在班学员数", "在班学员数、学员数、在班人数、班级人数", rows))
    if indexes.get("completion_rate") is None:
        raise ValueError(completion_upload_missing_error("完成度", "完成度、完课率、完成率、本月完课率", rows))

    snapshot_rows = []
    seen_class_ids = set()
    for row in rows[header_row_index + 1:]:
        class_name = row_value(row, indexes["class_name"])
        if not class_name:
            continue
        completion_rate = parse_completion(row[indexes["completion_rate"]] if indexes["completion_rate"] < len(row) else None)
        if completion_rate is None:
            continue
        teacher_text = row_value(row, indexes.get("teacher"))
        teacher_id = teacher_id_from_upload(teacher_text, class_name)
        class_id = completion_class_id(class_name)
        if class_id in seen_class_ids:
            continue
        seen_class_ids.add(class_id)
        last_month_index = indexes.get("last_month_completion")
        last_month_completion = (
            parse_completion(row[last_month_index] if last_month_index is not None and last_month_index < len(row) else None)
            if last_month_index is not None
            else None
        )
        snapshot_rows.append(
            {
                "id": class_id,
                "name": class_name,
                "teacher_id": teacher_id,
                "teacher_name": teacher_label(teacher_id) or teacher_text or "未分配",
                "student_count": parse_student_count(row[indexes["student_count"]] if indexes["student_count"] < len(row) else None),
                "completion_rate": completion_rate,
                "last_month_completion": last_month_completion,
            }
        )

    if not snapshot_rows:
        raise ValueError("没有读取到班级完课数据，请确认表头下面存在班级记录。")
    return snapshot_rows


def parse_completion_upload_csv(file_storage):
    raw = file_storage.stream.read()
    for encoding in ("utf-8-sig", "gb18030", "gbk"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("CSV 编码无法识别，请另存为 UTF-8 或 Excel .xlsx 后再上传。")
    return list(csv.reader(io.StringIO(text)))


def parse_completion_upload_xlsx(file_storage):
    workbook = load_workbook(file_storage, data_only=True)
    errors = []
    for sheet in workbook.worksheets:
        rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        try:
            return rows_to_completion_snapshot(rows)
        except ValueError as error:
            errors.append(f"{sheet.title}：{error}")
    detail = errors[0] if errors else "没有读取到工作表内容。"
    raise ValueError(f"未在 Excel 工作表中识别到可导入数据。{detail}")


def parse_completion_upload(file_storage):
    filename = (file_storage.filename or "").lower()
    if filename.endswith(".csv"):
        return rows_to_completion_snapshot(parse_completion_upload_csv(file_storage))
    if filename.endswith(".xlsx"):
        return parse_completion_upload_xlsx(file_storage)
    raise ValueError("仅支持 .xlsx 或 .csv 文件。")


def build_class_completion_summary(month_key):
    store = load_json(classes_file(), {"classes": []})
    teacher_lookup = {
        teacher["id"]: {
            "teacher_id": teacher["id"],
            "teacher_name": teacher["name"],
            "class_count": 0,
            "student_count": 0,
            "active_student_count": 0,
            "completion_values": [],
            "category_counts": blank_category_counts(),
        }
        for teacher in TEACHERS
    }

    classes = []
    total_values = []
    total_categories = blank_category_counts()
    total_students = 0
    active_students = 0

    for item in store.get("classes", []):
        students = item.get("students", [])
        teacher_id = class_teacher_id(item)
        category_counts = blank_category_counts()
        completion_values = []
        active_count = 0

        for student in students:
            weeks = get_student_weeks(student, month_key)
            completion = calculate_monthly_completion(weeks)
            category = classify_student_habit(weeks)
            category_counts[category] = category_counts.get(category, 0) + 1
            total_categories[category] = total_categories.get(category, 0) + 1
            if completion is not None:
                completion_values.append(completion)
                total_values.append(completion)
                active_count += 1

        student_count = len(students)
        total_students += student_count
        active_students += active_count

        teacher_summary = teacher_lookup.get(teacher_id)
        if teacher_summary:
            teacher_summary["class_count"] += 1
            teacher_summary["student_count"] += student_count
            teacher_summary["active_student_count"] += active_count
            teacher_summary["completion_values"].extend(completion_values)
            for category, count in category_counts.items():
                teacher_summary["category_counts"][category] += count

        classes.append(
            {
                "id": item.get("id", ""),
                "name": item.get("name", ""),
                "teacher_id": teacher_id,
                "teacher_name": teacher_label(teacher_id) or "未分配",
                "student_count": student_count,
                "active_student_count": active_count,
                "average_completion": average(completion_values),
                "category_counts": category_counts,
                "updated_at": item.get("updated_at", ""),
            }
        )

    teachers = []
    for teacher in teacher_lookup.values():
        values = teacher.pop("completion_values")
        teachers.append({**teacher, "average_completion": average(values)})

    return {
        "summary": {
            "class_count": len(classes),
            "student_count": total_students,
            "active_student_count": active_students,
            "average_completion": average(total_values),
            "category_counts": total_categories,
        },
        "classes": classes,
        "teachers": teachers,
    }


def snapshot_list(store=None):
    store = store or load_completion_snapshots()
    snapshots = []
    for date_key, snapshot in store.get("snapshots", {}).items():
        if isinstance(snapshot, dict):
            snapshots.append({**snapshot, "date": snapshot.get("date") or date_key})
    return sorted(snapshots, key=lambda item: item.get("date", ""))


def latest_completion_snapshot(month_key, report_date=None, store=None):
    end_date = report_date or month_end_date(month_key)
    candidates = [
        snapshot
        for snapshot in snapshot_list(store)
        if str(snapshot.get("date", "")).startswith(f"{month_key}-")
        and str(snapshot.get("date", "")) <= end_date
    ]
    return candidates[-1] if candidates else None


def completion_snapshots_until(month_key, report_date=None, store=None):
    end_date = report_date or month_end_date(month_key)
    return [
        snapshot
        for snapshot in snapshot_list(store)
        if str(snapshot.get("date", "")).startswith(f"{month_key}-")
        and str(snapshot.get("date", "")) <= end_date
    ]


def completion_snapshot_by_date(date_key, store=None):
    if not date_key:
        return None
    store = store or load_completion_snapshots()
    snapshot = store.get("snapshots", {}).get(str(date_key))
    if isinstance(snapshot, dict):
        return {**snapshot, "date": snapshot.get("date") or str(date_key)}
    return None


def previous_completion_snapshot(snapshot_date, store=None):
    candidates = [
        snapshot
        for snapshot in snapshot_list(store)
        if str(snapshot.get("date", "")) < snapshot_date
    ]
    return candidates[-1] if candidates else None


def last_month_completion_snapshot(month_key, store=None):
    store = store or load_completion_snapshots()
    target_month = previous_month_key(month_key)
    snapshot = store.get("last_month", {}).get(target_month)
    if isinstance(snapshot, dict):
        return {**snapshot, "month": snapshot.get("month") or target_month}
    return None


def rows_by_class_id(snapshot):
    if not snapshot:
        return {}
    return {row.get("id"): row for row in snapshot.get("rows", []) if row.get("id")}


def completion_lookup_key(class_name):
    return normalize_completion_header(class_name)


def alpha_zero_based_number(char):
    text = str(char or "").strip()
    if not text:
        return ""
    index = ord(text.upper()) - ord("A")
    return str(index) if index >= 0 else text


def alpha_one_based_number(char):
    text = str(char or "").strip()
    if not text:
        return ""
    index = ord(text.upper()) - ord("A") + 1
    return str(index) if index > 0 else text


def encode_period_token(token):
    return "".join(alpha_zero_based_number(char) if char.isalpha() else char for char in str(token or ""))


def encode_reminder_class_name(class_name):
    value = str(class_name or "")

    def replace_period(match):
        token = match.group(1)
        if not any(char.isalpha() for char in token):
            return match.group(0)
        return f"{encode_period_token(token)}期"

    value = re.sub(r"([A-Za-z0-9]+)期", replace_period, value)
    value = re.sub(r"-([A-Za-z])班", lambda match: f"-{alpha_one_based_number(match.group(1))}班", value)
    return value


def reminder_class_match_keys(class_name):
    raw_value = str(class_name or "").strip()
    variants = {raw_value, encode_reminder_class_name(raw_value)}
    for value in list(variants):
        variants.add(re.sub(r"L2-", "PU1-", value, flags=re.IGNORECASE))
    return {completion_lookup_key(value) for value in variants if completion_lookup_key(value)}


def rows_by_class_name(snapshot):
    if not snapshot:
        return {}
    lookup = {}
    for row in snapshot.get("rows", []):
        key = completion_lookup_key(row.get("name", ""))
        if key and key not in lookup:
            lookup[key] = row
    return lookup


def rate_delta(current, baseline):
    if current is None or baseline is None:
        return None
    return rounded_metric(float(current) - float(baseline))


def weighted_rate(rows, field):
    weighted_total = 0
    total_weight = 0
    values = []
    for row in rows:
        value = row.get(field)
        if value is None:
            continue
        values.append(value)
        weight = parse_student_count(row.get("student_count")) or 0
        if weight:
            weighted_total += float(value) * weight
            total_weight += weight
    if total_weight:
        return rounded_metric(weighted_total / total_weight)
    return average(values)


def weighted_delta(rows, field):
    weighted_total = 0
    total_weight = 0
    values = []
    for row in rows:
        value = row.get(field)
        if value is None:
            continue
        values.append(value)
        weight = parse_student_count(row.get("student_count")) or 0
        if weight:
            weighted_total += float(value) * weight
            total_weight += weight
    if total_weight:
        return rounded_metric(weighted_total / total_weight)
    return average(values)


def rounded_performance_percent(value):
    return int(float(value or 0) * 100 + 0.5) / 100


def completion_performance_thresholds(week_number):
    try:
        week = int(week_number)
    except (TypeError, ValueError):
        return []
    base_target = COMPLETION_PERFORMANCE_BASE_TARGETS.get(week)
    if base_target is None:
        return []
    return [
        {
            **tier,
            "target_rate": rounded_performance_percent(base_target * tier["factor"]),
        }
        for tier in COMPLETION_PERFORMANCE_TIERS
    ]


def completion_performance_local_lookup():
    store = load_json(classes_file(), {"classes": []})
    by_teacher = {}
    by_key = {}
    for item in store.get("classes", []):
        if not isinstance(item, dict):
            continue
        teacher_id = class_teacher_id(item)
        week_number = current_title_week_number(item)
        meta = {
            "id": item.get("id", ""),
            "name": item.get("name", ""),
            "note": str(item.get("note") or "").strip(),
            "teacher_id": teacher_id,
            "teacher_name": teacher_label(teacher_id) or "未分配",
            "student_count": len(item.get("students", []) if isinstance(item.get("students"), list) else []),
            "title_week_number": week_number,
            "title_week_label": f"W{week_number}" if week_number else "",
        }
        keys = set()
        for value in (item.get("name", ""), item.get("note", "")):
            keys.update(reminder_class_match_keys(value))
        for key in keys:
            if teacher_id:
                by_teacher.setdefault(teacher_id, {})[key] = meta
            by_key.setdefault(key, meta)
    return {"by_teacher": by_teacher, "by_key": by_key}


def completion_performance_local_meta(row, local_lookup):
    teacher_id = normalize_teacher_id(row.get("teacher_id"))
    keys = reminder_class_match_keys(row.get("name") or row.get("class_name", ""))
    for key in keys:
        meta = local_lookup.get("by_teacher", {}).get(teacher_id, {}).get(key)
        if meta:
            return meta
    for key in keys:
        meta = local_lookup.get("by_key", {}).get(key)
        if meta:
            return meta
    return None


def completion_performance_best_tier(completion_rate, thresholds):
    if completion_rate is None:
        return None
    try:
        current = float(completion_rate)
    except (TypeError, ValueError):
        return None
    for threshold in thresholds:
        if current + 0.0001 >= float(threshold["target_rate"]):
            return threshold
    return None


def completion_performance_next_tier(completion_rate, thresholds, best_tier):
    if completion_rate is None or not thresholds:
        return {"next_tier_label": "", "next_tier_target": None, "next_tier_gap": None}
    try:
        current = float(completion_rate)
    except (TypeError, ValueError):
        return {"next_tier_label": "", "next_tier_target": None, "next_tier_gap": None}
    if not best_tier:
        target = thresholds[-1]
    elif int(best_tier.get("level") or 0) <= 1:
        return {"next_tier_label": "已最高档", "next_tier_target": None, "next_tier_gap": 0}
    else:
        target = thresholds[int(best_tier["level"]) - 2]
    target_rate = float(target["target_rate"])
    return {
        "next_tier_label": target["label"],
        "next_tier_target": target["target_rate"],
        "next_tier_gap": rounded_performance_percent(max(0, target_rate - current)),
    }


def completion_performance_status(local_meta, week_number, completion_rate, thresholds, best_tier):
    if not local_meta:
        return "no_match", "未匹配我的班级"
    if not week_number:
        return "no_week", "缺少W"
    if not thresholds:
        return "not_assessed", "W41后不核算"
    if completion_rate is None:
        return "no_completion", "缺少完成度"
    if not best_tier:
        return "not_reached", "未达标"
    return "achieved", best_tier["label"]


def build_completion_performance_summary(completion):
    local_lookup = completion_performance_local_lookup()
    rows = []
    teacher_lookup = {}

    for item in completion.get("classes", []):
        local_meta = completion_performance_local_meta(item, local_lookup)
        week_number = local_meta.get("title_week_number") if local_meta else None
        thresholds = completion_performance_thresholds(week_number)
        best_tier = completion_performance_best_tier(item.get("completion_rate"), thresholds)
        next_tier = completion_performance_next_tier(item.get("completion_rate"), thresholds, best_tier)
        status, status_label = completion_performance_status(
            local_meta,
            week_number,
            item.get("completion_rate"),
            thresholds,
            best_tier,
        )
        teacher_id = normalize_teacher_id(item.get("teacher_id")) or (local_meta or {}).get("teacher_id") or ""
        teacher_name = item.get("teacher_name") or (local_meta or {}).get("teacher_name") or teacher_label(teacher_id) or "未分配"
        reward = int(best_tier.get("reward", 0)) if best_tier else 0
        counted = bool(thresholds and item.get("completion_rate") is not None)
        row = {
            "class_id": item.get("id", "") or (local_meta or {}).get("id", ""),
            "class_name": item.get("name", "") or (local_meta or {}).get("name", ""),
            "local_class_name": (local_meta or {}).get("name", ""),
            "teacher_id": teacher_id,
            "teacher_name": teacher_name,
            "student_count": parse_student_count(item.get("student_count") or (local_meta or {}).get("student_count")),
            "completion_rate": item.get("completion_rate"),
            "title_week_number": week_number,
            "title_week_label": f"W{week_number}" if week_number else "",
            "base_target": thresholds[0]["target_rate"] if thresholds else None,
            "target_rate": best_tier.get("target_rate") if best_tier else (thresholds[-1]["target_rate"] if thresholds else None),
            "tier_level": best_tier.get("level") if best_tier else None,
            "tier_label": best_tier.get("label") if best_tier else "",
            **next_tier,
            "reward": reward,
            "status": status,
            "status_label": status_label,
            "counted": counted,
            "matched": bool(local_meta),
        }
        rows.append(row)

        group_key = teacher_id or f"unknown-{hashlib.sha1(teacher_name.encode('utf-8')).hexdigest()[:10]}"
        group = teacher_lookup.setdefault(
            group_key,
            {
                "teacher_id": group_key,
                "teacher_name": teacher_name,
                "class_count": 0,
                "assessed_class_count": 0,
                "achieved_class_count": 0,
                "reward_total": 0,
                "completion_rows": [],
            },
        )
        group["class_count"] += 1
        if counted:
            group["assessed_class_count"] += 1
            group["completion_rows"].append(row)
        if best_tier:
            group["achieved_class_count"] += 1
            group["reward_total"] += reward

    teacher_order = {teacher["id"]: index for index, teacher in enumerate(TEACHERS)}
    teacher_rows = []
    for group in teacher_lookup.values():
        completion_rows = group.pop("completion_rows")
        teacher_rows.append(
            {
                **group,
                "average_completion": weighted_rate(completion_rows, "completion_rate"),
            }
        )
    teacher_rows.sort(key=lambda row: (teacher_order.get(row["teacher_id"], 999), row["teacher_name"]))
    rows.sort(
        key=lambda row: (
            teacher_order.get(row["teacher_id"], 999),
            row["teacher_name"],
            row["title_week_number"] or 999,
            row["class_name"],
        )
    )

    assessed_rows = [row for row in rows if row["counted"]]
    achieved_rows = [row for row in rows if row["tier_level"]]
    return {
        "tier_rules": COMPLETION_PERFORMANCE_TIERS,
        "summary": {
            "class_count": len(rows),
            "assessed_class_count": len(assessed_rows),
            "achieved_class_count": len(achieved_rows),
            "unassessed_class_count": len(rows) - len(assessed_rows),
            "reward_total": sum(row["reward"] for row in rows),
            "average_completion": weighted_rate(assessed_rows, "completion_rate"),
        },
        "teacher_rows": teacher_rows,
        "rows": rows,
    }


def completion_roster_classes(month_key):
    assignments = load_completion_assignments()
    if assignments:
        return [
            {
                "id": item.get("id") or completion_class_id(item.get("name", "")),
                "name": item.get("name", ""),
                "teacher_id": normalize_teacher_id(item.get("teacher_id")),
                "teacher_name": teacher_label(item.get("teacher_id")) or item.get("teacher_name") or "未分配",
                "student_count": None,
                "active_student_count": 0,
                "average_completion": None,
                "category_counts": blank_category_counts(),
                "updated_at": item.get("updated_at", ""),
            }
            for item in assignments
        ], True

    summary = build_class_completion_summary(month_key)
    by_name = {}
    for row in summary.get("classes", []):
        key = completion_lookup_key(row.get("name", ""))
        if not key:
            continue
        current = by_name.get(key)
        should_replace = (
            current is None
            or (not current.get("teacher_id") and row.get("teacher_id"))
            or parse_student_count(row.get("student_count")) > parse_student_count(current.get("student_count"))
        )
        if should_replace:
            by_name[key] = row
    return list(by_name.values()), False


def build_completion_snapshot_summary(month_key, snapshot, store=None, compare_date=None):
    store = store or load_completion_snapshots()
    snapshot_date = snapshot.get("date", "")
    history_snapshots = completion_snapshots_until(month_key, snapshot_date or month_end_date(month_key), store)
    if snapshot_date and not any(item.get("date") == snapshot_date for item in history_snapshots):
        history_snapshots.append(snapshot)
        history_snapshots = sorted(history_snapshots, key=lambda item: item.get("date", ""))
    history_dates = [item.get("date", "") for item in reversed(history_snapshots) if item.get("date")]
    previous_snapshot = history_snapshots[-2] if len(history_snapshots) >= 2 else None
    compare_snapshot = completion_snapshot_by_date(compare_date, store) if compare_date else None
    if (
        not compare_snapshot
        or not snapshot_date
        or compare_snapshot.get("date", "") >= snapshot_date
        or not compare_snapshot.get("date", "").startswith(f"{month_key}-")
    ):
        compare_snapshot = previous_snapshot
    previous_rows = rows_by_class_name(previous_snapshot)
    compare_rows = rows_by_class_name(compare_snapshot)
    last_month_snapshot = last_month_completion_snapshot(month_key, store)
    last_month_source_month = previous_month_key(month_key)
    if not last_month_snapshot:
        previous_month = previous_month_key(month_key)
        last_month_snapshot = latest_completion_snapshot(previous_month, month_end_date(previous_month), store)
        last_month_source_month = last_month_snapshot.get("date", "") if last_month_snapshot else previous_month
    previous_month_rows = rows_by_class_name(last_month_snapshot)
    uploaded_rows = rows_by_class_name(snapshot)
    history_row_maps = {
        item.get("date", ""): rows_by_class_name(item)
        for item in history_snapshots
        if item.get("date")
    }

    teacher_lookup = {
        teacher["id"]: {
            "teacher_id": teacher["id"],
            "teacher_name": teacher["name"],
            "class_count": 0,
            "student_count": 0,
            "active_student_count": 0,
            "completion_rows": [],
            "category_counts": blank_category_counts(),
        }
        for teacher in TEACHERS
    }

    classes = []
    total_students = 0
    active_students = 0
    matched_upload_keys = set()
    display_sources = []
    roster_classes, roster_is_authoritative = completion_roster_classes(month_key)

    for roster_row in roster_classes:
        key = completion_lookup_key(roster_row.get("name", ""))
        upload_row = uploaded_rows.get(key)
        if upload_row:
            matched_upload_keys.add(key)
        display_sources.append((roster_row, upload_row, key))

    if not roster_is_authoritative:
        for upload_row in snapshot.get("rows", []):
            key = completion_lookup_key(upload_row.get("name", ""))
            if not key or key in matched_upload_keys:
                continue
            display_sources.append((None, upload_row, key))
            matched_upload_keys.add(key)

    for roster_row, upload_row, lookup_key in display_sources:
        source_row = upload_row or roster_row or {}
        class_id = (roster_row or {}).get("id") or source_row.get("id", "")
        class_name = (roster_row or {}).get("name") or source_row.get("name", "")
        completion_rate = upload_row.get("completion_rate") if upload_row else None
        previous_row = previous_rows.get(lookup_key)
        compare_row = compare_rows.get(lookup_key)
        previous_month_row = previous_month_rows.get(lookup_key)
        previous_completion = previous_row.get("completion_rate") if previous_row else None
        compare_completion = compare_row.get("completion_rate") if compare_row else None
        last_month_completion = previous_month_row.get("completion_rate") if previous_month_row else None
        history_values = []
        for date_key in history_dates:
            history_row = history_row_maps.get(date_key, {}).get(lookup_key)
            history_values.append(
                {
                    "date": date_key,
                    "completion_rate": history_row.get("completion_rate") if history_row else None,
                }
            )

        student_count = parse_student_count(upload_row.get("student_count")) if upload_row else None
        teacher_id = normalize_teacher_id((roster_row or {}).get("teacher_id")) or normalize_teacher_id(source_row.get("teacher_id"))
        teacher_name = (
            teacher_label(teacher_id)
            or (roster_row or {}).get("teacher_name")
            or source_row.get("teacher_name")
            or "未分配"
        )
        row = {
            "id": class_id,
            "name": class_name,
            "teacher_id": teacher_id,
            "teacher_name": teacher_name,
            "student_count": student_count,
            "active_student_count": parse_student_count(student_count) if completion_rate is not None else 0,
            "average_completion": completion_rate,
            "completion_rate": completion_rate,
            "last_month_completion": last_month_completion,
            "previous_completion": previous_completion,
            "compare_completion": compare_completion,
            "compare_date": compare_snapshot.get("date", "") if compare_snapshot else "",
            "previous_snapshot_date": previous_snapshot.get("date") if previous_snapshot else "",
            "change_from_previous": rate_delta(completion_rate, previous_completion),
            "change_from_compare": rate_delta(completion_rate, compare_completion),
            "change_from_last_month": rate_delta(completion_rate, last_month_completion),
            "history": history_values,
            "lookup_matched": bool(upload_row),
            "category_counts": blank_category_counts(),
            "updated_at": snapshot.get("uploaded_at", ""),
        }
        classes.append(row)
        total_students += parse_student_count(student_count)
        if completion_rate is not None:
            active_students += parse_student_count(student_count)

        teacher_key = teacher_id or f"unknown-{hashlib.sha1(teacher_name.encode('utf-8')).hexdigest()[:10]}"
        if teacher_key not in teacher_lookup:
            teacher_lookup[teacher_key] = {
                "teacher_id": teacher_key,
                "teacher_name": teacher_name,
                "class_count": 0,
                "student_count": 0,
                "active_student_count": 0,
                "completion_rows": [],
                "category_counts": blank_category_counts(),
            }
        teacher_summary = teacher_lookup[teacher_key]
        teacher_summary["class_count"] += 1
        teacher_summary["student_count"] += parse_student_count(student_count)
        teacher_summary["active_student_count"] += row["active_student_count"]
        teacher_summary["completion_rows"].append(row)

    teachers = []
    for teacher in teacher_lookup.values():
        rows = teacher.pop("completion_rows")
        teachers.append({**teacher, "average_completion": weighted_rate(rows, "completion_rate")})

    return {
        "summary": {
            "class_count": len(classes),
            "student_count": total_students,
            "active_student_count": active_students,
            "average_completion": weighted_rate(classes, "completion_rate"),
            "category_counts": blank_category_counts(),
        },
        "classes": classes,
        "teachers": teachers,
        "source": "snapshot",
        "snapshot_date": snapshot_date,
        "uploaded_at": snapshot.get("uploaded_at", ""),
        "uploaded_by": snapshot.get("uploaded_by", ""),
        "can_upload": can_upload_completion_data(),
        "roster_source": "assignment" if roster_is_authoritative else "classes",
        "history_dates": history_dates,
        "visible_history_dates": history_dates[:2],
        "older_history_dates": history_dates[2:],
        "compare_dates": [date_key for date_key in history_dates if date_key != snapshot_date],
        "comparison": {
            "previous_snapshot_date": previous_snapshot.get("date") if previous_snapshot else "",
            "previous_change": weighted_delta(classes, "change_from_previous"),
            "compare_date": compare_snapshot.get("date", "") if compare_snapshot else "",
            "compare_change": weighted_delta(classes, "change_from_compare"),
            "last_month_source_date": last_month_snapshot.get("date") if last_month_snapshot else "",
            "last_month_source_month": last_month_source_month,
            "last_month_change": weighted_delta(classes, "change_from_last_month"),
        },
    }


def build_completion_summary(month_key, report_date=None, compare_date=None):
    store = load_completion_snapshots()
    snapshot = latest_completion_snapshot(month_key, report_date or month_end_date(month_key), store)
    if snapshot:
        return build_completion_snapshot_summary(month_key, snapshot, store, compare_date)

    if load_completion_assignments():
        summary = build_completion_snapshot_summary(month_key, {"date": "", "rows": []}, store, compare_date)
        summary["source"] = "assignment"
        summary["snapshot_date"] = ""
        summary["uploaded_at"] = ""
        summary["uploaded_by"] = ""
        summary["history_dates"] = []
        summary["visible_history_dates"] = []
        summary["older_history_dates"] = []
        summary["compare_dates"] = []
        summary["comparison"] = {
            "previous_snapshot_date": "",
            "previous_change": None,
            "compare_date": "",
            "compare_change": None,
            "last_month_source_date": "",
            "last_month_source_month": previous_month_key(month_key),
            "last_month_change": None,
        }
        return summary

    summary = build_class_completion_summary(month_key)
    summary.update(
        {
            "source": "classes",
            "snapshot_date": "",
            "uploaded_at": "",
            "uploaded_by": "",
            "can_upload": can_upload_completion_data(),
            "history_dates": [],
            "visible_history_dates": [],
            "older_history_dates": [],
            "compare_dates": [],
            "comparison": {
                "previous_snapshot_date": "",
                "previous_change": None,
                "compare_date": "",
                "compare_change": None,
                "last_month_source_date": "",
                "last_month_source_month": previous_month_key(month_key),
                "last_month_change": None,
            },
        }
    )
    return summary


def report_rows_by_teacher(report):
    return {
        str(row.get("teacher_id") or row.get("username") or "").strip().lower(): row
        for row in report.get("rows", [])
        if str(row.get("teacher_id") or row.get("username") or "").strip()
    }


def build_metric_summary(month_key, report_date, completion_teachers, metric_key):
    field = DATABASE_METRICS[metric_key]["field"]
    store = load_json(daily_report_file(), {"reports": {}})
    rows = {
        teacher["teacher_id"]: {
            "teacher_id": teacher["teacher_id"],
            "teacher_name": teacher["teacher_name"],
            "student_count": teacher["student_count"],
            "today": 0,
            "month_total": 0,
        }
        for teacher in completion_teachers
    }

    for date_key, report in store.get("reports", {}).items():
        date_text = str(date_key)
        if not date_text.startswith(f"{month_key}-") or date_text > report_date:
            continue
        saved_rows = report_rows_by_teacher(report)
        for teacher_id, output in rows.items():
            source = saved_rows.get(teacher_id)
            if not source:
                continue
            value = parse_count(field_value(source, field))
            output["month_total"] += value
            if date_text == report_date:
                output["today"] += value

    row_list = list(rows.values())
    return {
        "field": field,
        "label": DATABASE_METRICS[metric_key]["label"],
        "today_total": sum(row["today"] for row in row_list),
        "month_total": sum(row["month_total"] for row in row_list),
        "rows": row_list,
    }


def build_renewal_summary(month_key, report_date, completion_teachers):
    field = DATABASE_METRICS["renewal"]["field"]
    store = load_json(daily_report_file(), {"reports": {}})
    rows = {
        teacher["teacher_id"]: {
            "teacher_id": teacher["teacher_id"],
            "teacher_name": teacher["teacher_name"],
            "student_count": teacher["student_count"],
            "today": 0,
            "week_totals": [0, 0, 0, 0],
            "month_total": 0,
        }
        for teacher in completion_teachers
    }

    for date_key, report in store.get("reports", {}).items():
        date_text = str(date_key)
        if not date_text.startswith(f"{month_key}-") or date_text > report_date:
            continue
        week_index = month_week_index(date_text)
        saved_rows = report_rows_by_teacher(report)
        for teacher_id, output in rows.items():
            source = saved_rows.get(teacher_id)
            if not source:
                continue
            value = parse_count(field_value(source, field))
            output["week_totals"][week_index] += value
            output["month_total"] += value
            if date_text == report_date:
                output["today"] += value

    row_list = list(rows.values())
    week_totals = [
        sum(row["week_totals"][index] for row in row_list)
        for index in range(4)
    ]
    return {
        "field": field,
        "label": DATABASE_METRICS["renewal"]["label"],
        "today_total": sum(row["today"] for row in row_list),
        "month_total": sum(row["month_total"] for row in row_list),
        "week_totals": week_totals,
        "rows": row_list,
    }


def build_learning_summary(month_key, report_date, completion):
    class_completion = build_class_completion_summary(month_key)
    summary = build_metric_summary(month_key, report_date, class_completion["teachers"], "learning")
    summary["rows"] = [
        row
        for row in summary["rows"]
        if row["teacher_id"] not in LEARNING_EXCLUDED_TEACHER_IDS
    ]
    summary["today_total"] = sum(row["today"] for row in summary["rows"])
    summary["month_total"] = sum(row["month_total"] for row in summary["rows"])
    settings = load_database_settings().get("learning", {})
    class_settings = settings.get("classes", {})
    teacher_settings = settings.get("teachers", {})
    teacher_lookup = {row["teacher_id"]: row for row in summary["rows"]}
    for row in summary["rows"]:
        row["student_count"] = 0
        row["learning_base"] = 0
        row["classes"] = []
    class_rows = []

    for item in class_completion.get("classes", []):
        class_id = item.get("id", "")
        teacher_id = item.get("teacher_id", "")
        if teacher_id in LEARNING_EXCLUDED_TEACHER_IDS:
            continue
        coefficient = parse_float(class_settings.get(class_id, {}).get("coefficient"), 0)
        student_count = int(item.get("student_count") or 0)
        learning_base = rounded_metric(student_count * coefficient)
        class_row = {
            "class_id": class_id,
            "class_name": item.get("name", ""),
            "teacher_id": teacher_id,
            "teacher_name": item.get("teacher_name", ""),
            "student_count": student_count,
            "coefficient": coefficient,
            "learning_base": learning_base,
            "can_edit": setting_allowed_for_teacher(teacher_id),
        }
        class_rows.append(class_row)

        teacher_row = teacher_lookup.get(teacher_id)
        if teacher_row:
            teacher_row["student_count"] += student_count
            teacher_row["learning_base"] = rounded_metric(teacher_row.get("learning_base", 0) + learning_base)
            teacher_row["classes"].append(class_row)

    for row in summary["rows"]:
        teacher_id = row["teacher_id"]
        target_rate = normalize_target_rate(teacher_settings.get(teacher_id, {}).get("target_rate"))
        learning_base = rounded_metric(row.get("learning_base", 0))
        row["learning_base"] = learning_base
        row["target_rate"] = target_rate
        row["target_learning"] = rounded_metric(learning_base * target_rate)
        row["target_gap"] = rounded_metric(row.get("month_total", 0) - row["target_learning"])
        row["can_edit"] = setting_allowed_for_teacher(teacher_id)
        row.setdefault("classes", [])

    summary["learning_base_total"] = rounded_metric(sum(row["learning_base"] for row in summary["rows"]))
    summary["target_learning_total"] = rounded_metric(sum(row["target_learning"] for row in summary["rows"]))
    summary["target_gap_total"] = rounded_metric(sum(row["target_gap"] for row in summary["rows"]))
    summary["achievement_rate"] = (
        rounded_metric(summary["month_total"] / summary["learning_base_total"] * 100)
        if summary["learning_base_total"]
        else None
    )
    summary["target_rates"] = sorted(LEARNING_TARGET_RATES)
    summary["classes"] = class_rows
    return summary


def build_referral_summary(month_key, report_date, completion_teachers):
    store = load_json(daily_report_file(), {"reports": {}})
    rows = {
        teacher["teacher_id"]: {
            "teacher_id": teacher["teacher_id"],
            "teacher_name": teacher["teacher_name"],
            "student_count": teacher["student_count"],
            "leads_today": 0,
            "leads_month_total": 0,
            "conversions_today": 0,
            "conversions_month_total": 0,
        }
        for teacher in completion_teachers
    }

    for date_key, report in store.get("reports", {}).items():
        date_text = str(date_key)
        if not date_text.startswith(f"{month_key}-") or date_text > report_date:
            continue
        saved_rows = report_rows_by_teacher(report)
        for teacher_id, output in rows.items():
            source = saved_rows.get(teacher_id)
            if not source:
                continue
            leads = parse_count(field_value(source, REFERRAL_FIELDS["leads"]))
            conversions = parse_count(field_value(source, REFERRAL_FIELDS["conversions"]))
            output["leads_month_total"] += leads
            output["conversions_month_total"] += conversions
            if date_text == report_date:
                output["leads_today"] += leads
                output["conversions_today"] += conversions

    row_list = list(rows.values())
    return {
        "label": "转介绍",
        "leads_today_total": sum(row["leads_today"] for row in row_list),
        "leads_month_total": sum(row["leads_month_total"] for row in row_list),
        "conversions_today_total": sum(row["conversions_today"] for row in row_list),
        "conversions_month_total": sum(row["conversions_month_total"] for row in row_list),
        "rows": row_list,
    }


def gmv_month_settings(month_key):
    settings = load_database_settings()
    return (
        settings.get("gmv", {})
        .get(month_key, {})
    )


def gmv_month_adjustments(month_key, section_key):
    return gmv_month_settings(month_key).get(section_key, {})


def gmv_month_targets(month_key):
    targets = gmv_month_settings(month_key).get("targets", {})
    return targets if isinstance(targets, dict) else {}


def normalize_gmv_week_overrides(raw_values):
    if not isinstance(raw_values, list):
        raw_values = []
    output = []
    for index in range(4):
        value = raw_values[index] if index < len(raw_values) else None
        output.append(parse_gmv_amount(value))
    return output


def build_gmv_section(month_key, report_date, completion_teachers, section_key):
    config = GMV_SECTIONS[section_key]
    store = load_json(daily_report_file(), {"reports": {}})
    rows = {
        teacher["teacher_id"]: {
            "teacher_id": teacher["teacher_id"],
            "teacher_name": teacher["teacher_name"],
            "student_count": teacher["student_count"],
            "unit_count_week_totals": [0, 0, 0, 0],
            "default_week_totals": [0, 0, 0, 0],
        }
        for teacher in completion_teachers
    }

    for date_key, report in store.get("reports", {}).items():
        date_text = str(date_key)
        if not date_text.startswith(f"{month_key}-") or date_text > report_date:
            continue
        week_index = month_week_index(date_text)
        saved_rows = report_rows_by_teacher(report)
        for teacher_id, output in rows.items():
            source = saved_rows.get(teacher_id)
            if not source:
                continue
            unit_count = parse_count(field_value(source, config["field"]))
            output["unit_count_week_totals"][week_index] += unit_count
            output["default_week_totals"][week_index] = rounded_metric(
                output["default_week_totals"][week_index] + unit_count * GMV_UNIT_PRICE
            )

    adjustments = gmv_month_adjustments(month_key, section_key)
    row_list = []
    for row in rows.values():
        teacher_adjustment = adjustments.get(row["teacher_id"], {})
        overrides = normalize_gmv_week_overrides(teacher_adjustment.get("week_totals"))
        week_totals = []
        manual_week_flags = []
        for index, default_value in enumerate(row["default_week_totals"]):
            override_value = overrides[index]
            manual_week_flags.append(override_value is not None)
            week_totals.append(override_value if override_value is not None else default_value)
        row["week_totals"] = week_totals
        row["manual_week_flags"] = manual_week_flags
        row["month_total"] = rounded_metric(sum(week_totals))
        row["default_month_total"] = rounded_metric(sum(row["default_week_totals"]))
        row["is_adjusted"] = any(manual_week_flags)
        row_list.append(row)

    week_totals = [
        rounded_metric(sum(row["week_totals"][index] for row in row_list))
        for index in range(4)
    ]
    default_week_totals = [
        rounded_metric(sum(row["default_week_totals"][index] for row in row_list))
        for index in range(4)
    ]
    month_total = rounded_metric(sum(row["month_total"] for row in row_list))
    default_month_total = rounded_metric(sum(row["default_month_total"] for row in row_list))
    target_amount = parse_gmv_amount(gmv_month_targets(month_key).get(section_key))
    target_gap = rounded_metric(month_total - target_amount) if target_amount is not None else None

    return {
        "key": section_key,
        "label": config["label"],
        "field": config["field"],
        "unit_price": GMV_UNIT_PRICE,
        "can_edit": can_manage_gmv(),
        "week_totals": week_totals,
        "default_week_totals": default_week_totals,
        "month_total": month_total,
        "default_month_total": default_month_total,
        "target_amount": target_amount,
        "target_gap": target_gap,
        "rows": row_list,
    }


def build_gmv_summary(month_key, report_date, completion_teachers):
    renewal = build_gmv_section(month_key, report_date, completion_teachers, "renewal")
    referral = build_gmv_section(month_key, report_date, completion_teachers, "referral")
    target_values = [
        section["target_amount"]
        for section in (renewal, referral)
        if section.get("target_amount") is not None
    ]
    target_amount = rounded_metric(sum(target_values)) if target_values else None
    month_total = rounded_metric(renewal["month_total"] + referral["month_total"])
    return {
        "unit_price": GMV_UNIT_PRICE,
        "can_edit": can_manage_gmv(),
        "renewal": renewal,
        "referral": referral,
        "month_total": month_total,
        "default_month_total": rounded_metric(renewal["default_month_total"] + referral["default_month_total"]),
        "target_amount": target_amount,
        "target_gap": rounded_metric(month_total - target_amount) if target_amount is not None else None,
    }


def monthly_daily_reports(month_key, report_date):
    store = load_json(daily_report_file(), {"reports": {}})
    reports = {}
    for date_key, report in store.get("reports", {}).items():
        date_text = str(date_key)
        if date_text.startswith(f"{month_key}-") and date_text <= report_date:
            reports[date_text] = report
    return reports


def monthly_archive_settings(month_key):
    settings = load_database_settings()
    return {
        "learning": json.loads(json.dumps(settings.get("learning", {}), ensure_ascii=False)),
        "gmv": json.loads(json.dumps(settings.get("gmv", {}).get(month_key, {}), ensure_ascii=False)),
    }


def build_monthly_archive(month_key, report_date):
    completion = build_completion_summary(month_key, report_date)
    completion_teachers = completion["teachers"]
    archive = {
        "month": month_key,
        "date": report_date,
        "archived_at": datetime.now().isoformat(timespec="seconds"),
        "archived_by": g.user.get("username", ""),
        "next_month": next_month_key(month_key),
        "next_date": next_month_start_date(month_key),
        "daily_reports": monthly_daily_reports(month_key, report_date),
        "database": {
            "completion": completion,
            "learning": build_learning_summary(month_key, report_date, completion),
            "renewal": build_renewal_summary(month_key, report_date, completion_teachers),
            "referral": build_referral_summary(month_key, report_date, completion_teachers),
            "gmv": build_gmv_summary(month_key, report_date, completion_teachers),
            "completion_performance": build_completion_performance_summary(completion),
        },
        "settings": monthly_archive_settings(month_key),
    }
    archive["daily_report_count"] = len(archive["daily_reports"])
    archive["summary"] = {
        "learning_achievement_rate": archive["database"]["learning"].get("achievement_rate"),
        "learning_month_total": archive["database"]["learning"].get("month_total"),
        "renewal_orders": archive["database"]["renewal"].get("month_total"),
        "referral_leads": archive["database"]["referral"].get("leads_month_total"),
        "referral_conversions": archive["database"]["referral"].get("conversions_month_total"),
        "gmv_total": archive["database"]["gmv"].get("month_total"),
        "completion_performance_reward": archive["database"]["completion_performance"].get("summary", {}).get("reward_total"),
    }
    return archive


def reminder_cycle_key_for_date(date_text):
    try:
        day = datetime.strptime(str(date_text or ""), "%Y-%m-%d").date()
    except ValueError:
        day = datetime.now().date()
    monday = day - timedelta(days=day.weekday())
    return monday.isoformat()


def current_reminder_cycle_key():
    return reminder_cycle_key_for_date(datetime.now().strftime("%Y-%m-%d"))


def reminder_day_label(day_key):
    item = REMINDER_FLOW_BY_KEY.get(str(day_key or ""))
    return item["label"] if item else str(day_key or "")


def reminder_day_key_for_date(date_text):
    try:
        day = datetime.strptime(str(date_text or ""), "%Y-%m-%d").date()
    except ValueError:
        day = datetime.now().date()
    return {
        0: "monday",
        1: "tuesday",
        2: "wednesday",
        3: "thursday",
        4: "friday",
    }.get(day.weekday(), "")


def reminder_recovery_target(origin_day_key):
    target_key = REMINDER_RECOVERY_TARGET_BY_ORIGIN.get(str(origin_day_key or ""))
    if not target_key:
        return None
    return {"key": target_key, "label": reminder_day_label(target_key)}


def reminder_target_teacher_id(payload):
    return (
        normalize_teacher_id(payload.get("teacher_id"))
        or normalize_teacher_id(payload.get("class_teacher_id"))
        or current_user_teacher_id()
    )


def can_access_reminder_teacher(teacher_id):
    current_teacher = current_user_teacher_id()
    return current_teacher == COMPLETION_UPLOAD_TEACHER_ID or current_teacher == normalize_teacher_id(teacher_id)


def reminder_student_snapshot(student):
    weeks = student.get("weeks") if isinstance(student.get("weeks"), dict) else {}
    incomplete_days = student.get("incomplete_days") if isinstance(student.get("incomplete_days"), list) else []
    uploaded_days = student.get("uploaded_days") if isinstance(student.get("uploaded_days"), list) else []
    phone_call = student.get("phone_call") if isinstance(student.get("phone_call"), dict) else {}
    normalized_weeks = {}
    for week in range(1, 5):
        values = weeks.get(str(week)) if isinstance(weeks.get(str(week)), list) else []
        normalized_weeks[str(week)] = [
            parse_completion(values[index]) if index < len(values) else None
            for index in range(6)
        ]
    return {
        "id": str(student.get("id") or ""),
        "name": str(student.get("name") or "").strip(),
        "account": str(student.get("account") or "").strip(),
        "category": str(student.get("category") or student.get("habit_category") or "暂无数据").strip(),
        "monthly_completion": parse_completion(student.get("monthly_completion")),
        "renewal_enrolled": bool(student.get("renewal_enrolled")),
        "prompt": str(student.get("prompt") or "").strip(),
        "phone_call": {
            "completed": bool(phone_call.get("completed")),
            "completed_at": str(phone_call.get("completed_at") or "").strip(),
        },
        "weeks": normalized_weeks,
        "incomplete_days": [
            {
                "week": str(item.get("week") or ""),
                "day": parse_student_count(item.get("day")),
                "label": str(item.get("label") or "").strip(),
                "value": parse_completion(item.get("value")),
            }
            for item in incomplete_days
            if isinstance(item, dict)
        ],
        "uploaded_days": [
            {
                "week": str(item.get("week") or ""),
                "day": parse_student_count(item.get("day")),
                "label": str(item.get("label") or "").strip(),
                "value": parse_completion(item.get("value")),
            }
            for item in uploaded_days
            if isinstance(item, dict)
        ],
    }


def reminder_payload_class_keys(payload):
    keys = set()
    for field in ("class_name", "local_class_name"):
        keys.update(reminder_class_match_keys(payload.get(field, "")))
    return sorted(keys)


def reminder_record_matches(record, class_name, day_key, recover_from=""):
    if record.get("status") != "pending_recovery":
        return False
    if record.get("cycle_key") != current_reminder_cycle_key():
        return False
    if str(record.get("recovery_day_key") or "") != str(day_key or ""):
        return False
    if recover_from and str(record.get("origin_day_key") or "") != str(recover_from):
        return False
    class_keys = reminder_class_match_keys(class_name)
    record_keys = set(record.get("class_keys") or [])
    if not record_keys:
        record_keys = reminder_class_match_keys(record.get("class_name", ""))
    return bool(class_keys.intersection(record_keys))


def reminder_schedule_action_records(visible_teacher_ids):
    store = load_reminder_actions()
    cycle_key = current_reminder_cycle_key()
    records = []
    for record in store.get("records", []):
        teacher_id = normalize_teacher_id(record.get("teacher_id"))
        if visible_teacher_ids and teacher_id not in visible_teacher_ids:
            continue
        if record.get("cycle_key") != cycle_key:
            continue
        if record.get("status") not in {"pending_recovery", "completed", "recovered"}:
            continue
        records.append(record)
    return records


def reminder_record_class_intersects(record, row):
    row_keys = reminder_class_match_keys(row.get("class_name") or row.get("name", ""))
    record_keys = set(record.get("class_keys") or [])
    if not record_keys:
        record_keys = reminder_class_match_keys(record.get("class_name", ""))
    return bool(row_keys.intersection(record_keys))


def reminder_record_matches_schedule(record, row, day_key, task_label, recover_from=""):
    row_teacher_id = normalize_teacher_id(row.get("teacher_id"))
    record_teacher_id = normalize_teacher_id(record.get("teacher_id"))
    if row_teacher_id and record_teacher_id != row_teacher_id:
        return False
    if not reminder_record_class_intersects(record, row):
        return False

    if task_label == "回收":
        return (
            record.get("status") == "recovered"
            and str(record.get("recovery_day_key") or "") == str(day_key or "")
            and (not recover_from or str(record.get("origin_day_key") or "") == str(recover_from))
        )

    return (
        str(record.get("task_label") or "") == str(task_label or "")
        and str(record.get("origin_day_key") or "") == str(day_key or "")
        and record.get("status") in {"pending_recovery", "completed", "recovered"}
    )


def reminder_schedule_action_state(row, day_key, task_label, action_records, recover_from=""):
    matches = [
        record
        for record in action_records
        if reminder_record_matches_schedule(record, row, day_key, task_label, recover_from)
    ]
    if not matches:
        return {
            "completed": False,
            "completed_at": "",
            "record_count": 0,
            "student_count": 0,
            "label": "",
        }
    latest = max(
        matches,
        key=lambda record: record.get("recovered_at") or record.get("completed_at") or record.get("created_at") or "",
    )
    completed_at = latest.get("recovered_at") or latest.get("completed_at") or latest.get("created_at") or ""
    return {
        "completed": True,
        "completed_at": completed_at,
        "record_count": len(matches),
        "student_count": sum(len(record.get("students", [])) for record in matches),
        "label": "已完成回收" if task_label == "回收" else "已完成催课",
    }


def reminder_visible_teacher_ids():
    current_teacher = current_user_teacher_id()
    if current_teacher == COMPLETION_UPLOAD_TEACHER_ID:
        return {teacher["id"] for teacher in TEACHERS}
    return {current_teacher} if current_teacher else set()


def reminder_teacher_seed(teacher_id):
    teacher_name = teacher_label(teacher_id)
    return {
        "teacher_id": teacher_id,
        "teacher_name": teacher_name or "未分配",
        "class_count": 0,
        "database_count": 0,
        "included_count": 0,
        "ignored_count": 0,
        "extra_count": 0,
        "schedule_count": 0,
        "priorities": [],
        "database_fallbacks": [],
        "extra_classes": [],
        "schedule": [],
    }


def reminder_class_assessment_meta(item):
    week_number = current_title_week_number(item)
    completion_assessed = reminder_week_is_assessed(week_number)
    return {
        "title_week_number": week_number,
        "title_week_label": f"W{week_number}" if week_number else "",
        "completion_assessed": completion_assessed,
    }


def reminder_home_class_meta_by_teacher(home_classes):
    lookup = {}
    for item in home_classes:
        teacher_id = class_teacher_id(item)
        if not teacher_id:
            continue
        class_keys = reminder_local_class_keys(item)
        meta = reminder_class_assessment_meta(item)
        meta.update({
            "class_id": item.get("id", ""),
            "class_name": item.get("name", ""),
            "class_keys": sorted(class_keys),
        })
        teacher_lookup = lookup.setdefault(teacher_id, {})
        for key in class_keys:
            teacher_lookup.setdefault(key, meta)
    return lookup


def reminder_local_class_meta(row, meta_by_teacher):
    teacher_id = normalize_teacher_id(row.get("teacher_id"))
    keys = reminder_class_match_keys(row.get("name") or row.get("class_name", ""))
    for key in keys:
        meta = meta_by_teacher.get(teacher_id, {}).get(key)
        if meta:
            return meta
    return {}


def reminder_completion_row_score(row):
    return (
        bool(row.get("lookup_matched")),
        row.get("completion_rate") is not None,
        row.get("last_month_completion") is not None,
        row.get("change_from_last_month") is not None,
        parse_student_count(row.get("student_count")),
    )


def reminder_completion_rows_by_teacher(completion):
    lookup = {}
    for row in completion.get("classes", []):
        teacher_id = normalize_teacher_id(row.get("teacher_id"))
        if not teacher_id:
            continue
        teacher_lookup = lookup.setdefault(teacher_id, {})
        for key in reminder_class_match_keys(row.get("name") or row.get("class_name", "")):
            current = teacher_lookup.get(key)
            if current is None or reminder_completion_row_score(row) > reminder_completion_row_score(current):
                teacher_lookup[key] = row
    return lookup


def reminder_completion_row_for_home_class(item, completion_rows_by_teacher):
    teacher_id = class_teacher_id(item)
    teacher_lookup = completion_rows_by_teacher.get(teacher_id, {})
    for key in reminder_local_class_keys(item):
        row = teacher_lookup.get(key)
        if row:
            return row
    return {}


def reminder_home_class_plan_row(item, completion_row=None):
    row = reminder_home_class_row(item)
    row["id"] = item.get("id", "")
    row["name"] = item.get("name", "")
    row["source"] = "my_class"
    row["source_label"] = "考核班级" if row.get("completion_assessed", True) else "续费期"
    if completion_row:
        row["completion_data_matched"] = True
        row["database_class_id"] = completion_row.get("id", "")
        row["database_class_name"] = completion_row.get("name", "")
        row["student_count"] = parse_student_count(completion_row.get("student_count")) or row.get("student_count")
        row["completion_rate"] = completion_row.get("completion_rate")
        row["last_month_completion"] = completion_row.get("last_month_completion")
        row["change_from_last_month"] = completion_row.get("change_from_last_month")
        row["lookup_matched"] = bool(completion_row.get("lookup_matched"))
    else:
        row["completion_data_matched"] = False
        row["database_class_id"] = ""
        row["database_class_name"] = ""
        row["lookup_matched"] = False
    return row


def reminder_gap_from_last_month(row):
    change = row.get("change_from_last_month")
    if change is None:
        return None
    return rounded_metric(-float(change))


def reminder_row_requires_high_frequency(row):
    gap = reminder_gap_from_last_month(row)
    return bool(
        row.get("completion_assessed", True)
        and gap is not None
        and gap > REMINDER_HIGH_GAP_THRESHOLD
    )


def reminder_priority_row(row, index):
    change = row.get("change_from_last_month")
    gap = reminder_gap_from_last_month(row)
    high_frequency = reminder_row_requires_high_frequency(row)
    source_label = row.get("source_label") or ("高频考核" if high_frequency else "考核班级")
    if high_frequency:
        source_label = "高频考核"
    return {
        "rank": index + 1,
        "stars": max(1, 5 - index),
        "class_id": row.get("id", ""),
        "class_name": row.get("name", ""),
        "local_class_id": row.get("local_class_id", ""),
        "local_class_name": row.get("local_class_name", ""),
        "teacher_id": row.get("teacher_id", ""),
        "teacher_name": row.get("teacher_name", ""),
        "student_count": parse_student_count(row.get("student_count")),
        "completion_rate": row.get("completion_rate"),
        "last_month_completion": row.get("last_month_completion"),
        "change_from_last_month": change,
        "gap_from_last_month": gap,
        "completion_activity": bool(row.get("completion_activity")),
        "completion_assessed": bool(row.get("completion_assessed", True)),
        "title_week_number": row.get("title_week_number"),
        "title_week_label": row.get("title_week_label", ""),
        "requires_high_frequency": high_frequency,
        "source": row.get("source", "my_class"),
        "source_label": source_label,
    }


def reminder_database_fallback_row(row):
    source_label = row.get("source_label")
    if not source_label:
        source_label = "考核班级" if row.get("completion_assessed", True) else "续费期"
    return {
        "rank": None,
        "stars": None,
        "class_id": row.get("id", ""),
        "class_name": row.get("name", ""),
        "local_class_id": row.get("local_class_id", ""),
        "local_class_name": row.get("local_class_name", ""),
        "teacher_id": row.get("teacher_id", ""),
        "teacher_name": row.get("teacher_name", ""),
        "student_count": parse_student_count(row.get("student_count")),
        "completion_rate": row.get("completion_rate"),
        "last_month_completion": row.get("last_month_completion"),
        "change_from_last_month": row.get("change_from_last_month"),
        "gap_from_last_month": reminder_gap_from_last_month(row),
        "completion_activity": bool(row.get("completion_activity")),
        "completion_assessed": bool(row.get("completion_assessed", True)),
        "title_week_number": row.get("title_week_number"),
        "title_week_label": row.get("title_week_label", ""),
        "requires_high_frequency": False,
        "source": row.get("source", "my_class"),
        "source_label": source_label,
    }


def reminder_home_class_row(item):
    teacher_id = class_teacher_id(item)
    assessment_meta = reminder_class_assessment_meta(item)
    return {
        "rank": None,
        "stars": None,
        "id": item.get("id", ""),
        "name": item.get("name", ""),
        "class_id": item.get("id", ""),
        "class_name": item.get("name", ""),
        "local_class_id": item.get("id", ""),
        "local_class_name": item.get("name", ""),
        "teacher_id": teacher_id,
        "teacher_name": teacher_label(teacher_id) or "未分配",
        "student_count": len(item.get("students", [])),
        "completion_rate": None,
        "last_month_completion": None,
        "change_from_last_month": None,
        "gap_from_last_month": None,
        "completion_activity": bool(item.get("completion_activity")),
        "completion_assessed": assessment_meta["completion_assessed"],
        "title_week_number": assessment_meta["title_week_number"],
        "title_week_label": assessment_meta["title_week_label"],
        "requires_high_frequency": False,
        "source": "my_class",
        "source_label": "考核班级" if assessment_meta["completion_assessed"] else "续费期",
    }


def reminder_class_ref(
    row,
    day_key="",
    task_label="",
    action_records=None,
    recover_from="",
    local_upload_lookup=None,
    report_date="",
):
    output = {
        "class_id": row.get("class_id", ""),
        "class_name": row.get("class_name", ""),
        "local_class_id": row.get("local_class_id", ""),
        "local_class_name": row.get("local_class_name", ""),
        "teacher_id": row.get("teacher_id", ""),
        "teacher_name": row.get("teacher_name", ""),
        "rank": row.get("rank"),
        "stars": row.get("stars"),
        "student_count": row.get("student_count"),
        "completion_rate": row.get("completion_rate"),
        "last_month_completion": row.get("last_month_completion"),
        "change_from_last_month": row.get("change_from_last_month"),
        "completion_activity": bool(row.get("completion_activity")),
        "completion_assessed": bool(row.get("completion_assessed", True)),
        "title_week_number": row.get("title_week_number"),
        "title_week_label": row.get("title_week_label", ""),
        "requires_high_frequency": bool(row.get("requires_high_frequency")),
        "source": row.get("source", "database"),
        "source_label": row.get("source_label", "数据库"),
    }
    if action_records is not None:
        output["action_state"] = reminder_schedule_action_state(
            row,
            day_key,
            task_label,
            action_records,
            recover_from,
        )
    if local_upload_lookup is not None:
        output["upload_state"] = reminder_local_class_upload_info(output, local_upload_lookup, report_date)
    return output


def reminder_plan_local_meta_index():
    home_classes = reminder_home_classes(set())
    meta_by_teacher = reminder_home_class_meta_by_teacher(home_classes)
    meta_by_id = {}
    for teacher_lookup in meta_by_teacher.values():
        for meta in teacher_lookup.values():
            class_id = str(meta.get("class_id") or "").strip()
            if class_id:
                meta_by_id[class_id] = meta
    return meta_by_teacher, meta_by_id


def reminder_row_candidate_keys(row):
    keys = set(row.get("class_keys") or [])
    for value in (row.get("class_name", ""), row.get("local_class_name", ""), row.get("name", "")):
        keys.update(reminder_class_match_keys(value))
    return keys


def reminder_row_local_meta(row, meta_by_teacher, meta_by_id):
    for field in ("local_class_id", "class_id", "id"):
        class_id = str(row.get(field) or "").strip()
        if class_id and class_id in meta_by_id:
            return meta_by_id[class_id]

    teacher_id = normalize_teacher_id(row.get("teacher_id"))
    teacher_lookup = meta_by_teacher.get(teacher_id, {})
    for key in reminder_row_candidate_keys(row):
        meta = teacher_lookup.get(key)
        if meta:
            return meta
    return {}


def apply_local_meta_to_reminder_row(row, meta):
    changed = False

    def set_value(key, value):
        nonlocal changed
        if row.get(key) != value:
            row[key] = value
            changed = True

    if meta:
        set_value("local_class_id", meta.get("class_id", ""))
        set_value("local_class_name", meta.get("class_name", ""))
        set_value("class_name", meta.get("class_name", ""))
        set_value("title_week_number", meta.get("title_week_number"))
        set_value("title_week_label", meta.get("title_week_label", ""))
        set_value("completion_assessed", bool(meta.get("completion_assessed", True)))
        set_value("class_keys", sorted(meta.get("class_keys") or []))

    week_number = row.get("title_week_number")
    if week_number is not None:
        assessed = reminder_week_is_assessed(week_number)
        set_value("completion_assessed", assessed)
        if assessed:
            set_value("title_week_label", f"W{week_number}")

    return changed


def normalize_reminder_priority_rows(rows):
    normalized = []
    for index, row in enumerate(rows):
        next_row = {**row}
        next_row["rank"] = index + 1
        next_row["stars"] = max(1, 5 - index)
        next_row["requires_high_frequency"] = reminder_row_requires_high_frequency(next_row)
        normalized.append(next_row)
    return normalized


def reminder_rows_signature_set(rows):
    return {reminder_schedule_row_signature(row) for row in rows}


def append_unique_reminder_rows(target, rows):
    seen = reminder_rows_signature_set(target)
    for row in rows:
        signature = reminder_schedule_row_signature(row)
        if signature in seen:
            continue
        target.append(row)
        seen.add(signature)


def reminder_plan_source_context(plan):
    source_date = (
        str(plan.get("plan_source_date") or "").strip()
        or str(plan.get("date") or "").strip()
        or datetime.now().strftime("%Y-%m-%d")
    )
    month_key = str(plan.get("month") or "").strip() or source_date[:7] or datetime.now().strftime("%Y-%m")
    cycle_key = str(plan.get("cycle_key") or "").strip() or reminder_cycle_key_for_date(source_date)
    return month_key, source_date, cycle_key


def reminder_group_row_signatures(group):
    rows = []
    for key in ("priorities", "database_fallbacks", "extra_classes"):
        rows.extend([row for row in group.get(key, []) if isinstance(row, dict)])
    return reminder_rows_signature_set(rows)


def reminder_group_count_rows(group):
    rows = []
    seen = set()
    for key in ("priorities", "database_fallbacks", "extra_classes"):
        for row in group.get(key, []):
            if not isinstance(row, dict):
                continue
            signature = reminder_schedule_row_signature(row)
            if signature in seen:
                continue
            rows.append(row)
            seen.add(signature)
    return rows


def reminder_row_has_completion_data(row):
    return bool(
        row.get("completion_data_matched")
        or row.get("database_class_id")
        or row.get("lookup_matched")
        or row.get("completion_rate") is not None
        or row.get("last_month_completion") is not None
    )


def append_missing_current_home_classes(plan):
    month_key, source_date, cycle_key = reminder_plan_source_context(plan)
    source_plan = build_completion_reminder_plan_source(
        month_key,
        source_date,
        reminder_all_teacher_ids(),
        action_records=None,
        cycle_key=cycle_key,
    )
    target_groups = {
        normalize_teacher_id(group.get("teacher_id")): group
        for group in plan.get("groups", [])
        if isinstance(group, dict)
    }
    changed = False
    for source_group in source_plan.get("groups", []):
        teacher_id = normalize_teacher_id(source_group.get("teacher_id"))
        if not teacher_id:
            continue
        target_group = target_groups.get(teacher_id)
        if target_group is None:
            plan.setdefault("groups", []).append(source_group)
            target_groups[teacher_id] = source_group
            changed = True
            continue

        existing_signatures = reminder_group_row_signatures(target_group)
        for key in ("priorities", "database_fallbacks", "extra_classes"):
            target_rows = target_group.setdefault(key, [])
            for source_row in source_group.get(key, []):
                if not isinstance(source_row, dict):
                    continue
                signature = reminder_schedule_row_signature(source_row)
                if signature in existing_signatures:
                    continue
                target_rows.append({**source_row})
                existing_signatures.add(signature)
                changed = True
    return changed


def reconcile_weekly_reminder_plan(plan):
    if not isinstance(plan, dict) or not isinstance(plan.get("groups"), list):
        return False

    changed = append_missing_current_home_classes(plan)
    meta_by_teacher, meta_by_id = reminder_plan_local_meta_index()

    for group in plan.get("groups", []):
        if not isinstance(group, dict):
            continue

        priority_rows = [row for row in group.get("priorities", []) if isinstance(row, dict)]
        fallback_rows = [row for row in group.get("database_fallbacks", []) if isinstance(row, dict)]
        extra_rows = [row for row in group.get("extra_classes", []) if isinstance(row, dict)]
        for row in priority_rows + fallback_rows + extra_rows:
            meta = reminder_row_local_meta(row, meta_by_teacher, meta_by_id)
            if apply_local_meta_to_reminder_row(row, meta):
                changed = True

        assessed_priorities = []
        moved_database_rows = []
        moved_extra_rows = []
        for row in priority_rows:
            if row.get("completion_assessed", True):
                assessed_priorities.append(row)
                continue
            moved_row = {**row, "rank": None, "stars": None, "requires_high_frequency": False}
            if str(row.get("source") or "") == "my_class":
                moved_extra_rows.append(moved_row)
            else:
                moved_database_rows.append(moved_row)
            changed = True

        normalized_priorities = normalize_reminder_priority_rows(assessed_priorities)
        if normalized_priorities != priority_rows:
            changed = True
        priority_rows = normalized_priorities

        append_unique_reminder_rows(fallback_rows, moved_database_rows)
        append_unique_reminder_rows(extra_rows, moved_extra_rows)

        assessed_database_fallback_rows = [row for row in fallback_rows if row.get("completion_assessed", True)]
        renewal_database_fallback_rows = [row for row in fallback_rows if not row.get("completion_assessed", True)]
        assessed_extra_rows = [row for row in extra_rows if row.get("completion_assessed", True)]
        renewal_extra_rows = [row for row in extra_rows if not row.get("completion_assessed", True)]
        schedule_rows = (
            priority_rows
            + assessed_database_fallback_rows
            + assessed_extra_rows
            + renewal_database_fallback_rows
            + renewal_extra_rows
        )
        counted_rows = reminder_group_count_rows({
            "priorities": priority_rows,
            "database_fallbacks": fallback_rows,
            "extra_classes": extra_rows,
        })

        rebuilt_schedule = build_reminder_schedule(schedule_rows)
        next_values = {
            "priorities": priority_rows,
            "database_fallbacks": fallback_rows,
            "extra_classes": extra_rows,
            "class_count": len(counted_rows),
            "database_count": sum(1 for row in counted_rows if reminder_row_has_completion_data(row)),
            "included_count": len(priority_rows),
            "ignored_count": len(fallback_rows),
            "extra_count": len(extra_rows),
            "schedule_count": len(schedule_rows),
            "schedule": rebuilt_schedule,
        }
        for key, value in next_values.items():
            if group.get(key) != value:
                group[key] = value
                changed = True

    if changed:
        plan["updated_at"] = datetime.now().isoformat(timespec="seconds")
        plan["reconciled_at"] = plan["updated_at"]
    return changed


def build_reminder_schedule(schedule_rows, action_records=None):
    include_action_state = action_records is not None
    action_records = action_records or []
    buckets = {}
    output = []

    def dedupe(rows):
        seen = set()
        output_rows = []
        for row in rows:
            signature = reminder_schedule_row_signature(row)
            if signature in seen:
                continue
            seen.add(signature)
            output_rows.append(row)
        return output_rows

    all_rows = dedupe(schedule_rows)
    high_frequency_rows = dedupe([row for row in all_rows if row.get("requires_high_frequency")])
    activity_rows = dedupe([
        row
        for row in all_rows
        if row.get("completion_activity") and not row.get("requires_high_frequency")
    ])
    regular_rows = dedupe([
        row
        for row in all_rows
        if not row.get("requires_high_frequency") and not row.get("completion_activity")
    ])
    fill_rows = regular_rows + activity_rows + high_frequency_rows
    fill_cursor = 0

    def fill_daily_rows(existing_rows, target_count, excluded_signatures=None):
        nonlocal fill_cursor
        excluded_signatures = excluded_signatures or set()
        selected = [
            row
            for row in dedupe(existing_rows)
            if reminder_schedule_row_signature(row) not in excluded_signatures
        ]
        if len(selected) >= target_count or not fill_rows:
            return selected

        attempts = 0
        max_attempts = max(len(fill_rows) * 2, target_count)
        while len(selected) < target_count and attempts < max_attempts:
            row = fill_rows[fill_cursor % len(fill_rows)]
            fill_cursor += 1
            attempts += 1
            signature = reminder_schedule_row_signature(row)
            if signature in excluded_signatures:
                continue
            if any(reminder_schedule_row_signature(item) == signature for item in selected):
                continue
            selected.append(row)
        return selected

    for item in REMINDER_WEEK_FLOW:
        recover_classes = buckets.get(item["recover_from"] or "", [])
        recover_signatures = reminder_rows_signature_set(recover_classes)
        new_classes = []
        if item["key"] in {"monday", "tuesday", "friday"}:
            new_classes.extend(high_frequency_rows)
        if item["key"] in {"monday", "friday"}:
            new_classes.extend(activity_rows)
        new_classes = fill_daily_rows(
            new_classes,
            item.get("new_count", REMINDER_DAILY_MIN_CLASSES),
            recover_signatures,
        )
        if new_classes:
            buckets[item["key"]] = new_classes

        output.append(
            {
                "key": item["key"],
                "label": item["label"],
                "recover_from": item.get("recover_from"),
                "new_classes": [
                    reminder_class_ref(row, item["key"], "催课", action_records if include_action_state else None)
                    for row in new_classes
                ],
                "recover_classes": [
                    reminder_class_ref(
                        row,
                        item["key"],
                        "回收",
                        action_records if include_action_state else None,
                        item.get("recover_from"),
                    )
                    for row in recover_classes
                ],
                "focus_classes": [],
            }
        )
    return output


def reminder_schedule_row_signature(row):
    class_keys = sorted(reminder_class_match_keys(row.get("class_name") or row.get("name", "")))
    if class_keys:
        return "|".join(class_keys)
    class_id = str(row.get("class_id") or row.get("id") or "").strip()
    teacher_id = normalize_teacher_id(row.get("teacher_id"))
    class_name = str(row.get("class_name") or row.get("name") or "").strip()
    return f"{teacher_id}|{class_id}|{class_name}"


def reminder_class_is_in_activity(class_name, activity_keys):
    class_keys = reminder_class_match_keys(class_name)
    return bool(class_keys and class_keys & activity_keys)


def reminder_home_classes(visible_teacher_ids):
    store = load_json(classes_file(), {"classes": []})
    rows = []
    for item in store.get("classes", []):
        teacher_id = class_teacher_id(item)
        if visible_teacher_ids and teacher_id not in visible_teacher_ids:
            continue
        if not str(item.get("name", "")).strip():
            continue
        rows.append(item)
    return rows


def reminder_local_class_upload_lookup(visible_teacher_ids):
    store = load_json(classes_file(), {"classes": []})
    lookup = {}
    for item in store.get("classes", []):
        teacher_id = class_teacher_id(item)
        if visible_teacher_ids and teacher_id not in visible_teacher_ids:
            continue
        updated_at = str(item.get("updated_at") or "").strip()
        class_info = {
            "class_id": item.get("id", ""),
            "class_name": item.get("name", ""),
            "updated_at": updated_at,
            "updated_date": updated_at[:10],
            "student_count": len(item.get("students", [])),
        }
        teacher_lookup = lookup.setdefault(teacher_id, {})
        for key in reminder_local_class_keys(item):
            teacher_lookup.setdefault(key, class_info)
    return lookup


def reminder_local_class_upload_info(task, local_lookup, report_date):
    teacher_id = normalize_teacher_id(task.get("teacher_id"))
    teacher_lookup = local_lookup.get(teacher_id, {})
    for key in reminder_class_match_keys(task.get("class_name", "")):
        info = teacher_lookup.get(key)
        if info:
            return {
                **info,
                "matched": True,
                "is_fresh": info.get("updated_date") == report_date,
            }
    return {
        "matched": False,
        "class_id": "",
        "class_name": "",
        "updated_at": "",
        "updated_date": "",
        "student_count": 0,
        "is_fresh": False,
    }


def build_completion_reminder_plan_source(month_key, report_date, visible_teacher_ids, action_records=None, cycle_key=None):
    action_records_for_summary = action_records or []
    completion = build_completion_summary(month_key, report_date)
    home_classes = reminder_home_classes(visible_teacher_ids)
    completion_rows_by_teacher = reminder_completion_rows_by_teacher(completion)
    activity_keys_by_teacher = {teacher_id: set() for teacher_id in visible_teacher_ids}
    for item in home_classes:
        if not item.get("completion_activity"):
            continue
        teacher_id = class_teacher_id(item)
        activity_keys_by_teacher.setdefault(teacher_id, set()).update(reminder_local_class_keys(item))
    teacher_lookup = {
        teacher_id: reminder_teacher_seed(teacher_id)
        for teacher_id in visible_teacher_ids
    }

    for item in home_classes:
        teacher_id = class_teacher_id(item)
        if visible_teacher_ids and teacher_id not in visible_teacher_ids:
            continue
        if teacher_id not in teacher_lookup:
            teacher_lookup[teacher_id] = reminder_teacher_seed(teacher_id)

        completion_row = reminder_completion_row_for_home_class(item, completion_rows_by_teacher)
        row = reminder_home_class_plan_row(item, completion_row)
        row["completion_activity"] = bool(item.get("completion_activity")) or reminder_class_is_in_activity(
            row.get("class_name", ""),
            activity_keys_by_teacher.get(teacher_id, set()),
        )

        teacher_lookup[teacher_id]["class_count"] += 1
        if completion_row:
            teacher_lookup[teacher_id]["database_count"] += 1
        else:
            teacher_lookup[teacher_id]["extra_count"] += 1

        has_comparison = (
            row.get("completion_assessed", True)
            and completion_row
            and row.get("completion_rate") is not None
            and row.get("last_month_completion") is not None
            and row.get("change_from_last_month") is not None
        )
        if has_comparison:
            teacher_lookup[teacher_id]["priorities"].append(row)
        elif completion_row:
            teacher_lookup[teacher_id]["ignored_count"] += 1
            teacher_lookup[teacher_id]["database_fallbacks"].append(row)
        else:
            teacher_lookup[teacher_id]["extra_classes"].append(row)

    teacher_order = {teacher["id"]: index for index, teacher in enumerate(TEACHERS)}
    ordered_groups = []
    for teacher_id, group in sorted(
        teacher_lookup.items(),
        key=lambda item: (teacher_order.get(item[0], 999), item[1]["teacher_name"]),
    ):
        sorted_rows = sorted(
            group["priorities"],
            key=lambda row: (
                not reminder_row_requires_high_frequency(row),
                -(reminder_gap_from_last_month(row) or -999),
                row.get("completion_rate") if row.get("completion_rate") is not None else 101,
                row.get("title_week_number") or 999,
                row.get("class_name", "") or row.get("name", ""),
            ),
        )
        priority_rows = [reminder_priority_row(row, index) for index, row in enumerate(sorted_rows)]
        database_fallback_rows = [
            reminder_database_fallback_row(row)
            for row in sorted(
                group["database_fallbacks"],
                key=lambda row: (
                    not row.get("completion_assessed", True),
                    row.get("completion_rate") if row.get("completion_rate") is not None else 101,
                    row.get("title_week_number") or 999,
                    row.get("class_name", "") or row.get("name", ""),
                ),
            )
        ]
        extra_rows = [
            reminder_database_fallback_row(row)
            for row in sorted(
                group["extra_classes"],
                key=lambda row: (
                    not row.get("completion_assessed", True),
                    row.get("title_week_number") or 999,
                    row.get("class_name", "") or row.get("name", ""),
                ),
            )
        ]
        assessed_database_fallback_rows = [row for row in database_fallback_rows if row.get("completion_assessed", True)]
        renewal_database_fallback_rows = [row for row in database_fallback_rows if not row.get("completion_assessed", True)]
        assessed_extra_rows = [row for row in extra_rows if row.get("completion_assessed", True)]
        renewal_extra_rows = [row for row in extra_rows if not row.get("completion_assessed", True)]
        schedule_rows = (
            priority_rows
            + assessed_database_fallback_rows
            + assessed_extra_rows
            + renewal_database_fallback_rows
            + renewal_extra_rows
        )
        group["priorities"] = priority_rows
        group["database_fallbacks"] = database_fallback_rows
        group["extra_classes"] = extra_rows
        group["included_count"] = len(priority_rows)
        group["schedule_count"] = len(schedule_rows)
        group["schedule"] = build_reminder_schedule(schedule_rows, action_records)
        ordered_groups.append(group)

    return {
        "month": month_key,
        "date": report_date,
        "cycle_key": cycle_key or reminder_cycle_key_for_date(report_date),
        "source": completion.get("source", ""),
        "snapshot_date": completion.get("snapshot_date", ""),
        "last_month_source_month": completion.get("comparison", {}).get("last_month_source_month", ""),
        "last_month_source_date": completion.get("comparison", {}).get("last_month_source_date", ""),
        "groups": ordered_groups,
        "summary": {
            "teacher_count": len(ordered_groups),
            "class_count": sum(group["class_count"] for group in ordered_groups),
            "database_count": sum(group["database_count"] for group in ordered_groups),
            "included_count": sum(group["included_count"] for group in ordered_groups),
            "ignored_count": sum(group["ignored_count"] for group in ordered_groups),
            "extra_count": sum(group["extra_count"] for group in ordered_groups),
            "schedule_count": sum(group["schedule_count"] for group in ordered_groups),
            "completed_action_count": len(action_records_for_summary),
        },
    }


def reminder_all_teacher_ids():
    return {teacher["id"] for teacher in TEACHERS}


def reminder_plan_source_date(cycle_key):
    store = load_completion_snapshots()
    monday_snapshot = completion_snapshot_by_date(cycle_key, store)
    if monday_snapshot:
        return monday_snapshot.get("date") or cycle_key

    source_month = str(cycle_key or "")[:7]
    previous_snapshot = latest_completion_snapshot(source_month, cycle_key, store)
    if previous_snapshot:
        return previous_snapshot.get("date") or cycle_key
    return cycle_key


def generate_weekly_reminder_plan(cycle_key, source_date, generated_by=""):
    source_month = str(source_date or cycle_key)[:7]
    plan = build_completion_reminder_plan_source(
        source_month,
        source_date,
        reminder_all_teacher_ids(),
        action_records=None,
        cycle_key=cycle_key,
    )
    now_text = datetime.now().isoformat(timespec="seconds")
    plan.update(
        {
            "cycle_key": cycle_key,
            "plan_source_date": source_date,
            "generated_at": now_text,
            "generated_by": generated_by,
            "rule_version": REMINDER_PLAN_RULE_VERSION,
            "is_frozen": True,
        }
    )
    return plan


def save_weekly_reminder_plan(cycle_key, source_date, generated_by=""):
    plan = generate_weekly_reminder_plan(cycle_key, source_date, generated_by)
    store = load_reminder_plans()
    previous_plan = store.get("plans", {}).get(cycle_key)
    if previous_plan:
        store.setdefault("history", []).append(previous_plan)
    store.setdefault("plans", {})[cycle_key] = plan
    store["updated_at"] = plan["generated_at"]
    save_reminder_plans(store)
    return plan


def ensure_weekly_reminder_plan(report_date):
    cycle_key = reminder_cycle_key_for_date(report_date)
    store = load_reminder_plans()
    plan = store.get("plans", {}).get(cycle_key)
    if plan:
        if plan.get("rule_version") != REMINDER_PLAN_RULE_VERSION:
            source_date = plan.get("plan_source_date") or reminder_plan_source_date(cycle_key)
            with REMINDER_PLANS_LOCK:
                return save_weekly_reminder_plan(cycle_key, source_date, "system-rule-update")
        if reconcile_weekly_reminder_plan(plan):
            store.setdefault("plans", {})[cycle_key] = plan
            store["updated_at"] = plan.get("updated_at", datetime.now().isoformat(timespec="seconds"))
            save_reminder_plans(store)
        return plan

    source_date = reminder_plan_source_date(cycle_key)
    with REMINDER_PLANS_LOCK:
        store = load_reminder_plans()
        plan = store.get("plans", {}).get(cycle_key)
        if plan:
            if plan.get("rule_version") != REMINDER_PLAN_RULE_VERSION:
                source_date = plan.get("plan_source_date") or source_date
                return save_weekly_reminder_plan(cycle_key, source_date, "system-rule-update")
            if reconcile_weekly_reminder_plan(plan):
                store.setdefault("plans", {})[cycle_key] = plan
                store["updated_at"] = plan.get("updated_at", datetime.now().isoformat(timespec="seconds"))
                save_reminder_plans(store)
            return plan
        return save_weekly_reminder_plan(cycle_key, source_date, "system")


def refresh_weekly_reminder_plan_from_snapshot(snapshot_date):
    if datetime.strptime(snapshot_date, "%Y-%m-%d").weekday() != 0:
        return None
    cycle_key = reminder_cycle_key_for_date(snapshot_date)
    with REMINDER_PLANS_LOCK:
        return save_weekly_reminder_plan(cycle_key, snapshot_date, g.user.get("username", ""))


def reminder_schedule_with_action_states(schedule, action_records, local_upload_lookup=None, report_date=""):
    output = []
    for day in schedule or []:
        day_key = day.get("key", "")
        recover_from = day.get("recover_from")
        output.append(
            {
                **day,
                "new_classes": [
                    reminder_class_ref(row, day_key, "催课", action_records, "", local_upload_lookup, report_date)
                    for row in day.get("new_classes", [])
                ],
                "recover_classes": [
                    reminder_class_ref(row, day_key, "回收", action_records, recover_from, local_upload_lookup, report_date)
                    for row in day.get("recover_classes", [])
                ],
                "focus_classes": [
                    reminder_class_ref(row, day_key, "重点复催", action_records, "", local_upload_lookup, report_date)
                    for row in day.get("focus_classes", [])
                ],
            }
        )
    return output


def reminder_plan_summary(groups, action_records):
    return {
        "teacher_count": len(groups),
        "class_count": sum(group.get("class_count", 0) for group in groups),
        "database_count": sum(group.get("database_count", 0) for group in groups),
        "included_count": sum(group.get("included_count", 0) for group in groups),
        "ignored_count": sum(group.get("ignored_count", 0) for group in groups),
        "extra_count": sum(group.get("extra_count", 0) for group in groups),
        "schedule_count": sum(group.get("schedule_count", 0) for group in groups),
        "completed_action_count": len(action_records),
    }


def build_completion_reminder_plan(month_key, report_date):
    visible_teacher_ids = reminder_visible_teacher_ids()
    action_records = reminder_schedule_action_records(visible_teacher_ids)
    local_upload_lookup = reminder_local_class_upload_lookup(visible_teacher_ids)
    plan = ensure_weekly_reminder_plan(report_date)
    day_key = reminder_day_key_for_date(report_date)
    groups = [
        {
            **group,
            "schedule": reminder_schedule_with_action_states(
                group.get("schedule", []),
                action_records,
                local_upload_lookup,
                report_date,
            ),
        }
        for group in plan.get("groups", [])
        if not visible_teacher_ids or normalize_teacher_id(group.get("teacher_id")) in visible_teacher_ids
    ]
    return {
        "month": plan.get("month") or month_key,
        "date": plan.get("date") or report_date,
        "cycle_key": plan.get("cycle_key") or reminder_cycle_key_for_date(report_date),
        "day_key": day_key,
        "day_label": reminder_day_label(day_key) if day_key else "周末",
        "is_workday": bool(day_key),
        "can_manage_all": current_user_teacher_id() == COMPLETION_UPLOAD_TEACHER_ID,
        "source": plan.get("source", ""),
        "snapshot_date": plan.get("snapshot_date", ""),
        "plan_source_date": plan.get("plan_source_date") or plan.get("snapshot_date", ""),
        "generated_at": plan.get("generated_at", ""),
        "generated_by": plan.get("generated_by", ""),
        "is_frozen": True,
        "last_month_source_month": plan.get("last_month_source_month", ""),
        "last_month_source_date": plan.get("last_month_source_date", ""),
        "groups": groups,
        "summary": reminder_plan_summary(groups, action_records),
    }


def reminder_day_tasks(day):
    if not day:
        return []
    tasks = []
    for task_label, key in (("催课", "new_classes"), ("回收", "recover_classes"), ("重点复催", "focus_classes")):
        for item in day.get(key, []) or []:
            tasks.append({**item, "task_label": task_label})
    return tasks


def build_completion_reminder_supervision(month_key, report_date):
    plan = build_completion_reminder_plan(month_key, report_date)
    visible_teacher_ids = reminder_visible_teacher_ids()
    local_lookup = reminder_local_class_upload_lookup(visible_teacher_ids)
    day_key = reminder_day_key_for_date(report_date)
    day_label = reminder_day_label(day_key) if day_key else "周末"
    groups = []
    summary = {
        "teacher_count": 0,
        "task_count": 0,
        "completed_count": 0,
        "pending_count": 0,
        "fresh_upload_count": 0,
        "stale_upload_count": 0,
        "missing_local_count": 0,
    }

    for group in plan.get("groups", []):
        today = next((item for item in group.get("schedule", []) if item.get("key") == day_key), None)
        tasks = []
        for task in reminder_day_tasks(today):
            upload_info = reminder_local_class_upload_info(task, local_lookup, report_date)
            action_state = task.get("action_state") or {}
            is_completed = bool(action_state.get("completed"))
            is_fresh = bool(upload_info.get("is_fresh"))
            is_matched = bool(upload_info.get("matched"))
            tasks.append(
                {
                    "class_id": task.get("class_id", ""),
                    "class_name": task.get("class_name", ""),
                    "teacher_id": task.get("teacher_id", ""),
                    "teacher_name": task.get("teacher_name", ""),
                    "task_label": task.get("task_label", ""),
                    "source": task.get("source", ""),
                    "source_label": task.get("source_label", ""),
                    "completion_assessed": bool(task.get("completion_assessed", True)),
                    "requires_high_frequency": bool(task.get("requires_high_frequency")),
                    "title_week_label": task.get("title_week_label", ""),
                    "completed": is_completed,
                    "completed_at": action_state.get("completed_at", ""),
                    "completed_label": action_state.get("label", ""),
                    "upload": upload_info,
                }
            )
            summary["task_count"] += 1
            summary["completed_count" if is_completed else "pending_count"] += 1
            if is_fresh:
                summary["fresh_upload_count"] += 1
            elif is_matched:
                summary["stale_upload_count"] += 1
            else:
                summary["missing_local_count"] += 1

        if tasks:
            summary["teacher_count"] += 1
            groups.append(
                {
                    "teacher_id": group.get("teacher_id", ""),
                    "teacher_name": group.get("teacher_name", ""),
                    "tasks": tasks,
                    "task_count": len(tasks),
                    "completed_count": sum(1 for task in tasks if task.get("completed")),
                    "pending_count": sum(1 for task in tasks if not task.get("completed")),
                    "fresh_upload_count": sum(1 for task in tasks if task.get("upload", {}).get("is_fresh")),
                    "stale_upload_count": sum(
                        1
                        for task in tasks
                        if task.get("upload", {}).get("matched") and not task.get("upload", {}).get("is_fresh")
                    ),
                    "missing_local_count": sum(1 for task in tasks if not task.get("upload", {}).get("matched")),
                }
            )

    return {
        "month": plan.get("month") or month_key,
        "date": report_date,
        "cycle_key": plan.get("cycle_key") or reminder_cycle_key_for_date(report_date),
        "day_key": day_key,
        "day_label": day_label,
        "is_workday": bool(day_key),
        "plan_source_date": plan.get("plan_source_date") or plan.get("snapshot_date", ""),
        "is_frozen": bool(plan.get("is_frozen", True)),
        "groups": groups,
        "summary": summary,
    }


def editable_teacher_id():
    return current_user_teacher_id()


def class_teacher_lookup():
    store = load_json(classes_file(), {"classes": []})
    lookup = {}
    for item in store.get("classes", []):
        class_id = str(item.get("id", "")).strip()
        if not class_id:
            continue
        lookup[class_id] = class_teacher_id(item)
    for item in load_completion_assignments():
        class_id = str(item.get("id") or completion_class_id(item.get("name", ""))).strip()
        teacher_id = normalize_teacher_id(item.get("teacher_id"))
        if class_id and teacher_id:
            lookup[class_id] = teacher_id
    for snapshot in snapshot_list():
        for row in snapshot.get("rows", []):
            class_id = str(row.get("id", "")).strip()
            teacher_id = normalize_teacher_id(row.get("teacher_id"))
            if class_id and teacher_id:
                lookup[class_id] = teacher_id
    return lookup


def setting_allowed_for_teacher(teacher_id):
    current_teacher = editable_teacher_id()
    return bool(current_teacher and teacher_id == current_teacher)


@database_bp.put("/learning-settings")
@login_required
def update_learning_settings():
    payload = request.get_json(silent=True) or {}
    class_lookup = class_teacher_lookup()

    with DATABASE_SETTINGS_LOCK:
        settings = load_database_settings()
        learning = settings.setdefault("learning", {})
        class_settings = learning.setdefault("classes", {})
        teacher_settings = learning.setdefault("teachers", {})

        for item in payload.get("classes", []):
            class_id = str(item.get("class_id") or item.get("id") or "").strip()
            teacher_id = class_lookup.get(class_id)
            if not class_id or not teacher_id or not setting_allowed_for_teacher(teacher_id):
                continue
            class_settings[class_id] = {
                "coefficient": rounded_metric(parse_float(item.get("coefficient"), 0))
            }

        for item in payload.get("teachers", []):
            teacher_id = normalize_teacher_id(item.get("teacher_id"))
            if not teacher_id or not setting_allowed_for_teacher(teacher_id):
                continue
            teacher_settings[teacher_id] = {
                "target_rate": normalize_target_rate(item.get("target_rate"))
            }

        save_database_settings(settings)

    return jsonify({"ok": True})


@database_bp.post("/completion-upload")
@login_required
def upload_completion_snapshot():
    if not can_upload_completion_data():
        return jsonify({"error": "只有文云Joanna账号可以上传完课数据。"}), 403

    file_storage = request.files.get("file")
    if not file_storage:
        return jsonify({"error": "请上传 Excel 或 CSV 文件。"}), 400

    try:
        snapshot_date = normalize_date(request.form.get("date"))
        rows = parse_completion_upload(file_storage)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    snapshot = {
        "date": snapshot_date,
        "month": snapshot_date[:7],
        "uploaded_at": datetime.now().isoformat(timespec="seconds"),
        "uploaded_by": g.user.get("username", ""),
        "rows": rows,
    }

    with COMPLETION_SNAPSHOTS_LOCK:
        store = load_completion_snapshots()
        store.setdefault("snapshots", {})[snapshot_date] = snapshot
        save_completion_snapshots(store)

    reminder_plan = refresh_weekly_reminder_plan_from_snapshot(snapshot_date)

    return jsonify(
        {
            "ok": True,
            "snapshot": {
                "date": snapshot_date,
                "row_count": len(rows),
                "uploaded_at": snapshot["uploaded_at"],
            },
            "reminder_plan": {
                "updated": bool(reminder_plan),
                "cycle_key": reminder_plan.get("cycle_key", "") if reminder_plan else "",
                "source_date": reminder_plan.get("plan_source_date", "") if reminder_plan else "",
            },
        }
    )


@database_bp.post("/completion-last-month-upload")
@login_required
def upload_completion_last_month():
    if not can_upload_completion_data():
        return jsonify({"error": "只有文云Joanna账号可以上传上月完课数据。"}), 403

    file_storage = request.files.get("file")
    if not file_storage:
        return jsonify({"error": "请上传 Excel 或 CSV 文件。"}), 400

    try:
        target_month = normalize_month(request.form.get("month"))
        rows = parse_completion_upload(file_storage)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    snapshot = {
        "month": target_month,
        "uploaded_at": datetime.now().isoformat(timespec="seconds"),
        "uploaded_by": g.user.get("username", ""),
        "rows": rows,
    }

    with COMPLETION_SNAPSHOTS_LOCK:
        store = load_completion_snapshots()
        store.setdefault("last_month", {})[target_month] = snapshot
        save_completion_snapshots(store)

    return jsonify(
        {
            "ok": True,
            "snapshot": {
                "month": target_month,
                "row_count": len(rows),
                "uploaded_at": snapshot["uploaded_at"],
            },
        }
    )


@database_bp.put("/gmv-adjustments")
@login_required
def update_gmv_adjustments():
    if not can_manage_gmv():
        return jsonify({"error": "只有文云Joanna账号可以修正GMV数据。"}), 403

    payload = request.get_json(silent=True) or {}
    try:
        month_key = normalize_month(payload.get("month"))
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    if isinstance(payload.get("sections"), dict):
        sections_payload = payload.get("sections")
    else:
        section_key = str(payload.get("section") or "").strip()
        sections_payload = {section_key: payload.get("rows", [])}
    targets_payload = payload.get("targets")

    with DATABASE_SETTINGS_LOCK:
        settings = load_database_settings()
        gmv_settings = settings.setdefault("gmv", {})
        month_settings = gmv_settings.setdefault(month_key, {})

        for section_key, rows in sections_payload.items():
            if section_key not in GMV_SECTIONS:
                return jsonify({"error": "GMV板块不存在。"}), 400
            if not isinstance(rows, list):
                return jsonify({"error": "GMV修正数据格式不正确。"}), 400

            section_settings = {}
            for item in rows:
                teacher_id = normalize_teacher_id(item.get("teacher_id"))
                if not teacher_id:
                    continue
                week_overrides = normalize_gmv_week_overrides(item.get("week_totals"))
                if any(value is not None for value in week_overrides):
                    section_settings[teacher_id] = {"week_totals": week_overrides}
            if section_settings:
                month_settings[section_key] = section_settings
            else:
                month_settings.pop(section_key, None)

        if isinstance(targets_payload, dict):
            target_settings = {}
            for section_key in GMV_SECTIONS:
                if section_key not in targets_payload:
                    continue
                target_value = parse_gmv_amount(targets_payload.get(section_key))
                if target_value is not None:
                    target_settings[section_key] = target_value
            if target_settings:
                month_settings["targets"] = target_settings
            else:
                month_settings.pop("targets", None)

        if month_settings:
            gmv_settings[month_key] = month_settings
        else:
            gmv_settings.pop(month_key, None)

        save_database_settings(settings)

    return jsonify({"ok": True})


@database_bp.post("/monthly-archives")
@login_required
def create_monthly_archive():
    if not can_upload_completion_data():
        return jsonify({"error": "只有文云Joanna账号可以进行月度存档。"}), 403

    payload = request.get_json(silent=True) or {}
    try:
        month_key = normalize_month(payload.get("month"))
        report_date = normalize_date(payload.get("date") or month_end_date(month_key))
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    if not report_date.startswith(f"{month_key}-"):
        report_date = month_end_date(month_key)

    archive = build_monthly_archive(month_key, report_date)
    with MONTHLY_ARCHIVES_LOCK:
        store = load_monthly_archives()
        previous_archive = store.get("archives", {}).get(month_key)
        if previous_archive:
            store.setdefault("history", []).append(previous_archive)
        store.setdefault("archives", {})[month_key] = archive
        store["updated_at"] = archive["archived_at"]
        save_monthly_archives(store)

    return jsonify(
        {
            "ok": True,
            "archive": {
                "month": archive["month"],
                "date": archive["date"],
                "archived_at": archive["archived_at"],
                "archived_by": archive["archived_by"],
                "daily_report_count": archive["daily_report_count"],
                "summary": archive["summary"],
            },
            "next_month": archive["next_month"],
            "next_date": archive["next_date"],
            "message": f"{archive['month']} 已完成存档，已为下个月重新开始统计。",
        }
    )


@database_bp.get("/summary")
@login_required
def database_summary():
    try:
        month_key = normalize_month(request.args.get("month"))
        report_date = normalize_date(request.args.get("date"))
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    if not report_date.startswith(f"{month_key}-"):
        current_date = datetime.now().strftime("%Y-%m-%d")
        report_date = current_date if current_date.startswith(f"{month_key}-") else month_end_date(month_key)

    completion = build_completion_summary(month_key, report_date, request.args.get("compare_date"))
    completion_teachers = completion["teachers"]
    return jsonify(
        {
            "month": month_key,
            "date": report_date,
            "completion": completion,
            "learning": build_learning_summary(month_key, report_date, completion),
            "renewal": build_renewal_summary(month_key, report_date, completion_teachers),
            "referral": build_referral_summary(month_key, report_date, completion_teachers),
            "gmv": build_gmv_summary(month_key, report_date, completion_teachers),
            "completion_performance": build_completion_performance_summary(completion),
            "permissions": {
                "can_upload_completion": can_upload_completion_data(),
                "can_manage_gmv": can_manage_gmv(),
            },
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }
    )


@database_bp.get("/completion-reminders")
@login_required
def completion_reminders():
    try:
        month_key = normalize_month(request.args.get("month"))
        report_date = normalize_date(request.args.get("date"))
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    if not report_date.startswith(f"{month_key}-"):
        current_date = datetime.now().strftime("%Y-%m-%d")
        report_date = current_date if current_date.startswith(f"{month_key}-") else month_end_date(month_key)

    return jsonify(build_completion_reminder_plan(month_key, report_date))


@database_bp.get("/completion-reminders/supervision")
@login_required
def completion_reminder_supervision():
    try:
        month_key = normalize_month(request.args.get("month"))
        report_date = normalize_date(request.args.get("date"))
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    if not report_date.startswith(f"{month_key}-"):
        current_date = datetime.now().strftime("%Y-%m-%d")
        report_date = current_date if current_date.startswith(f"{month_key}-") else month_end_date(month_key)

    return jsonify(build_completion_reminder_supervision(month_key, report_date))


def serialize_reminder_record(record):
    return {
        "id": record.get("id", ""),
        "cycle_key": record.get("cycle_key", ""),
        "teacher_id": record.get("teacher_id", ""),
        "teacher_name": record.get("teacher_name", ""),
        "class_name": record.get("class_name", ""),
        "local_class_name": record.get("local_class_name", ""),
        "origin_day_key": record.get("origin_day_key", ""),
        "origin_day_label": record.get("origin_day_label", ""),
        "recovery_day_key": record.get("recovery_day_key", ""),
        "recovery_day_label": record.get("recovery_day_label", ""),
        "task_label": record.get("task_label", ""),
        "student_count": len(record.get("students", [])),
        "students": record.get("students", []),
        "created_at": record.get("created_at", ""),
        "completed_at": record.get("completed_at", ""),
        "completed_by": record.get("completed_by", ""),
        "recovered_at": record.get("recovered_at", ""),
        "recovered_by": record.get("recovered_by", ""),
        "status": record.get("status", ""),
    }


@database_bp.get("/completion-reminders/action-records")
@login_required
def completion_reminder_action_records():
    class_name = str(request.args.get("class_name") or "").strip()
    day_key = str(request.args.get("day_key") or "").strip()
    task_label = str(request.args.get("task_label") or "催课").strip()
    recover_from = str(request.args.get("recover_from") or "").strip()
    teacher_id = normalize_teacher_id(request.args.get("teacher_id"))

    if not class_name or not day_key:
        return jsonify({"records": [], "student_count": 0})
    if teacher_id and not can_access_reminder_teacher(teacher_id):
        return jsonify({"error": "只能查看自己有权限的催课记录。"}), 403

    store = load_reminder_actions()
    records = []
    for record in store.get("records", []):
        record_teacher_id = normalize_teacher_id(record.get("teacher_id"))
        if teacher_id and record_teacher_id != teacher_id:
            continue
        if not can_access_reminder_teacher(record_teacher_id):
            continue
        row = {"class_name": class_name, "teacher_id": teacher_id or record_teacher_id}
        if reminder_record_matches_schedule(record, row, day_key, task_label, recover_from):
            records.append(serialize_reminder_record(record))

    records = sorted(
        records,
        key=lambda item: item.get("recovered_at") or item.get("completed_at") or item.get("created_at") or "",
        reverse=True,
    )
    return jsonify(
        {
            "records": records,
            "student_count": sum(record.get("student_count", 0) for record in records),
        }
    )


@database_bp.get("/completion-reminders/recovery-records")
@login_required
def completion_reminder_recovery_records():
    class_name = str(request.args.get("class_name") or "").strip()
    day_key = str(request.args.get("day_key") or "").strip()
    recover_from = str(request.args.get("recover_from") or "").strip()
    teacher_id = normalize_teacher_id(request.args.get("teacher_id"))

    if not class_name or not day_key:
        return jsonify({"records": [], "student_count": 0})
    if teacher_id and not can_access_reminder_teacher(teacher_id):
        return jsonify({"error": "只能查看自己的催课回收记录。"}), 403

    store = load_reminder_actions()
    records = []
    for record in store.get("records", []):
        record_teacher_id = normalize_teacher_id(record.get("teacher_id"))
        if teacher_id and record_teacher_id != teacher_id:
            continue
        if not can_access_reminder_teacher(record_teacher_id):
            continue
        if reminder_record_matches(record, class_name, day_key, recover_from):
            records.append(serialize_reminder_record(record))

    return jsonify(
        {
            "records": records,
            "student_count": sum(record.get("student_count", 0) for record in records),
        }
    )


@database_bp.post("/completion-reminders/actions")
@login_required
def save_completion_reminder_action():
    payload = request.get_json(silent=True) or {}
    class_name = str(payload.get("class_name") or "").strip()
    day_key = str(payload.get("day_key") or "").strip()
    task_label = str(payload.get("task_label") or "催课").strip()
    target_teacher_id = reminder_target_teacher_id(payload)

    if not class_name:
        return jsonify({"error": "缺少催课班级。"}), 400
    if not day_key:
        return jsonify({"error": "缺少催课日期。"}), 400
    if not target_teacher_id or not can_access_reminder_teacher(target_teacher_id):
        return jsonify({"error": "只能保存自己的催课记录。"}), 403

    cycle_key = current_reminder_cycle_key()
    now_text = datetime.now().isoformat(timespec="seconds")
    class_keys = reminder_payload_class_keys(payload)
    if not class_keys:
        class_keys = sorted(reminder_class_match_keys(class_name))

    if task_label == "回收":
        recover_from = str(payload.get("recover_from") or "").strip()
        def mark_recovered(store):
            completed_count = 0
            store.setdefault("records", [])
            for record in store.get("records", []):
                if normalize_teacher_id(record.get("teacher_id")) != target_teacher_id:
                    continue
                if not reminder_record_matches(record, class_name, day_key, recover_from):
                    continue
                record["status"] = "recovered"
                record["recovered_at"] = now_text
                record["recovered_by"] = g.user.get("username", "")
                completed_count += 1
            store["updated_at"] = now_text
            return completed_count

        with REMINDER_ACTIONS_LOCK:
            completed_count = update_json(
                completion_reminder_actions_file(),
                {"records": []},
                mark_recovered,
            )
        return jsonify(
            {
                "ok": True,
                "action": "recovered",
                "completed_count": completed_count,
                "message": "已完成回收，待回收名单已清空。",
            }
        )

    students = [
        reminder_student_snapshot(student)
        for student in payload.get("students", [])
        if isinstance(student, dict)
    ]
    recovery_target = reminder_recovery_target(day_key) if task_label == "催课" else None
    status = "pending_recovery" if recovery_target and students else "completed"

    record = {
        "id": uuid.uuid4().hex,
        "cycle_key": cycle_key,
        "teacher_id": target_teacher_id,
        "teacher_name": teacher_label(target_teacher_id),
        "class_id": str(payload.get("class_id") or ""),
        "class_name": class_name,
        "local_class_id": str(payload.get("local_class_id") or ""),
        "local_class_name": str(payload.get("local_class_name") or "").strip(),
        "class_keys": class_keys,
        "source": str(payload.get("source") or ""),
        "origin_day_key": day_key,
        "origin_day_label": str(payload.get("day_label") or reminder_day_label(day_key)).strip(),
        "recovery_day_key": recovery_target["key"] if recovery_target else "",
        "recovery_day_label": recovery_target["label"] if recovery_target else "",
        "task_label": task_label,
        "completion_rate": parse_completion(payload.get("completion_rate")),
        "last_month_completion": parse_completion(payload.get("last_month_completion")),
        "change_from_last_month": parse_completion(payload.get("change_from_last_month")),
        "students": students,
        "status": status,
        "created_at": now_text,
        "completed_at": now_text,
        "completed_by": g.user.get("username", ""),
    }

    def same_action(existing):
        if existing.get("cycle_key") != cycle_key:
            return False
        if normalize_teacher_id(existing.get("teacher_id")) != target_teacher_id:
            return False
        if str(existing.get("origin_day_key") or "") != day_key:
            return False
        if str(existing.get("task_label") or "") != task_label:
            return False
        existing_keys = set(existing.get("class_keys") or reminder_class_match_keys(existing.get("class_name", "")))
        return bool(existing_keys.intersection(class_keys))

    def append_action_record(store):
        store.setdefault("records", [])
        store["records"] = [
            existing
            for existing in store.get("records", [])
            if not same_action(existing)
        ]
        store["records"].append(record)
        store["updated_at"] = now_text

    with REMINDER_ACTIONS_LOCK:
        update_json(
            completion_reminder_actions_file(),
            {"records": []},
            append_action_record,
        )

    if status == "pending_recovery":
        message = f"已完成催课，{len(students)}名学员已同步到{recovery_target['label']}回收。"
    elif students:
        message = f"已记录本次{task_label}完成。"
    else:
        message = "本次没有需要催课的学员，已记录完成。"

    return jsonify(
        {
            "ok": True,
            "action": "saved",
            "status": status,
            "record": serialize_reminder_record(record),
            "recovery_target": recovery_target,
            "message": message,
        }
    )
