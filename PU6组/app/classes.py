import csv
import io
import json
import re
import uuid
from datetime import datetime, timedelta

from flask import Blueprint, current_app, g, jsonify, request, send_from_directory, url_for
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from app.auth import can_manage_accounts, login_required
from app.teachers import (
    infer_teacher_id_from_class_name,
    normalize_teacher_id,
    teacher_id_for_username,
    teacher_label,
    teacher_options,
)


classes_bp = Blueprint("classes", __name__, url_prefix="/api/classes")

WEEK_COUNT = 4
DAY_COUNT = 6
WEEK_SECONDS = 7 * 24 * 60 * 60
HEADER_SCAN_LIMIT = 20
ACTIVITY_STATUSES = {"draft", "active", "ended"}
ACTIVITY_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}
DEFAULT_ACTIVITY_STAGE_LABELS = ["盲盒种子", "发芽", "抽枝", "神秘小树苗", "惊喜绽放"]
DEFAULT_ACTIVITY_RESULT_LABELS = ["彩虹花", "向日葵", "樱花树", "蓝绣球", "小橘树", "紫铃兰"]
ACTIVITY_STAGE_LIMIT = 8
ACTIVITY_RESULT_LIMIT = 12
DEFAULT_COMPLETION_ACTIVITY = {
    "eyebrow": "6月完课活动",
    "title": "盲盒种子成长计划",
    "description": "参与班级的每位学员获得 1 颗虚拟盲盒种子。每周任务完成可获得 1 次浇水机会，满 4 次后植物长大；补交完成后也会自动补上进度。",
}
CHINESE_NUMBERS = {
    1: "一",
    2: "二",
    3: "三",
    4: "四",
    5: "五",
    6: "六",
}

NAME_COLUMNS = {
    "姓名",
    "名字",
    "学员",
    "学生",
    "学员姓名",
    "学生姓名",
    "孩子姓名",
    "宝贝姓名",
    "用户姓名",
    "昵称",
    "name",
    "studentname",
}
ACCOUNT_COLUMNS = {
    "账号",
    "帐号",
    "学号",
    "学员账号",
    "学员帐号",
    "学生账号",
    "学生帐号",
    "用户账号",
    "用户帐号",
    "账号id",
    "学员id",
    "学生id",
    "account",
    "studentaccount",
    "studentid",
    "userid",
    "id",
}


def classes_file():
    return current_app.config["CLASSES_FILE"]


def renewal_projects_file():
    return current_app.config["RENEWAL_PROJECTS_FILE"]


def completion_activities_file():
    return current_app.config["COMPLETION_ACTIVITIES_FILE"]


def completion_activity_asset_dir():
    return current_app.config["COMPLETION_ACTIVITY_ASSET_DIR"]


def now_iso():
    return datetime.now().isoformat(timespec="seconds")


def current_month_key():
    return datetime.now().strftime("%Y-%m")


def local_date_key(value=None):
    target = value or datetime.now()
    return target.strftime("%Y-%m-%d")


def export_cycle_anchor(value=None):
    target = value or datetime.now()
    anchor = datetime(target.year, target.month, target.day)
    monday_based_day = anchor.weekday()
    offset = 4 - monday_based_day if monday_based_day >= 4 else -(monday_based_day + 3)
    return anchor + timedelta(days=offset)


def parse_title_week_number(value):
    text = str(value or "").strip()
    if not text:
        return None
    text = re.sub(r"^[wW]\s*", "", text)
    if not re.fullmatch(r"\d+", text):
        raise ValueError("周数 W 请输入数字，例如 23。")
    number = int(text)
    if number < 1 or number > 99:
        raise ValueError("周数 W 请输入 1-99 之间的数字。")
    return number


def parse_date_key(value):
    try:
        return datetime.strptime(str(value or ""), "%Y-%m-%d")
    except ValueError:
        return None


def current_title_week_number(item):
    try:
        base_number = parse_title_week_number(
            item.get("title_week_base_number") or item.get("title_week_number")
        )
    except ValueError:
        return None
    if not base_number:
        return None
    anchor = parse_date_key(item.get("title_week_anchor"))
    if anchor is None:
        return base_number
    elapsed_weeks = max(0, int((export_cycle_anchor() - anchor).total_seconds() // WEEK_SECONDS))
    return base_number + elapsed_weeks


def load_store():
    path = classes_file()
    if not path.exists():
        return {"classes": []}
    with path.open("r", encoding="utf-8") as file:
        return json.load(file)


def save_store(store):
    path = classes_file()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump(store, file, ensure_ascii=False, indent=2)


def load_renewal_store():
    path = renewal_projects_file()
    if not path.exists():
        return {"projects": []}
    try:
        with path.open("r", encoding="utf-8") as file:
            store = json.load(file)
    except (json.JSONDecodeError, OSError):
        return {"projects": []}
    if not isinstance(store, dict) or not isinstance(store.get("projects"), list):
        return {"projects": []}
    return store


def save_renewal_store(store):
    path = renewal_projects_file()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump(store, file, ensure_ascii=False, indent=2)


def prune_renewal_followups_for_class(class_id, valid_student_ids):
    class_key = str(class_id or "").strip()
    if not class_key:
        return 0

    valid_ids = {str(student_id) for student_id in valid_student_ids if str(student_id or "").strip()}
    store = load_renewal_store()
    removed_count = 0
    changed = False

    for project in store.get("projects", []):
        if str(project.get("class_id") or "").strip() != class_key:
            continue

        project_changed = False
        followups = project.get("student_followups")
        if isinstance(followups, dict):
            for student_id in list(followups.keys()):
                if str(student_id) not in valid_ids:
                    followups.pop(student_id, None)
                    removed_count += 1
                    changed = True
                    project_changed = True

        enrolled_ids = project.get("enrolled_student_ids")
        if isinstance(enrolled_ids, list):
            next_enrolled_ids = [
                student_id
                for student_id in enrolled_ids
                if str(student_id) in valid_ids
            ]
            if len(next_enrolled_ids) != len(enrolled_ids):
                project["enrolled_student_ids"] = next_enrolled_ids
                changed = True
                project_changed = True

        if project_changed:
            project["updated_at"] = now_iso()

    if changed:
        save_renewal_store(store)
    return removed_count


def renewal_enrolled_student_ids(class_id):
    class_key = str(class_id or "").strip()
    if not class_key:
        return set()
    enrolled_ids = set()
    for project in load_renewal_store().get("projects", []):
        if str(project.get("class_id") or "").strip() != class_key:
            continue
        followups = project.get("student_followups") if isinstance(project.get("student_followups"), dict) else {}
        for student_id, record in followups.items():
            if isinstance(record, dict) and record.get("enrolled"):
                enrolled_ids.add(str(student_id))
        for student_id in project.get("enrolled_student_ids", []):
            student_key = str(student_id or "").strip()
            if student_key and student_key not in followups:
                enrolled_ids.add(student_key)
    return enrolled_ids


def normalize_activity_status(value):
    status = str(value or "").strip().lower()
    return status if status in ACTIVITY_STATUSES else "draft"


def activity_participant_ids(item):
    values = item.get("participant_class_ids")
    if not isinstance(values, list):
        return []
    return [str(value) for value in values if str(value or "").strip()]


def split_activity_labels(value, defaults, limit):
    if isinstance(value, list):
        labels = [str(item or "").strip() for item in value]
    else:
        labels = re.split(r"[,，、\n]+", str(value or ""))
        labels = [item.strip() for item in labels]
    labels = [item[:30] for item in labels if item]
    return labels[:limit] or defaults[:limit]


def normalize_activity_image_map(value, limit):
    if not isinstance(value, dict):
        return {}
    output = {}
    for raw_key, raw_asset in value.items():
        try:
            index = int(raw_key)
        except (TypeError, ValueError):
            continue
        if index < 0 or index >= limit or not isinstance(raw_asset, dict):
            continue
        filename = secure_filename(str(raw_asset.get("filename") or ""))
        if not filename:
            continue
        output[str(index)] = {
            "filename": filename,
            "updated_at": str(raw_asset.get("updated_at") or ""),
        }
    return output


def default_activity_visuals():
    return {
        "stage_labels": DEFAULT_ACTIVITY_STAGE_LABELS[:],
        "result_labels": DEFAULT_ACTIVITY_RESULT_LABELS[:],
        "stage_images": {},
        "result_images": {},
        "image_note": "每完成一周任务即可推进一次进度，满进度后揭晓专属结果。",
        "image_footer": "本图统计本月当前累计活动进度，补交完成后会自动更新。",
    }


def normalize_activity_visuals(value=None):
    visuals = default_activity_visuals()
    if not isinstance(value, dict):
        return visuals
    visuals["stage_labels"] = split_activity_labels(
        value.get("stage_labels"),
        DEFAULT_ACTIVITY_STAGE_LABELS,
        ACTIVITY_STAGE_LIMIT,
    )
    visuals["result_labels"] = split_activity_labels(
        value.get("result_labels"),
        DEFAULT_ACTIVITY_RESULT_LABELS,
        ACTIVITY_RESULT_LIMIT,
    )
    visuals["stage_images"] = normalize_activity_image_map(
        value.get("stage_images"),
        len(visuals["stage_labels"]),
    )
    visuals["result_images"] = normalize_activity_image_map(
        value.get("result_images"),
        len(visuals["result_labels"]),
    )
    visuals["image_note"] = str(value.get("image_note") or visuals["image_note"] or "").strip()[:120]
    visuals["image_footer"] = str(value.get("image_footer") or visuals["image_footer"] or "").strip()[:160]
    return visuals


def serialize_activity_visuals(item):
    visuals = normalize_activity_visuals(item.get("visuals"))
    for image_group in ("stage_images", "result_images"):
        for asset in visuals[image_group].values():
            asset["url"] = url_for("classes.activity_asset", filename=asset["filename"])
    return visuals


def build_completion_activity(fields=None, status="draft", participant_ids=None):
    timestamp = now_iso()
    data = {
        "id": uuid.uuid4().hex,
        **DEFAULT_COMPLETION_ACTIVITY,
        "visuals": default_activity_visuals(),
        "status": normalize_activity_status(status),
        "participant_class_ids": list(dict.fromkeys(participant_ids or [])),
        "created_at": timestamp,
        "updated_at": timestamp,
        "published_at": timestamp if normalize_activity_status(status) == "active" else "",
        "ended_at": "",
    }
    update_activity_fields(data, fields or {})
    return data


def load_activity_store(class_store=None):
    path = completion_activities_file()
    if not path.exists():
        source_store = class_store if class_store is not None else load_store()
        participant_ids = [
            item.get("id")
            for item in source_store.get("classes", [])
            if item.get("id") and item.get("completion_activity")
        ]
        store = {
            "activities": [
                build_completion_activity(
                    status="active",
                    participant_ids=participant_ids,
                )
            ]
        }
        save_activity_store(store)
        return store
    try:
        with path.open("r", encoding="utf-8") as file:
            store = json.load(file)
    except (json.JSONDecodeError, OSError):
        store = {"activities": []}
    if not isinstance(store, dict):
        store = {"activities": []}
    if not isinstance(store.get("activities"), list):
        store["activities"] = []
    return store


def save_activity_store(store):
    path = completion_activities_file()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump(store, file, ensure_ascii=False, indent=2)


def active_completion_activity(activity_store=None):
    store = activity_store if activity_store is not None else load_activity_store()
    for item in store.get("activities", []):
        if item.get("status") == "active":
            return item
    return None


def find_completion_activity(activity_store, activity_id):
    for item in activity_store.get("activities", []):
        if item.get("id") == activity_id:
            return item
    return None


def update_activity_fields(item, payload):
    if "eyebrow" in payload:
        item["eyebrow"] = str(payload.get("eyebrow") or "").strip()[:40] or "完课活动"
    if "title" in payload:
        title = str(payload.get("title") or "").strip()
        if not title:
            raise ValueError("请输入活动标题。")
        item["title"] = title[:80]
    if "description" in payload:
        item["description"] = str(payload.get("description") or "").strip()[:500]
    visuals_payload = payload.get("visuals") if isinstance(payload.get("visuals"), dict) else {}
    if "stage_labels" in payload:
        visuals_payload["stage_labels"] = payload.get("stage_labels")
    if "result_labels" in payload:
        visuals_payload["result_labels"] = payload.get("result_labels")
    if visuals_payload:
        visuals = normalize_activity_visuals(item.get("visuals"))
        if "stage_labels" in visuals_payload:
            visuals["stage_labels"] = split_activity_labels(
                visuals_payload.get("stage_labels"),
                DEFAULT_ACTIVITY_STAGE_LABELS,
                ACTIVITY_STAGE_LIMIT,
            )
        if "result_labels" in visuals_payload:
            visuals["result_labels"] = split_activity_labels(
                visuals_payload.get("result_labels"),
                DEFAULT_ACTIVITY_RESULT_LABELS,
                ACTIVITY_RESULT_LIMIT,
            )
        if "image_note" in visuals_payload:
            visuals["image_note"] = str(visuals_payload.get("image_note") or "").strip()[:120]
        if "image_footer" in visuals_payload:
            visuals["image_footer"] = str(visuals_payload.get("image_footer") or "").strip()[:160]
        item["visuals"] = visuals
    item["updated_at"] = now_iso()


def serialize_completion_activity(item, include_private=False):
    output = {
        "id": item.get("id", ""),
        "eyebrow": item.get("eyebrow") or DEFAULT_COMPLETION_ACTIVITY["eyebrow"],
        "title": item.get("title") or DEFAULT_COMPLETION_ACTIVITY["title"],
        "description": item.get("description") or DEFAULT_COMPLETION_ACTIVITY["description"],
        "status": normalize_activity_status(item.get("status")),
        "visuals": serialize_activity_visuals(item),
        "participant_class_ids": activity_participant_ids(item),
        "published_at": item.get("published_at", ""),
        "ended_at": item.get("ended_at", ""),
        "updated_at": item.get("updated_at", ""),
    }
    if include_private:
        output["created_at"] = item.get("created_at", "")
    return output


def completion_activity_payload(activity_store=None):
    store = activity_store if activity_store is not None else load_activity_store()
    active = active_completion_activity(store)
    payload = {
        "activity": serialize_completion_activity(active) if active else None,
        "can_manage_activity": can_manage_accounts(),
    }
    if can_manage_accounts():
        payload["activities"] = [
            serialize_completion_activity(item, include_private=True)
            for item in store.get("activities", [])
        ]
    return payload


def set_class_activity_participation(target_class, activity_store, enabled):
    active = active_completion_activity(activity_store)
    if enabled and not active:
        raise ValueError("当前没有进行中的完课活动，请管理员先发布活动。")
    if not active:
        target_class["completion_activity"] = False
        return

    participants = activity_participant_ids(active)
    class_id = target_class.get("id")
    if enabled:
        if class_id not in participants:
            participants.append(class_id)
        target_class["completion_activity"] = True
    else:
        participants = [item for item in participants if item != class_id]
        target_class["completion_activity"] = False
    active["participant_class_ids"] = participants
    active["updated_at"] = now_iso()


def sync_active_activity_to_classes(class_store, activity_store):
    active = active_completion_activity(activity_store)
    active_ids = set(activity_participant_ids(active)) if active else set()
    for item in class_store.get("classes", []):
        item["completion_activity"] = bool(active and item.get("id") in active_ids)


def publish_completion_activity(class_store, activity_store, activity):
    timestamp = now_iso()
    was_ended = activity.get("status") == "ended"
    for item in activity_store.get("activities", []):
        if item.get("status") == "active" and item.get("id") != activity.get("id"):
            item["status"] = "ended"
            item["ended_at"] = timestamp
            item["updated_at"] = timestamp
    activity["status"] = "active"
    activity["published_at"] = timestamp
    activity["ended_at"] = ""
    activity["updated_at"] = timestamp
    if was_ended:
        activity["participant_class_ids"] = []
    sync_active_activity_to_classes(class_store, activity_store)


def end_completion_activity(class_store, activity):
    timestamp = now_iso()
    activity["status"] = "ended"
    activity["ended_at"] = timestamp
    activity["updated_at"] = timestamp
    for item in class_store.get("classes", []):
        item["completion_activity"] = False


def create_completion_activity(activity_store, payload):
    source_id = str(payload.get("source_id") or "").strip()
    source = find_completion_activity(activity_store, source_id) if source_id else None
    fields = {}
    for key in ("eyebrow", "title", "description"):
        value = payload.get(key) if key in payload else (source or {}).get(key)
        if value is not None:
            fields[key] = value
    source_visuals = normalize_activity_visuals((source or {}).get("visuals"))
    if source:
        fields["visuals"] = source_visuals
    if isinstance(payload.get("visuals"), dict):
        fields["visuals"] = payload.get("visuals")
    activity = build_completion_activity(fields=fields, status="draft", participant_ids=[])
    activity_store.setdefault("activities", []).insert(0, activity)
    return activity


def activity_image_extension(filename):
    text = str(filename or "").strip().lower()
    match = re.search(r"(\.[a-z0-9]+)$", text)
    if not match:
        return ""
    extension = match.group(1)
    return extension if extension in ACTIVITY_IMAGE_EXTENSIONS else ""


def current_owner():
    return g.user["username"]


def current_teacher_id():
    return (
        normalize_teacher_id(g.user.get("teacher_id"))
        or normalize_teacher_id(g.user.get("username"))
        or teacher_id_for_username(g.user.get("username"))
    )


def class_teacher_id(item):
    return (
        normalize_teacher_id(item.get("teacher_id"))
        or teacher_id_for_username(item.get("owner"))
        or infer_teacher_id_from_class_name(item.get("name"))
    )


def can_read_class(item):
    return item.get("owner") == current_owner() or can_manage_accounts()


def blank_week():
    return [None for _ in range(DAY_COUNT)]


def normalize_week_values(values):
    output = blank_week()
    if not isinstance(values, list):
        return output
    for index in range(min(DAY_COUNT, len(values))):
        output[index] = parse_completion(values[index])
    return output


def normalize_weeks(weeks):
    output = {str(week): blank_week() for week in range(1, WEEK_COUNT + 1)}
    if not isinstance(weeks, dict):
        return output
    for week in range(1, WEEK_COUNT + 1):
        output[str(week)] = normalize_week_values(weeks.get(str(week)))
    return output


def normalize_reminder_class_name(value):
    return re.sub(r"\s+", "", str(value or "").lower())


def alpha_zero_based_number(char):
    text = str(char or "").strip()
    if not text:
        return ""
    index = ord(text.upper()) - ord("A")
    return str(index) if index >= 0 else text


def alpha_one_based_number(char):
    text = str(char or "").strip()
    if not text:
        return ""
    index = ord(text.upper()) - ord("A") + 1
    return str(index) if index > 0 else text


def encode_period_token(token):
    return "".join(alpha_zero_based_number(char) if char.isalpha() else char for char in str(token or ""))


def encode_reminder_class_name(class_name):
    value = str(class_name or "")

    def replace_period(match):
        token = match.group(1)
        if not any(char.isalpha() for char in token):
            return match.group(0)
        return f"{encode_period_token(token)}期"

    value = re.sub(r"([A-Za-z0-9]+)期", replace_period, value)
    value = re.sub(r"-([A-Za-z])班", lambda match: f"-{alpha_one_based_number(match.group(1))}班", value)
    return value


def reminder_class_match_keys(class_name):
    raw_value = str(class_name or "").strip()
    variants = {raw_value, encode_reminder_class_name(raw_value)}
    for value in list(variants):
        variants.add(re.sub(r"L2-", "PU1-", value, flags=re.IGNORECASE))
    return {normalize_reminder_class_name(value) for value in variants if normalize_reminder_class_name(value)}


def reminder_class_names_match(first_name, second_name):
    first_keys = reminder_class_match_keys(first_name)
    second_keys = reminder_class_match_keys(second_name)
    return bool(first_keys and first_keys & second_keys)


def get_student_account(student):
    return str(student.get("account") or student.get("phone") or "").strip()


def get_student_weeks(student, month_key):
    months = student.get("months")
    if isinstance(months, dict):
        month_data = months.get(month_key, {})
        return normalize_weeks(month_data.get("weeks", {}))
    return normalize_weeks(student.get("weeks", {}))


def ensure_month_data(student, month_key):
    months = student.setdefault("months", {})
    month_data = months.setdefault(month_key, {})
    month_data["weeks"] = normalize_weeks(month_data.get("weeks", {}))
    return month_data


def calculate_monthly_completion(weeks):
    values = []
    for week_values in weeks.values():
        values.extend(value for value in week_values if value is not None)
    if not values:
        return None
    return round(sum(values) / len(values), 2)


def uploaded_day_values(weeks):
    values = []
    for week_key in sorted(weeks, key=lambda value: int(value)):
        week_values = weeks[week_key]
        for day_index, value in enumerate(week_values, start=1):
            if value is not None:
                values.append({"week": int(week_key), "day": day_index, "value": value})
    return values


def has_abnormal_break(values):
    zero_streak = 0
    zero_streak_start = None

    for index, item in enumerate(values):
        if item["value"] <= 0:
            if zero_streak == 0:
                zero_streak_start = index
            zero_streak += 1
            if zero_streak >= 2:
                previous_values = values[:zero_streak_start]
                if previous_values and all(previous["value"] >= 100 for previous in previous_values):
                    return True
        else:
            zero_streak = 0
            zero_streak_start = None
    return False


def classify_student_habit(weeks):
    values = uploaded_day_values(weeks)
    if not values:
        return "暂无数据"

    if all(item["value"] == 0 for item in values):
        return "长期不上课"

    if all(item["value"] >= 100 for item in values):
        return "完课超赞"

    if has_abnormal_break(values):
        return "异常断课"

    incomplete_values = [item for item in values if item["value"] < 100]
    if incomplete_values and all(item["day"] == DAY_COUNT for item in incomplete_values):
        return "周末欠缺"

    average_completion = sum(item["value"] for item in values) / len(values)
    if average_completion <= 60:
        return "断续上课"

    full_count = len([item for item in values if item["value"] >= 100])
    if full_count > len(values) / 2:
        return "偶尔断课"

    return "断续上课"


def serialize_student(student, enrolled_ids=None):
    month_key = current_month_key()
    weeks = get_student_weeks(student, month_key)
    student_id = str(student.get("id") or "")
    enrolled_lookup = enrolled_ids or set()
    return {
        "id": student_id,
        "name": str(student.get("name") or "").strip(),
        "account": get_student_account(student),
        "month": month_key,
        "monthly_completion": calculate_monthly_completion(weeks),
        "habit_category": classify_student_habit(weeks),
        "renewal_enrolled": student_id in enrolled_lookup,
        "weeks": weeks,
        "updated_at": student.get("updated_at", ""),
    }


def serialize_class(item, include_students=False, active_activity=True):
    students = item.get("students", [])
    teacher_id = class_teacher_id(item)
    title_week_number = current_title_week_number(item)
    activity_enabled = bool(item.get("completion_activity")) and bool(active_activity)
    output = {
        "id": item["id"],
        "name": item["name"],
        "note": str(item.get("note") or "").strip(),
        "title_week_number": title_week_number,
        "title_week_label": f"W{title_week_number}" if title_week_number else "",
        "title_week_anchor": str(item.get("title_week_anchor") or ""),
        "teacher_id": teacher_id,
        "teacher_name": teacher_label(teacher_id),
        "owner": item.get("owner", ""),
        "can_edit": item.get("owner") == current_owner(),
        "student_count": len(students),
        "completion_activity": activity_enabled,
        "created_at": item.get("created_at", ""),
        "updated_at": item.get("updated_at", ""),
    }
    if include_students:
        enrolled_ids = renewal_enrolled_student_ids(item.get("id"))
        output["students"] = [serialize_student(student, enrolled_ids) for student in students]
        output["month"] = current_month_key()
    return output


def team_class_summary(store):
    teacher_lookup = {teacher["id"]: {**teacher, "classes": [], "class_count": 0, "student_count": 0} for teacher in teacher_options()}
    unknown_groups = {}
    for item in store.get("classes", []):
        teacher_id = class_teacher_id(item)
        group = teacher_lookup.get(teacher_id)
        if group is None:
            owner_key = item.get("owner") or "unknown"
            group = unknown_groups.setdefault(owner_key, {
                "id": owner_key,
                "name": teacher_label(teacher_id) or f"未识别账号 {owner_key}",
                "classes": [],
                "class_count": 0,
                "student_count": 0,
            })
        student_count = len(item.get("students", []))
        group["classes"].append({
            "id": item.get("id", ""),
            "name": item.get("name", ""),
            "note": str(item.get("note") or "").strip(),
            "student_count": student_count,
            "owner": item.get("owner", ""),
        })
        group["class_count"] += 1
        group["student_count"] += student_count

    groups = [
        group
        for group in [*teacher_lookup.values(), *unknown_groups.values()]
        if group["class_count"] > 0
    ]
    for group in groups:
        group["classes"] = sorted(group["classes"], key=lambda item: item.get("name", ""))
    return {
        "teacher_count": len([group for group in groups if group["class_count"] > 0]),
        "class_count": sum(group["class_count"] for group in groups),
        "student_count": sum(group["student_count"] for group in groups),
        "groups": groups,
    }


def find_owned_class(store, class_id):
    for item in store["classes"]:
        if item["id"] == class_id and item["owner"] == current_owner():
            return item
    return None


def find_readable_class(store, class_id):
    for item in store["classes"]:
        if item["id"] == class_id and can_read_class(item):
            return item
    return None


def find_student(target_class, student_id):
    for student in target_class.get("students", []):
        if student.get("id") == student_id:
            return student
    return None


def normalize_header(value):
    text = str(value or "").replace("\u3000", " ").strip().lower()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[：:（）()\[\]【】{}<>《》\"“”'‘’、,，.。/\\|_\-]+", "", text)
    text = text.replace("必填", "").replace("选填", "")
    return text


def pick_column(headers, candidates):
    normalized = [normalize_header(header) for header in headers]
    candidate_values = [normalize_header(candidate) for candidate in candidates]

    lookup = {header: index for index, header in enumerate(normalized) if header}
    for candidate in candidate_values:
        index = lookup.get(candidate)
        if index is not None:
            return index

    for index, header in enumerate(normalized):
        if not header:
            continue
        for candidate in candidate_values:
            if candidate and (candidate in header or header in candidate):
                return index
    return None


def has_day_label(header, day):
    normalized_header = normalize_header(header)
    return any(normalize_header(candidate) in normalized_header for candidate in day_aliases(day))


def is_daily_completion_header(header):
    normalized_header = normalize_header(header)
    return any(
        keyword in normalized_header
        for keyword in (
            "当日完成度",
            "当日完成率",
            "当日完课度",
            "当日完课率",
            "当天完成度",
            "当天完成率",
            "当天完课度",
            "当天完课率",
            "本日完成度",
            "本日完成率",
            "本日完课度",
            "本日完课率",
        )
    )


def day_aliases(day):
    chinese = CHINESE_NUMBERS[day]
    return {
        f"day{day}",
        f"d{day}",
        f"{day}天",
        f"{day}日",
        f"第{day}天",
        f"第{day}日",
        f"{day}天当日完成度",
        f"{day}日当日完成度",
        f"第{day}天当日完成度",
        f"第{day}日当日完成度",
        f"第{day}天完成度",
        f"第{day}日完成度",
        f"{chinese}天",
        f"{chinese}日",
        f"第{chinese}天",
        f"第{chinese}日",
        f"{chinese}天当日完成度",
        f"{chinese}日当日完成度",
        f"第{chinese}天当日完成度",
        f"第{chinese}日当日完成度",
        f"第{chinese}天完成度",
        f"第{chinese}日完成度",
    }


def row_texts(row, width):
    return [str(row[index] or "").strip() if index < len(row) else "" for index in range(width)]


def fill_merged_parent_cells(values):
    filled = []
    current = ""
    for value in values:
        if value:
            current = value
        filled.append(current)
    return filled


def build_headers(rows, row_index):
    start_index = max(0, row_index - 2)
    context_rows = rows[start_index:row_index + 1]
    width = max((len(row) for row in context_rows), default=0)
    prepared_rows = []

    for offset, row in enumerate(context_rows):
        values = row_texts(row, width)
        if start_index + offset < row_index:
            values = fill_merged_parent_cells(values)
        prepared_rows.append(values)

    headers = []
    for column_index in range(width):
        parts = []
        for values in prepared_rows:
            value = values[column_index]
            if value and (not parts or parts[-1] != value):
                parts.append(value)
        headers.append(" ".join(parts))
    return headers


def pick_day_columns(headers):
    normalized = [normalize_header(header) for header in headers]
    day_columns = {}
    used_indexes = set()
    has_daily_completion_row = any(is_daily_completion_header(header) for header in headers)

    for day in range(1, DAY_COUNT + 1):
        candidates = [normalize_header(candidate) for candidate in day_aliases(day)]
        for index, header in enumerate(headers):
            if index in used_indexes:
                continue
            if has_day_label(header, day) and is_daily_completion_header(header):
                day_columns[day] = index
                used_indexes.add(index)
                break
        if day in day_columns:
            continue
        if has_daily_completion_row:
            continue
        for index, header in enumerate(normalized):
            if index in used_indexes or not header:
                continue
            if header in candidates:
                day_columns[day] = index
                used_indexes.add(index)
                break
        if day in day_columns:
            continue
        for index, header in enumerate(normalized):
            if index in used_indexes or not header:
                continue
            if any(candidate and candidate in header for candidate in candidates):
                day_columns[day] = index
                used_indexes.add(index)
                break
    return day_columns


def non_empty_cells(row):
    return [str(value).strip() for value in row if str(value or "").strip()]


def readable_headers(rows):
    candidates = []
    for row in rows[:HEADER_SCAN_LIMIT]:
        cells = non_empty_cells(row)
        if cells:
            candidates.append("、".join(cells[:10]))
    return "；".join(candidates[:5])


def find_header_row(rows):
    best = None
    for row_index, row in enumerate(rows[:HEADER_SCAN_LIMIT]):
        headers = build_headers(rows, row_index)
        if not any(headers):
            continue
        day_columns = pick_day_columns(headers)
        has_daily_completion_row = any(is_daily_completion_header(header) for header in headers)
        indexes = {
            "name": pick_column(headers, NAME_COLUMNS),
            "account": pick_column(headers, ACCOUNT_COLUMNS),
            "days": day_columns,
        }
        score = 0
        score += 5 if indexes["name"] is not None else 0
        score += 5 if indexes["account"] is not None else 0
        score += len(day_columns) * 2
        score += 20 if has_daily_completion_row and day_columns else 0
        if best is None or score > best["score"]:
            best = {"score": score, "row_index": row_index, "headers": headers, "indexes": indexes}
    return best


def missing_column_error(column_name, examples, rows):
    preview = readable_headers(rows)
    message = f"表格需要包含“{column_name}”列，常见列名：{examples}。"
    if preview:
        message += f" 已读取到的前几行内容：{preview}"
    return message


def parse_completion(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        rate = float(value)
        if 0 <= rate <= 1:
            rate *= 100
        return max(0, min(100, round(rate, 2)))

    text = str(value).strip().replace("％", "%")
    if not text:
        return None

    normalized_text = normalize_header(text)
    if normalized_text in {"完成", "已完成", "是", "yes", "y", "true", "√", "对"}:
        return 100
    if normalized_text in {"未完成", "没完成", "否", "no", "n", "false", "x", "×", "错", "缺课"}:
        return 0

    if text.endswith("%"):
        text = text[:-1]
    try:
        rate = float(text)
    except ValueError:
        return None
    if 0 <= rate <= 1:
        rate *= 100
    return max(0, min(100, round(rate, 2)))


def row_value(row, index):
    if index is None or index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else str(value).strip()


def rows_to_students(rows):
    if not rows:
        return []

    header = find_header_row(rows)
    indexes = header["indexes"] if header else {}
    header_row_index = header["row_index"] if header else 0
    name_index = indexes.get("name")
    account_index = indexes.get("account")
    day_columns = indexes.get("days", {})

    if name_index is None:
        raise ValueError(missing_column_error("学员姓名", "学员姓名、学生姓名、孩子姓名、姓名", rows))
    if account_index is None:
        raise ValueError(missing_column_error("学员账号", "学员账号、学生账号、账号、学号、ID", rows))
    if not day_columns:
        raise ValueError(missing_column_error("Day1-Day6", "第一天当日完成度、第二天当日完成度、Day1、Day2", rows))

    students = []
    for row in rows[header_row_index + 1:]:
        name = row_value(row, name_index)
        account = row_value(row, account_index)
        if not name:
            continue
        days = {}
        for day, column_index in day_columns.items():
            value = row[column_index] if column_index < len(row) else None
            days[str(day)] = parse_completion(value)
        students.append(
            {
                "name": name,
                "account": account,
                "days": days,
                "updated_at": now_iso(),
            }
        )

    if not students:
        raise ValueError("没有读取到学员数据，请确认表头下面存在学员记录。")
    return students


def parse_csv(file_storage):
    raw = file_storage.stream.read()
    for encoding in ("utf-8-sig", "gb18030", "gbk"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("CSV 编码无法识别，请另存为 UTF-8 或 Excel .xlsx 后再上传。")
    return list(csv.reader(io.StringIO(text)))


def parse_xlsx_sheets(file_storage):
    workbook = load_workbook(file_storage, data_only=True)
    for sheet in workbook.worksheets:
        rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        yield sheet.title, rows


def parse_upload(file_storage):
    filename = (file_storage.filename or "").lower()
    if filename.endswith(".csv"):
        rows = parse_csv(file_storage)
        return rows_to_students(rows)
    if filename.endswith(".xlsx"):
        errors = []
        for sheet_name, rows in parse_xlsx_sheets(file_storage):
            try:
                return rows_to_students(rows)
            except ValueError as error:
                errors.append(f"{sheet_name}：{error}")
        detail = errors[0] if errors else "没有读取到工作表内容。"
        raise ValueError(f"未在 Excel 工作表中识别到可导入数据。{detail}")
    raise ValueError("仅支持 .xlsx 或 .csv 文件。")


def parse_week_number(value):
    try:
        week = int(value)
    except (TypeError, ValueError):
        raise ValueError("请选择本月第几周。") from None
    if week < 1 or week > WEEK_COUNT:
        raise ValueError("周数只能选择第一周到第四周。")
    return str(week)


def normalize_identity(value):
    return str(value or "").strip().lower()


def sync_students_from_upload(target_class, imported_students, week_number):
    students = target_class.setdefault("students", [])
    existing_by_account = {
        normalize_identity(get_student_account(student)): student
        for student in students
        if normalize_identity(get_student_account(student))
    }
    existing_by_name = {
        normalize_identity(student.get("name")): student
        for student in students
        if normalize_identity(student.get("name"))
    }

    month_key = current_month_key()
    updated_at = now_iso()
    updated = 0
    created = 0
    matched_student_ids = set()
    synced_students = []

    for imported in imported_students:
        account_key = normalize_identity(imported.get("account"))
        name_key = normalize_identity(imported.get("name"))
        current = None
        if account_key:
            current = existing_by_account.get(account_key)
        if current is None and name_key:
            current = existing_by_name.get(name_key)
        if current is not None and current.get("id") in matched_student_ids:
            current = None

        if current:
            updated += 1
        else:
            current = {
                "id": uuid.uuid4().hex,
                "name": "",
                "account": "",
                "months": {},
                "created_at": updated_at,
            }
            created += 1

        current["name"] = imported.get("name", current.get("name", "")).strip()
        current["account"] = imported.get("account", current.get("account", "")).strip()
        current["updated_at"] = updated_at

        if imported.get("days"):
            month_data = ensure_month_data(current, month_key)
            week_values = blank_week()
            for day, value in imported["days"].items():
                day_index = int(day) - 1
                if 0 <= day_index < DAY_COUNT:
                    week_values[day_index] = value
            month_data["weeks"][week_number] = week_values

        if normalize_identity(current.get("account")):
            existing_by_account[normalize_identity(current.get("account"))] = current
        if normalize_identity(current.get("name")):
            existing_by_name[normalize_identity(current.get("name"))] = current
        matched_student_ids.add(current["id"])
        synced_students.append(current)

    removed = len([student for student in students if student.get("id") not in matched_student_ids])
    target_class["students"] = synced_students
    target_class["updated_at"] = updated_at
    renewal_removed = prune_renewal_followups_for_class(target_class.get("id"), matched_student_ids)
    return {
        "created": created,
        "updated": updated,
        "removed": removed,
        "renewal_removed": renewal_removed,
        "week": week_number,
    }


def clear_current_month_data(target_class):
    month_key = current_month_key()
    cleared = 0
    for student in target_class.get("students", []):
        had_data = False
        months = student.get("months")
        if isinstance(months, dict) and month_key in months:
            months.pop(month_key, None)
            had_data = True
        if "weeks" in student:
            student.pop("weeks", None)
            had_data = True
        if "completion_rate" in student:
            student.pop("completion_rate", None)
            had_data = True
        if had_data:
            student["updated_at"] = now_iso()
            cleared += 1
    target_class["updated_at"] = now_iso()
    return {"cleared": cleared, "month": month_key}


def clear_current_week_data(target_class, week_number):
    month_key = current_month_key()
    week_key = str(week_number)
    cleared = 0
    for student in target_class.get("students", []):
        had_data = False
        months = student.get("months")
        if isinstance(months, dict):
            month_data = months.get(month_key)
            if isinstance(month_data, dict):
                weeks = normalize_weeks(month_data.get("weeks", {}))
                if any(value is not None for value in weeks.get(week_key, [])):
                    had_data = True
                weeks[week_key] = blank_week()
                month_data["weeks"] = weeks
        if isinstance(student.get("weeks"), dict):
            weeks = normalize_weeks(student.get("weeks", {}))
            if any(value is not None for value in weeks.get(week_key, [])):
                had_data = True
            weeks[week_key] = blank_week()
            student["weeks"] = weeks
        if had_data:
            student.pop("completion_rate", None)
            student["updated_at"] = now_iso()
            cleared += 1
    target_class["updated_at"] = now_iso()
    return {"cleared": cleared, "month": month_key, "week": week_number}


@classes_bp.get("")
@login_required
def list_classes():
    store = load_store()
    activity_store = load_activity_store(store)
    active_activity = active_completion_activity(activity_store)
    classes = [
        serialize_class(item, active_activity=active_activity)
        for item in store["classes"]
        if item["owner"] == current_owner()
    ]
    payload = {
        "classes": classes,
        "teachers": teacher_options(),
        **completion_activity_payload(activity_store),
    }
    if can_manage_accounts():
        payload["team_summary"] = team_class_summary(store)
    return jsonify(payload)


@classes_bp.post("")
@login_required
def create_class():
    payload = request.get_json(silent=True) or {}
    name = str(payload.get("name", "")).strip()
    if not name:
        return jsonify({"error": "请输入班级名称。"}), 400
    teacher_id = current_teacher_id()

    store = load_store()
    item = {
        "id": uuid.uuid4().hex,
        "owner": current_owner(),
        "name": name,
        "note": "",
        "teacher_id": teacher_id,
        "students": [],
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    store["classes"].append(item)
    save_store(store)
    activity_store = load_activity_store(store)
    active_activity = active_completion_activity(activity_store)
    return jsonify({
        "class": serialize_class(item, include_students=True, active_activity=active_activity),
        **completion_activity_payload(activity_store),
    }), 201


@classes_bp.post("/activities")
@login_required
def create_activity():
    if not can_manage_accounts():
        return jsonify({"error": "只有管理员可以管理完课活动。"}), 403
    payload = request.get_json(silent=True) or {}
    class_store = load_store()
    activity_store = load_activity_store(class_store)
    try:
        activity = create_completion_activity(activity_store, payload)
        action = str(payload.get("action") or "").strip().lower()
        status = normalize_activity_status(payload.get("status"))
        if action == "publish" or status == "active":
            publish_completion_activity(class_store, activity_store, activity)
            save_store(class_store)
        save_activity_store(activity_store)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400
    return jsonify(completion_activity_payload(activity_store)), 201


@classes_bp.patch("/activities/<activity_id>")
@login_required
def update_activity(activity_id):
    if not can_manage_accounts():
        return jsonify({"error": "只有管理员可以管理完课活动。"}), 403
    payload = request.get_json(silent=True) or {}
    class_store = load_store()
    activity_store = load_activity_store(class_store)
    activity = find_completion_activity(activity_store, activity_id)
    if activity is None:
        return jsonify({"error": "活动不存在。"}), 404
    action = str(payload.get("action") or "").strip().lower()
    status = normalize_activity_status(activity.get("status"))
    if status == "active" and action != "end":
        return jsonify({"error": "进行中的活动不能编辑，请复制为新草稿，或结束后再发起新活动。"}), 400
    if status == "ended":
        return jsonify({"error": "已存档活动不能直接修改，请复制为新草稿。"}), 400
    try:
        if action == "end":
            if status != "active":
                return jsonify({"error": "只有进行中的活动可以结束。"}), 400
        else:
            update_activity_fields(activity, payload)
        if action == "publish":
            publish_completion_activity(class_store, activity_store, activity)
            save_store(class_store)
        elif action == "end":
            end_completion_activity(class_store, activity)
            save_store(class_store)
        save_activity_store(activity_store)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400
    return jsonify(completion_activity_payload(activity_store))


@classes_bp.post("/activities/<activity_id>/visuals")
@login_required
def upload_activity_visual(activity_id):
    if not can_manage_accounts():
        return jsonify({"error": "只有管理员可以上传活动图片。"}), 403
    class_store = load_store()
    activity_store = load_activity_store(class_store)
    activity = find_completion_activity(activity_store, activity_id)
    if activity is None:
        return jsonify({"error": "活动不存在。"}), 404
    if normalize_activity_status(activity.get("status")) != "draft":
        return jsonify({"error": "只有草稿活动可以上传或修改图片素材。"}), 400

    slot_type = str(request.form.get("slot_type") or "").strip()
    try:
        slot_index = int(request.form.get("slot_index"))
    except (TypeError, ValueError):
        return jsonify({"error": "图片位置无效。"}), 400
    if slot_type not in {"stage", "result"}:
        return jsonify({"error": "图片类型无效。"}), 400
    visuals = normalize_activity_visuals(activity.get("visuals"))
    visuals["stage_labels"] = split_activity_labels(
        request.form.get("stage_labels"),
        visuals["stage_labels"],
        ACTIVITY_STAGE_LIMIT,
    )
    visuals["result_labels"] = split_activity_labels(
        request.form.get("result_labels"),
        visuals["result_labels"],
        ACTIVITY_RESULT_LIMIT,
    )
    visuals["image_note"] = str(request.form.get("image_note") or visuals.get("image_note") or "").strip()[:120]
    visuals["image_footer"] = str(request.form.get("image_footer") or visuals.get("image_footer") or "").strip()[:160]
    limit = len(visuals["stage_labels"]) if slot_type == "stage" else len(visuals["result_labels"])
    if slot_index < 0 or slot_index >= limit:
        return jsonify({"error": "图片位置无效。"}), 400

    file_storage = request.files.get("file")
    if not file_storage or not file_storage.filename:
        return jsonify({"error": "请选择要上传的图片。"}), 400
    extension = activity_image_extension(file_storage.filename)
    if not extension:
        return jsonify({"error": "仅支持 png、jpg、jpeg、webp 图片。"}), 400

    asset_dir = completion_activity_asset_dir()
    asset_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{activity_id}_{slot_type}_{slot_index}_{uuid.uuid4().hex}{extension}"
    file_storage.save(asset_dir / filename)

    image_key = "stage_images" if slot_type == "stage" else "result_images"
    old_filename = visuals[image_key].get(str(slot_index), {}).get("filename")
    if old_filename and old_filename != filename:
        old_path = asset_dir / secure_filename(old_filename)
        if old_path.exists() and old_path.is_file():
            try:
                old_path.unlink()
            except OSError:
                pass
    visuals[image_key][str(slot_index)] = {
        "filename": filename,
        "updated_at": now_iso(),
    }
    activity["visuals"] = visuals
    activity["updated_at"] = now_iso()
    save_activity_store(activity_store)
    return jsonify(completion_activity_payload(activity_store))


@classes_bp.get("/activity-assets/<path:filename>")
@login_required
def activity_asset(filename):
    return send_from_directory(completion_activity_asset_dir(), filename)


@classes_bp.get("/reminder-match")
@login_required
def reminder_match_class():
    class_id = str(request.args.get("class_id") or "").strip()
    class_name = str(request.args.get("class_name") or "").strip()
    teacher_id = normalize_teacher_id(request.args.get("teacher_id"))

    store = load_store()
    activity_store = load_activity_store(store)
    active_activity = active_completion_activity(activity_store)
    candidates = [item for item in store["classes"] if can_read_class(item)]

    if class_id:
        direct_match = next((item for item in candidates if item.get("id") == class_id), None)
        if direct_match:
            return jsonify({"class": serialize_class(direct_match, include_students=True, active_activity=active_activity)})

    matches = [
        item
        for item in candidates
        if class_name and reminder_class_names_match(item.get("name"), class_name)
    ]
    if teacher_id:
        teacher_matches = [item for item in matches if class_teacher_id(item) == teacher_id]
        if teacher_matches:
            matches = teacher_matches

    if not matches:
        return jsonify({"error": "未匹配到该催课班级的学员明细。"}), 404

    matches = sorted(
        matches,
        key=lambda item: (
            0 if teacher_id and class_teacher_id(item) == teacher_id else 1,
            item.get("name", ""),
        ),
    )
    return jsonify({"class": serialize_class(matches[0], include_students=True, active_activity=active_activity)})


@classes_bp.patch("/<class_id>")
@login_required
def update_class(class_id):
    payload = request.get_json(silent=True) or {}
    store = load_store()
    activity_store = load_activity_store(store)
    active_activity = active_completion_activity(activity_store)
    item = find_owned_class(store, class_id)
    if item is None:
        return jsonify({"error": "班级不存在。"}), 404

    if "name" in payload:
        name = str(payload.get("name") or "").strip()
        if not name:
            return jsonify({"error": "请输入班级名称。"}), 400
        item["name"] = name
    if "note" in payload:
        item["note"] = str(payload.get("note") or "").strip()[:300]
    if "title_week_number" in payload:
        try:
            title_week_number = parse_title_week_number(payload.get("title_week_number"))
        except ValueError as error:
            return jsonify({"error": str(error)}), 400
        if title_week_number:
            item["title_week_base_number"] = title_week_number
            item["title_week_anchor"] = local_date_key(export_cycle_anchor())
        else:
            item.pop("title_week_base_number", None)
            item.pop("title_week_anchor", None)
            item.pop("title_week_number", None)
    if "teacher_id" in payload:
        item["teacher_id"] = normalize_teacher_id(payload.get("teacher_id"))
    if "completion_activity" in payload:
        try:
            set_class_activity_participation(item, activity_store, bool(payload.get("completion_activity")))
            active_activity = active_completion_activity(activity_store)
        except ValueError as error:
            return jsonify({"error": str(error)}), 400

    item["updated_at"] = now_iso()
    save_store(store)
    save_activity_store(activity_store)
    return jsonify({
        "class": serialize_class(item, include_students=True, active_activity=active_activity),
        **completion_activity_payload(activity_store),
    })


@classes_bp.get("/<class_id>")
@login_required
def get_class(class_id):
    store = load_store()
    activity_store = load_activity_store(store)
    active_activity = active_completion_activity(activity_store)
    item = find_readable_class(store, class_id)
    if item is None:
        return jsonify({"error": "班级不存在。"}), 404
    return jsonify({
        "class": serialize_class(item, include_students=True, active_activity=active_activity),
        **completion_activity_payload(activity_store),
    })


@classes_bp.delete("/<class_id>")
@login_required
def delete_class(class_id):
    store = load_store()
    before = len(store["classes"])
    store["classes"] = [
        item
        for item in store["classes"]
        if not (item["id"] == class_id and item["owner"] == current_owner())
    ]
    if len(store["classes"]) == before:
        return jsonify({"error": "班级不存在。"}), 404
    save_store(store)
    return jsonify({"ok": True})


@classes_bp.delete("/<class_id>/month-data")
@login_required
def clear_month_data(class_id):
    store = load_store()
    activity_store = load_activity_store(store)
    active_activity = active_completion_activity(activity_store)
    item = find_owned_class(store, class_id)
    if item is None:
        return jsonify({"error": "班级不存在。"}), 404

    result = clear_current_month_data(item)
    save_store(store)
    return jsonify({
        "result": result,
        "class": serialize_class(item, include_students=True, active_activity=active_activity),
        **completion_activity_payload(activity_store),
    })


@classes_bp.delete("/<class_id>/week-data")
@login_required
def clear_week_data(class_id):
    try:
        week_number = parse_week_number(request.args.get("week"))
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    store = load_store()
    activity_store = load_activity_store(store)
    active_activity = active_completion_activity(activity_store)
    item = find_owned_class(store, class_id)
    if item is None:
        return jsonify({"error": "班级不存在。"}), 404

    result = clear_current_week_data(item, week_number)
    save_store(store)
    return jsonify({
        "result": result,
        "class": serialize_class(item, include_students=True, active_activity=active_activity),
        **completion_activity_payload(activity_store),
    })


@classes_bp.patch("/<class_id>/students/<student_id>")
@login_required
def update_student(class_id, student_id):
    payload = request.get_json(silent=True) or {}
    name = str(payload.get("name") or "").strip()
    if not name:
        return jsonify({"error": "请输入学员姓名。"}), 400

    store = load_store()
    activity_store = load_activity_store(store)
    active_activity = active_completion_activity(activity_store)
    item = find_owned_class(store, class_id)
    if item is None:
        return jsonify({"error": "班级不存在。"}), 404

    student = find_student(item, student_id)
    if student is None:
        return jsonify({"error": "学员不存在。"}), 404

    updated_at = now_iso()
    student["name"] = name
    student["updated_at"] = updated_at
    item["updated_at"] = updated_at
    save_store(store)
    return jsonify({
        "student": serialize_student(student),
        "class": serialize_class(item, include_students=True, active_activity=active_activity),
        **completion_activity_payload(activity_store),
    })


@classes_bp.post("/<class_id>/upload")
@login_required
def upload_students(class_id):
    file_storage = request.files.get("file")
    if not file_storage:
        return jsonify({"error": "请上传 Excel 或 CSV 文件。"}), 400

    try:
        week_number = parse_week_number(request.form.get("week"))
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    store = load_store()
    activity_store = load_activity_store(store)
    active_activity = active_completion_activity(activity_store)
    item = find_owned_class(store, class_id)
    if item is None:
        return jsonify({"error": "班级不存在。"}), 404

    try:
        imported_students = parse_upload(file_storage)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    result = sync_students_from_upload(item, imported_students, week_number)
    save_store(store)
    return jsonify({
        "result": result,
        "class": serialize_class(item, include_students=True, active_activity=active_activity),
        **completion_activity_payload(activity_store),
    })
