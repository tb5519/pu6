import csv
import io
import json
import re
import uuid
from datetime import datetime

from openpyxl import load_workbook

from flask import Blueprint, current_app, g, jsonify, request

from app.auth import can_manage_accounts, login_required
from app.classes import (
    calculate_monthly_completion,
    class_teacher_id,
    current_month_key,
    get_student_weeks,
    load_store as load_class_store,
    save_store as save_class_store,
)
from app.teachers import TEACHERS, normalize_match_text, normalize_teacher_id, teacher_id_for_username, teacher_label


renewal_bp = Blueprint("renewal", __name__, url_prefix="/api/renewal")

RENEWAL_STAGES = ["铺垫阶段", "续报首月", "续报次月", "结营续报"]
FOLLOWUP_STATUSES = ["愿意继续学", "需要考虑", "拒绝", "未接听"]
RENEWAL_FIRST_MONTH_STAGE = "续报首月"
RENEWAL_SECOND_MONTH_STAGE = "续报次月"
RENEWAL_FOUR_WEEK_STAGES = {RENEWAL_FIRST_MONTH_STAGE, RENEWAL_SECOND_MONTH_STAGE}
RENEWAL_CLOSING_STAGE = "结营续报"
RENEWAL_SINGLE_FOLLOWUP_STAGES = {RENEWAL_CLOSING_STAGE}
RENEWAL_PRIORITY_STAGES = {*RENEWAL_FOUR_WEEK_STAGES, *RENEWAL_SINGLE_FOLLOWUP_STAGES}
FOLLOWUP_METHODS = ["私信", "电话"]
LEADER_ACTION_TYPES = ["留言", "去电", "跟进"]
BLOCKER_OPTIONS = ["升初中", "时间紧张", "经济", "学员问题", "线下", "效果不满意", "不知道顾虑", "不回复", "拒绝早报"]
RENEWAL_WEEK_COUNT = 4
FOLLOWUP_STATUS_PRIORITY = {
    "愿意继续学": 0,
    "需要考虑": 1,
    "未接听": 2,
    "拒绝": 3,
    "": 4,
}


def renewal_file():
    return current_app.config["RENEWAL_PROJECTS_FILE"]


def now_iso():
    return datetime.now().isoformat(timespec="seconds")


def current_owner():
    return g.user["username"]


def current_teacher_id():
    return (
        normalize_teacher_id(g.user.get("teacher_id"))
        or normalize_teacher_id(g.user.get("username"))
        or teacher_id_for_username(g.user.get("username"))
    )


def load_store():
    path = renewal_file()
    if not path.exists():
        return {"projects": [], "blocker_options": []}
    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)
    if not isinstance(data, dict):
        return {"projects": [], "blocker_options": []}
    projects = data.get("projects")
    if not isinstance(projects, list):
        data["projects"] = []
    if not isinstance(data.get("blocker_options"), list):
        data["blocker_options"] = []
    return data


def save_store(store):
    path = renewal_file()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump(store, file, ensure_ascii=False, indent=2)


def normalize_stage(value):
    stage = str(value or "").strip()
    return stage if stage in RENEWAL_STAGES else RENEWAL_STAGES[0]


def normalize_followup_status(value):
    status = str(value or "").strip()
    return status if status in FOLLOWUP_STATUSES else ""


def normalize_leader_action_type(value):
    action_type = str(value or "").strip()
    return action_type if action_type in LEADER_ACTION_TYPES else LEADER_ACTION_TYPES[0]


def normalize_blocker_option(value):
    return str(value or "").strip()[:24]


def blocker_options(store=None):
    data = store if isinstance(store, dict) else load_store()
    options = []
    for option in [*BLOCKER_OPTIONS, *data.get("blocker_options", [])]:
        normalized = normalize_blocker_option(option)
        if normalized and normalized not in options:
            options.append(normalized)
    return options


def normalize_blocker(value, store=None):
    blocker = str(value or "").strip()
    return blocker if blocker in blocker_options(store) else ""


def normalize_followup_methods(value):
    if isinstance(value, list):
        raw_methods = value
    else:
        raw_methods = str(value or "").replace("、", ",").split(",")
    methods = []
    for item in raw_methods:
        method = str(item or "").strip()
        if method in FOLLOWUP_METHODS and method not in methods:
            methods.append(method)
    return methods


def normalize_followup_date(value):
    text = str(value or "").strip()[:10]
    if not text:
        return datetime.now().strftime("%Y-%m-%d")
    try:
        return datetime.strptime(text, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        return datetime.now().strftime("%Y-%m-%d")


def format_followup_date(value):
    text = str(value or "").strip()[:10]
    if not text:
        return ""
    try:
        parsed = datetime.strptime(text, "%Y-%m-%d")
    except ValueError:
        return text
    return f"{parsed.month}.{parsed.day}"


def normalize_weekly_followups(value):
    output = {str(week): [] for week in range(1, RENEWAL_WEEK_COUNT + 1)}
    if not isinstance(value, dict):
        return output
    for week in range(1, RENEWAL_WEEK_COUNT + 1):
        records = value.get(str(week), [])
        if not isinstance(records, list):
            continue
        collapsed = {}
        for record in records:
            if not isinstance(record, dict):
                continue
            methods = normalize_followup_methods(record.get("methods"))
            if not methods:
                continue
            date_key = normalize_followup_date(record.get("date"))
            current = collapsed.setdefault(date_key, {
                "date": date_key,
                "methods": [],
                "created_at": "",
            })
            for method in methods:
                if method not in current["methods"]:
                    current["methods"].append(method)
            current["created_at"] = str(record.get("created_at") or record.get("createdAt") or current["created_at"])
        output[str(week)] = [
            {
                "date": item.get("date", ""),
                "methods": methods,
                "created_at": item.get("created_at", ""),
            }
            for item in sorted(
                collapsed.values(),
                key=lambda value: (value.get("date", ""), value.get("created_at", "")),
            )
            for methods in [item.get("methods", [])]
        ]
    return output


def normalize_general_followups(value):
    if not isinstance(value, list):
        return []
    collapsed = {}
    for record in value:
        if not isinstance(record, dict):
            continue
        methods = normalize_followup_methods(record.get("methods"))
        if not methods:
            continue
        date_key = normalize_followup_date(record.get("date"))
        current = collapsed.setdefault(date_key, {
            "date": date_key,
            "methods": [],
            "created_at": "",
        })
        for method in methods:
            if method not in current["methods"]:
                current["methods"].append(method)
        current["created_at"] = str(record.get("created_at") or record.get("createdAt") or current["created_at"])
    return [
        {
            "date": item.get("date", ""),
            "methods": item.get("methods", []),
            "created_at": item.get("created_at", ""),
        }
        for item in sorted(
            collapsed.values(),
            key=lambda value: (value.get("date", ""), value.get("created_at", "")),
        )
    ]


def serialize_weekly_followups(record):
    weekly_followups = normalize_weekly_followups(record.get("weekly_followups"))
    output = {}
    for week in range(1, RENEWAL_WEEK_COUNT + 1):
        week_key = str(week)
        records = [
            {
                "date": item.get("date", ""),
                "date_label": format_followup_date(item.get("date")),
                "methods": item.get("methods", []),
                "created_at": format_followup_time(item.get("created_at")),
            }
            for item in sorted(
                weekly_followups.get(week_key, []),
                key=lambda value: (value.get("date", ""), value.get("created_at", "")),
            )
        ]
        latest = records[-1] if records else {}
        output[week_key] = {
            "latest_date": latest.get("date", ""),
            "latest_date_label": latest.get("date_label", ""),
            "latest_methods": latest.get("methods", []),
            "count": len(records),
            "records": records,
        }
    return output


def serialize_general_followups(record, include_weekly=False):
    source_records = normalize_general_followups(record.get("general_followups"))
    if include_weekly:
        for week_records in normalize_weekly_followups(record.get("weekly_followups")).values():
            source_records.extend(week_records)
        source_records = normalize_general_followups(source_records)
    records = [
        {
            "date": item.get("date", ""),
            "date_label": format_followup_date(item.get("date")),
            "methods": item.get("methods", []),
            "created_at": format_followup_time(item.get("created_at")),
        }
        for item in source_records
    ]
    latest = records[-1] if records else {}
    return {
        "latest_date": latest.get("date", ""),
        "latest_date_label": latest.get("date_label", ""),
        "latest_methods": latest.get("methods", []),
        "count": len(records),
        "records": records,
    }


def class_lookup():
    return {item.get("id"): item for item in load_class_store().get("classes", []) if item.get("id")}


def can_read_class(item):
    return can_manage_accounts() or item.get("owner") == current_owner()


def can_add_class(item):
    return item.get("owner") == current_owner()


def can_edit_project(project):
    return can_manage_accounts() or project.get("owner") == current_owner()


def can_read_project(project):
    return can_edit_project(project)


def source_class_student_count(source_class):
    return len(source_class.get("students", [])) if source_class else 0


def normalize_locked_student_count(value, fallback=0):
    try:
        count = int(value)
    except (TypeError, ValueError):
        count = int(fallback or 0)
    return max(0, min(count, 9999))


def ensure_project_student_count_lock(project, source_class):
    if "locked_student_count" in project:
        if "student_count_note" not in project:
            project["student_count_note"] = ""
            return True
        return False
    project["locked_student_count"] = source_class_student_count(source_class)
    project["student_count_note"] = str(project.get("student_count_note") or "").strip()[:300]
    return True


def project_student_count(project, source_class):
    return normalize_locked_student_count(
        project.get("locked_student_count"),
        fallback=source_class_student_count(source_class),
    )


def find_project(store, project_id):
    return next((item for item in store.get("projects", []) if item.get("id") == project_id), None)


def enrolled_student_ids(project, source_class=None):
    followups = project.get("student_followups", {})
    ids = {
        str(student_id)
        for student_id, record in followups.items()
        if bool((record or {}).get("enrolled"))
    }
    ids.update(
        str(value)
        for value in project.get("enrolled_student_ids", [])
        if str(value or "").strip() and str(value) not in followups
    )
    if source_class:
        valid_ids = {str(student.get("id")) for student in source_class.get("students", []) if student.get("id")}
        ids = ids & valid_ids
    return ids


def prune_project_followups(project, source_class):
    if not source_class:
        return False
    valid_ids = {
        str(student.get("id"))
        for student in source_class.get("students", [])
        if student.get("id")
    }
    changed = False

    followups = project.get("student_followups")
    if isinstance(followups, dict):
        for student_id in list(followups.keys()):
            if str(student_id) not in valid_ids:
                followups.pop(student_id, None)
                changed = True

    enrolled_ids = project.get("enrolled_student_ids")
    if isinstance(enrolled_ids, list):
        next_enrolled_ids = [
            student_id
            for student_id in enrolled_ids
            if str(student_id) in valid_ids
        ]
        if len(next_enrolled_ids) != len(enrolled_ids):
            project["enrolled_student_ids"] = next_enrolled_ids
            changed = True

    if changed:
        project["updated_at"] = now_iso()
    return changed


def prune_store_followups(store, classes_by_id):
    changed = False
    for project in store.get("projects", []):
        source_class = classes_by_id.get(project.get("class_id"))
        changed = prune_project_followups(project, source_class) or changed
    return changed


def student_followup_record(project, student_id):
    followups = project.setdefault("student_followups", {})
    student_key = str(student_id)
    record = followups.setdefault(student_key, {})
    if student_key in {str(value) for value in project.get("enrolled_student_ids", [])} and "enrolled" not in record:
        record["enrolled"] = True
    record["status"] = normalize_followup_status(record.get("status"))
    record["enrolled"] = bool(record.get("enrolled"))
    record["current_blocker"] = normalize_blocker(record.get("current_blocker"))
    record["weekly_followups"] = normalize_weekly_followups(record.get("weekly_followups"))
    record["general_followups"] = normalize_general_followups(record.get("general_followups"))
    record["notes"] = normalize_note_entries(record)
    record["note"] = note_history_text(record)
    record["leader_action_type"] = normalize_leader_action_type(record.get("leader_action_type"))
    record["leader_note"] = str(record.get("leader_note") or "").strip()[:500]
    record["leader_talk_keyword"] = str(record.get("leader_talk_keyword") or "").strip()[:120]
    record["leader_talk_type"] = str(record.get("leader_talk_type") or "").strip()[:40]
    record["leader_talk_title"] = str(record.get("leader_talk_title") or "").strip()[:180]
    record["leader_talk_text"] = str(record.get("leader_talk_text") or "").strip()[:5000]
    record["leader_note_done"] = bool(record.get("leader_note_done"))
    record["leader_note_updated_at"] = str(record.get("leader_note_updated_at") or "")
    record["leader_note_done_at"] = str(record.get("leader_note_done_at") or "")
    record["followed_at"] = str(record.get("followed_at") or "")
    return record


def format_followup_time(value):
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        return datetime.fromisoformat(text).strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return text


def normalize_note_entries(record):
    entries = []
    seen = set()
    raw_entries = record.get("notes")
    if isinstance(raw_entries, list):
        for item in raw_entries:
            if isinstance(item, dict):
                text = str(item.get("text") or "").strip()
                created_at = str(item.get("created_at") or "")
            else:
                text = str(item or "").strip()
                created_at = ""
            if not text:
                continue
            updated_at = str(item.get("updated_at") or "") if isinstance(item, dict) else ""
            key = (text, created_at, updated_at)
            if key in seen:
                continue
            seen.add(key)
            entries.append({
                "text": text[:500],
                "created_at": created_at,
                "updated_at": updated_at,
            })
    legacy_note = str(record.get("note") or "").strip()
    if legacy_note and not entries:
        for line in [item.strip() for item in legacy_note.replace("\r", "\n").split("\n")]:
            if not line:
                continue
            key = (line, "", "")
            if key in seen:
                continue
            seen.add(key)
            entries.append({
                "text": line[:500],
                "created_at": str(record.get("followed_at") or ""),
                "updated_at": "",
            })
    return entries[-30:]


def note_history_text(record):
    return "\n".join(item.get("text", "") for item in normalize_note_entries(record) if item.get("text"))


def append_followup_note(record, value):
    text = str(value or "").strip()
    if not text:
        return False
    entries = normalize_note_entries(record)
    entries.append({
        "text": text[:500],
        "created_at": now_iso(),
    })
    entries = entries[-30:]
    record["notes"] = entries
    record["note"] = "\n".join(item.get("text", "") for item in entries if item.get("text"))
    return True


def replace_followup_note(record, value):
    text = str(value or "").strip()
    current_text = note_history_text(record).strip()
    if text == current_text:
        return False
    if not text:
        return set_followup_note_entries(record, [])
    entries = normalize_note_entries(record)
    created_at = entries[0].get("created_at") if entries else now_iso()
    return set_followup_note_entries(record, [{
        "text": text[:500],
        "created_at": created_at or now_iso(),
        "updated_at": now_iso() if entries else "",
    }])


def set_followup_note_entries(record, entries):
    clean_entries = []
    for item in entries:
        text = str((item or {}).get("text") or "").strip()
        if not text:
            continue
        clean_entries.append({
            "text": text[:500],
            "created_at": str((item or {}).get("created_at") or ""),
            "updated_at": str((item or {}).get("updated_at") or ""),
        })
    clean_entries = clean_entries[-30:]
    record["notes"] = clean_entries
    record["note"] = "\n".join(item.get("text", "") for item in clean_entries if item.get("text"))
    return True


def update_followup_note(record, payload):
    if not isinstance(payload, dict):
        return False
    try:
        index = int(payload.get("index"))
    except (TypeError, ValueError):
        return False
    text = str(payload.get("text") or "").strip()
    if not text:
        return False
    entries = normalize_note_entries(record)
    if index < 0 or index >= len(entries):
        return False
    if entries[index].get("text") == text[:500]:
        return False
    entries[index]["text"] = text[:500]
    entries[index]["updated_at"] = now_iso()
    return set_followup_note_entries(record, entries)


def delete_followup_note(record, payload):
    try:
        index = int(payload.get("index") if isinstance(payload, dict) else payload)
    except (TypeError, ValueError):
        return False
    entries = normalize_note_entries(record)
    if index < 0 or index >= len(entries):
        return False
    entries.pop(index)
    return set_followup_note_entries(record, entries)


def append_weekly_followup(record, payload):
    try:
        week = int(payload.get("week"))
    except (TypeError, ValueError):
        return False
    if week < 1 or week > RENEWAL_WEEK_COUNT:
        return False
    methods = normalize_followup_methods(payload.get("methods"))
    if not methods:
        return False
    entry = {
        "date": normalize_followup_date(payload.get("date")),
        "methods": methods,
        "created_at": now_iso(),
    }
    weekly_followups = normalize_weekly_followups(record.get("weekly_followups"))
    week_records = weekly_followups.setdefault(str(week), [])
    existing = next((item for item in week_records if item.get("date") == entry["date"]), None)
    if existing is None:
        week_records.append(entry)
    else:
        existing["methods"] = methods
        existing["created_at"] = entry["created_at"]
    record["weekly_followups"] = weekly_followups
    record["followed_at"] = entry["created_at"]
    return True


def append_general_followup(record, payload):
    methods = normalize_followup_methods(payload.get("methods"))
    if not methods:
        return False
    entry = {
        "date": normalize_followup_date(payload.get("date")),
        "methods": methods,
        "created_at": now_iso(),
    }
    general_followups = normalize_general_followups(record.get("general_followups"))
    existing = next((item for item in general_followups if item.get("date") == entry["date"]), None)
    if existing is None:
        general_followups.append(entry)
    else:
        existing["methods"] = methods
        existing["created_at"] = entry["created_at"]
    record["general_followups"] = normalize_general_followups(general_followups)
    record["followed_at"] = entry["created_at"]
    return True


RENEWAL_UPLOAD_COLUMN_ALIASES = {
    "teacher": ["班主任", "老师", "带班老师", "组员"],
    "class_name": ["班级名称", "班级名", "班级"],
    "stage": ["续费阶段", "阶段", "当前阶段"],
    "student_name": ["学员姓名", "学生姓名", "姓名", "学员"],
    "student_account": ["学员账号", "学习账号", "学生账号", "账号", "手机号"],
    "followup_status": ["铺垫情况", "跟进情况", "意向情况", "意向", "铺垫电话"],
    "enrolled": ["是否报名", "已报名", "报名状态"],
    "blocker": ["当前卡点", "卡点", "顾虑", "当前顾虑"],
    "week": ["跟进周数", "周数", "第几周", "续费周数"],
    "date": ["跟进时间", "跟进日期", "时间", "日期"],
    "methods": ["跟进方式", "沟通方式", "方式"],
    "note": ["备注", "跟进备注", "沟通记录", "记录"],
}


def upload_cell_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_upload_header(value):
    return re.sub(r"[\s\ufeff:：()（）【】\[\]_-]+", "", str(value or "").lower())


def renewal_upload_column_key(header):
    normalized = normalize_upload_header(header)
    if not normalized:
        return ""
    for key, aliases in RENEWAL_UPLOAD_COLUMN_ALIASES.items():
        for alias in aliases:
            alias_key = normalize_upload_header(alias)
            if alias_key and (normalized == alias_key or alias_key in normalized):
                return key
    return ""


def mapped_upload_rows(rows, sheet_name=""):
    best_score = 0
    header_index = None
    header_map = {}
    for index, row in enumerate(rows[:20]):
        current_map = {}
        for column_index, value in enumerate(row):
            key = renewal_upload_column_key(value)
            if key and key not in current_map:
                current_map[key] = column_index
        score = len(current_map)
        if "class_name" in current_map:
            score += 3
        if "student_account" in current_map or "student_name" in current_map:
            score += 3
        if score > best_score:
            best_score = score
            header_index = index
            header_map = current_map
    if header_index is None or "class_name" not in header_map or not (
        "student_account" in header_map or "student_name" in header_map
    ):
        return []

    records = []
    for row_number, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        if not any(upload_cell_text(value) for value in row):
            continue
        record = {
            key: upload_cell_text(row[column_index]) if column_index < len(row) else ""
            for key, column_index in header_map.items()
        }
        if not (
            record.get("class_name")
            or record.get("student_account")
            or record.get("student_name")
            or record.get("note")
        ):
            continue
        record["_row_number"] = row_number
        record["_sheet_name"] = sheet_name
        records.append(record)
    return records


def parse_renewal_upload_file(file_storage):
    filename = (file_storage.filename or "").lower()
    file_storage.stream.seek(0)
    if filename.endswith(".csv"):
        raw_content = file_storage.stream.read()
        try:
            text = raw_content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw_content.decode("gb18030", errors="ignore")
        rows = [
            [upload_cell_text(value) for value in row]
            for row in csv.reader(io.StringIO(text))
        ]
        return mapped_upload_rows(rows, "CSV")

    if filename.endswith(".xlsx"):
        workbook = load_workbook(file_storage.stream, read_only=True, data_only=True)
        records = []
        for worksheet in workbook.worksheets:
            rows = [
                [upload_cell_text(value) for value in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            records.extend(mapped_upload_rows(rows, worksheet.title))
        return records

    raise ValueError("请上传 .xlsx 或 .csv 格式的续费历史数据表。")


def teacher_id_from_upload(value):
    text = str(value or "").strip()
    if not text:
        return ""
    normalized_id = normalize_teacher_id(text) or teacher_id_for_username(text)
    if normalized_id:
        return normalized_id
    text_key = normalize_match_text(text)
    for teacher in TEACHERS:
        candidates = [teacher.get("id"), teacher.get("username"), teacher.get("name"), *teacher.get("aliases", [])]
        if any(normalize_match_text(candidate) == text_key for candidate in candidates):
            return teacher.get("id", "")
    return ""


def normalize_upload_stage(value):
    stage = str(value or "").strip()
    if stage in RENEWAL_STAGES:
        return stage
    if "次月" in stage:
        return RENEWAL_SECOND_MONTH_STAGE
    if "首月" in stage:
        return RENEWAL_FIRST_MONTH_STAGE
    if "结营" in stage:
        return "结营续报"
    if "铺垫" in stage:
        return "铺垫阶段"
    return ""


def normalize_upload_status(value):
    status = str(value or "").strip()
    if status in FOLLOWUP_STATUSES:
        return status
    if any(keyword in status for keyword in ("愿意", "继续", "高意向", "想报")):
        return "愿意继续学"
    if "考虑" in status:
        return "需要考虑"
    if "拒绝" in status:
        return "拒绝"
    if "未接" in status or "没接" in status or "未通" in status:
        return "未接听"
    return ""


def normalize_upload_methods(value):
    text = str(value or "").strip()
    if not text:
        return []
    methods = []
    if any(keyword in text for keyword in ("私信", "微信", "私聊")):
        methods.append("私信")
    if any(keyword in text for keyword in ("电话", "去电", "拨打")):
        methods.append("电话")
    if not methods:
        methods = normalize_followup_methods(text)
    return methods


def parse_upload_bool(value):
    text = str(value or "").strip().lower()
    if not text:
        return None
    if text in {"1", "true", "yes", "y", "是", "已报名", "报名", "已报"}:
        return True
    if text in {"0", "false", "no", "n", "否", "未报名", "未报", "没有"}:
        return False
    return None


def parse_upload_week(value):
    text = str(value or "").strip()
    if not text:
        return None
    chinese_weeks = {"一": 1, "二": 2, "三": 3, "四": 4}
    for key, week in chinese_weeks.items():
        if key in text:
            return week
    match = re.search(r"[1-4]", text)
    if match:
        return int(match.group(0))
    return None


def clean_upload_note(value):
    text = str(value or "").strip()
    if not text:
        return ""
    table_markers = ["学员姓名", "学员账号", "平均完课", "跟进时间", "跟进情况", "是否报名"]
    marker_count = sum(1 for marker in table_markers if marker in text)
    if marker_count >= 3:
        return ""
    return text[:500]


def normalize_upload_blocker(value, store):
    option = normalize_blocker_option(value)
    if not option:
        return ""
    if option not in blocker_options(store):
        custom_options = [
            normalize_blocker_option(item)
            for item in store.get("blocker_options", [])
            if normalize_blocker_option(item)
        ]
        custom_options.append(option)
        store["blocker_options"] = list(dict.fromkeys(custom_options))
    return option


def upload_class_match_keys(item):
    keys = set()
    for value in (item.get("name"), item.get("note")):
        normalized = normalize_match_text(value)
        if normalized:
            keys.add(normalized)
    return keys


def find_upload_class(classes_by_id, class_name, teacher_id=""):
    class_key = normalize_match_text(class_name)
    if not class_key:
        return None, "缺少班级名称"
    candidates = [
        item
        for item in classes_by_id.values()
        if class_key in upload_class_match_keys(item)
    ]
    if teacher_id:
        teacher_matches = [item for item in candidates if class_teacher_id(item) == teacher_id]
        if teacher_matches:
            candidates = teacher_matches
    if len(candidates) == 1:
        return candidates[0], ""
    if not candidates:
        return None, "没有在完课班级中匹配到班级"
    return None, "匹配到多个同名班级，请补充班主任"


def find_student_for_upload(source_class, account, name):
    account_key = normalize_match_text(account)
    name_key = normalize_match_text(name)
    students = source_class.get("students", [])
    if account_key:
        for student in students:
            student_account = student.get("account") or student.get("phone")
            if normalize_match_text(student_account) == account_key:
                return student, ""
    if name_key:
        matches = [
            student
            for student in students
            if normalize_match_text(student.get("name")) == name_key
        ]
        if len(matches) == 1:
            return matches[0], ""
        if len(matches) > 1:
            return None, "匹配到多个同名学员，请填写学员账号"
    return None, "没有在班级中匹配到学员"


def ensure_upload_project(store, source_class, stage):
    class_id = source_class.get("id")
    existing = next(
        (project for project in store.get("projects", []) if project.get("class_id") == class_id),
        None,
    )
    if existing:
        return existing, False

    teacher_id = class_teacher_id(source_class)
    project = {
        "id": uuid.uuid4().hex,
        "class_id": class_id,
        "class_name": source_class.get("name", ""),
        "owner": source_class.get("owner", current_owner()),
        "teacher_id": teacher_id,
        "stage": stage or RENEWAL_STAGES[0],
        "locked_student_count": source_class_student_count(source_class),
        "student_count_note": "",
        "student_followups": {},
        "note": "",
        "created_by": current_owner(),
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    store.setdefault("projects", []).append(project)
    return project, True


def skipped_upload_label(record):
    location = str(record.get("_sheet_name") or "").strip()
    row_number = record.get("_row_number")
    if location and row_number:
        return f"{location} 第{row_number}行"
    if row_number:
        return f"第{row_number}行"
    return "未知行"


def blocker_priority(value):
    blocker = str(value or "").strip()
    if not blocker:
        return (1, len(BLOCKER_OPTIONS), "")
    try:
        option_index = BLOCKER_OPTIONS.index(blocker)
    except ValueError:
        option_index = len(BLOCKER_OPTIONS)
    return (0, option_index, blocker)


def student_priority_key(student):
    status = student.get("followup_status") or ""
    blocker_bucket, blocker_index, blocker = blocker_priority(student.get("current_blocker"))
    return (
        1 if student.get("enrolled") else 0,
        blocker_bucket,
        blocker_index,
        blocker,
        FOLLOWUP_STATUS_PRIORITY.get(status, FOLLOWUP_STATUS_PRIORITY[""]),
        str(student.get("name") or ""),
        str(student.get("account") or ""),
    )


def serialize_source_class(item):
    teacher_id = class_teacher_id(item)
    return {
        "id": item.get("id", ""),
        "name": item.get("name", ""),
        "note": str(item.get("note") or "").strip(),
        "owner": item.get("owner", ""),
        "teacher_id": teacher_id,
        "teacher_name": teacher_label(teacher_id),
        "student_count": len(item.get("students", [])),
        "completion_activity": bool(item.get("completion_activity")),
    }


def leader_plan_counts(project, source_class=None):
    valid_ids = None
    if source_class:
        valid_ids = {str(student.get("id")) for student in source_class.get("students", []) if student.get("id")}
    total = 0
    pending = 0
    for student_id, record in (project.get("student_followups") or {}).items():
        if valid_ids is not None and str(student_id) not in valid_ids:
            continue
        if not isinstance(record, dict):
            continue
        has_plan = bool(
            str(record.get("leader_note") or "").strip()
            or str(record.get("leader_talk_text") or "").strip()
            or normalize_leader_action_type(record.get("leader_action_type")) == "去电"
        )
        if not has_plan:
            continue
        total += 1
        if not bool(record.get("leader_note_done")):
            pending += 1
    return {
        "leader_plan_count": total,
        "pending_leader_plan_count": pending,
    }


def serialize_project(project, classes_by_id):
    source_class = classes_by_id.get(project.get("class_id"))
    source_count = source_class_student_count(source_class)
    locked_count = project_student_count(project, source_class)
    teacher_id = (
        normalize_teacher_id(project.get("teacher_id"))
        or (class_teacher_id(source_class) if source_class else "")
    )
    output = {
        "id": project.get("id", ""),
        "class_id": project.get("class_id", ""),
        "class_name": project.get("class_name", ""),
        "class_note": "",
        "class_missing": source_class is None,
        "owner": project.get("owner", ""),
        "teacher_id": teacher_id,
        "teacher_name": teacher_label(teacher_id),
        "student_count": locked_count,
        "source_student_count": source_count,
        "student_count_note": str(project.get("student_count_note") or "").strip(),
        "completion_activity": False,
        "stage": normalize_stage(project.get("stage")),
        "note": str(project.get("note") or "").strip(),
        "can_edit": can_edit_project(project),
        "created_at": project.get("created_at", ""),
        "updated_at": project.get("updated_at", ""),
    }
    if source_class:
        output.update({
            "class_name": source_class.get("name", output["class_name"]),
            "class_note": str(source_class.get("note") or "").strip(),
            "source_student_count": source_count,
            "completion_activity": bool(source_class.get("completion_activity")),
            "owner": source_class.get("owner", output["owner"]),
            "teacher_id": class_teacher_id(source_class),
            "teacher_name": teacher_label(class_teacher_id(source_class)),
        })
    student_count = int(output.get("student_count") or 0)
    enrolled_count = len(enrolled_student_ids(project, source_class))
    output["enrolled_count"] = enrolled_count
    output["renewal_rate"] = round(enrolled_count / student_count * 100, 2) if student_count else None
    output.update(leader_plan_counts(project, source_class))
    return output


def serialize_project_detail(project, classes_by_id):
    output = serialize_project(project, classes_by_id)
    source_class = classes_by_id.get(project.get("class_id"))
    month_key = current_month_key()
    include_weekly_in_general = output.get("stage") in RENEWAL_SINGLE_FOLLOWUP_STAGES
    output["students"] = []
    if source_class:
        students = []
        for student in source_class.get("students", []):
            student_id = str(student.get("id") or "")
            if not student_id:
                continue
            followup = student_followup_record(project, student_id)
            students.append({
                "id": student.get("id", ""),
                "name": str(student.get("name") or "").strip(),
                "account": str(student.get("account") or student.get("phone") or "").strip(),
                "average_completion": calculate_monthly_completion(get_student_weeks(student, month_key)),
                "followup_time": format_followup_time(followup.get("followed_at")),
                "followup_status": normalize_followup_status(followup.get("status")),
                "current_blocker": normalize_blocker(followup.get("current_blocker")),
                "weekly_followups": serialize_weekly_followups(followup),
                "general_followup": serialize_general_followups(followup, include_weekly=include_weekly_in_general),
                "enrolled": bool(followup.get("enrolled")),
                "followup_note": note_history_text(followup),
                "followup_notes": normalize_note_entries(followup),
                "leader_action_type": normalize_leader_action_type(followup.get("leader_action_type")),
                "leader_note": str(followup.get("leader_note") or "").strip(),
                "leader_talk_keyword": str(followup.get("leader_talk_keyword") or "").strip(),
                "leader_talk_type": str(followup.get("leader_talk_type") or "").strip(),
                "leader_talk_title": str(followup.get("leader_talk_title") or "").strip(),
                "leader_talk_text": str(followup.get("leader_talk_text") or "").strip(),
                "leader_note_done": bool(followup.get("leader_note_done")),
                "leader_note_updated_at": format_followup_time(followup.get("leader_note_updated_at")),
                "leader_note_done_at": format_followup_time(followup.get("leader_note_done_at")),
            })
        if output.get("stage") in RENEWAL_PRIORITY_STAGES:
            students.sort(key=student_priority_key)
        output["students"] = students
    return output


def visible_projects(store):
    if can_manage_accounts():
        return store.get("projects", [])
    owner = current_owner()
    return [project for project in store.get("projects", []) if project.get("owner") == owner]


def project_summary(projects):
    counts = {stage: 0 for stage in RENEWAL_STAGES}
    for project in projects:
        counts[normalize_stage(project.get("stage"))] += 1
    return {
        "total": len(projects),
        "stage_counts": counts,
    }


def teacher_overview(projects, classes_by_id):
    if not can_manage_accounts():
        return []
    teachers = {}
    for item in classes_by_id.values():
        teacher_id = class_teacher_id(item)
        if not teacher_id:
            continue
        entry = teachers.setdefault(teacher_id, {
            "teacher_id": teacher_id,
            "teacher_name": teacher_label(teacher_id),
            "class_count": 0,
            "project_count": 0,
            "student_count": 0,
            "enrolled_count": 0,
            "pending_leader_plan_count": 0,
            "stage_counts": {stage: 0 for stage in RENEWAL_STAGES},
        })
        entry["class_count"] += 1
    for project in projects:
        teacher_id = normalize_teacher_id(project.get("teacher_id"))
        if not teacher_id:
            continue
        entry = teachers.setdefault(teacher_id, {
            "teacher_id": teacher_id,
            "teacher_name": teacher_label(teacher_id),
            "class_count": 0,
            "project_count": 0,
            "student_count": 0,
            "enrolled_count": 0,
            "pending_leader_plan_count": 0,
            "stage_counts": {stage: 0 for stage in RENEWAL_STAGES},
        })
        entry["teacher_name"] = project.get("teacher_name") or entry["teacher_name"]
        entry["project_count"] += 1
        entry["student_count"] += int(project.get("student_count") or 0)
        entry["enrolled_count"] += int(project.get("enrolled_count") or 0)
        entry["pending_leader_plan_count"] += int(project.get("pending_leader_plan_count") or 0)
        entry["stage_counts"][normalize_stage(project.get("stage"))] += 1
    return sorted(
        teachers.values(),
        key=lambda item: (item.get("teacher_name", ""), item.get("teacher_id", "")),
    )


def build_payload():
    store = load_store()
    classes_by_id = class_lookup()
    changed = prune_store_followups(store, classes_by_id)
    for project in store.get("projects", []):
        changed = ensure_project_student_count_lock(
            project,
            classes_by_id.get(project.get("class_id")),
        ) or changed
    if changed:
        save_store(store)
    projects = [
        serialize_project(project, classes_by_id)
        for project in visible_projects(store)
    ]
    tracked_class_ids = {project.get("class_id") for project in store.get("projects", []) if project.get("class_id")}
    available_classes = [
        serialize_source_class(item)
        for item in classes_by_id.values()
        if can_add_class(item) and item.get("id") not in tracked_class_ids
    ]
    available_classes.sort(key=lambda item: (item.get("teacher_name", ""), item.get("name", "")))
    projects.sort(key=lambda item: (RENEWAL_STAGES.index(item["stage"]), item.get("teacher_name", ""), item.get("class_name", "")))
    return {
        "stages": RENEWAL_STAGES,
        "followup_statuses": FOLLOWUP_STATUSES,
        "followup_methods": FOLLOWUP_METHODS,
        "blocker_options": blocker_options(store),
        "projects": projects,
        "available_classes": available_classes,
        "summary": project_summary(projects),
        "teacher_overview": teacher_overview(projects, classes_by_id),
        "can_manage_all": can_manage_accounts(),
        "current_teacher_id": current_teacher_id(),
    }


@renewal_bp.get("")
@login_required
def renewal_home():
    return jsonify(build_payload())


@renewal_bp.post("/blockers")
@login_required
def create_blocker_option():
    if not can_manage_accounts():
        return jsonify({"error": "只有管理员可以新增当前卡点选项。"}), 403
    payload = request.get_json(silent=True) or {}
    option = normalize_blocker_option(payload.get("option"))
    if not option:
        return jsonify({"error": "请先填写要新增的当前卡点。"}), 400

    store = load_store()
    existing_options = blocker_options(store)
    if option not in existing_options:
        custom_options = [
            normalize_blocker_option(item)
            for item in store.get("blocker_options", [])
            if normalize_blocker_option(item)
        ]
        custom_options.append(option)
        store["blocker_options"] = list(dict.fromkeys(custom_options))
        save_store(store)
    return jsonify(build_payload()), 201


@renewal_bp.post("/projects")
@login_required
def create_project():
    payload = request.get_json(silent=True) or {}
    class_id = str(payload.get("class_id") or "").strip()
    if not class_id:
        return jsonify({"error": "请先选择要加入续费项目的班级。"}), 400

    classes_by_id = class_lookup()
    source_class = classes_by_id.get(class_id)
    if source_class is None or not can_add_class(source_class):
        return jsonify({"error": "只能添加完课-我的班级里的班级。"}), 404

    store = load_store()
    if any(project.get("class_id") == class_id for project in store.get("projects", [])):
        return jsonify({"error": "这个班级已经在续费项目里了。"}), 400

    teacher_id = class_teacher_id(source_class)
    project = {
        "id": uuid.uuid4().hex,
        "class_id": class_id,
        "class_name": source_class.get("name", ""),
        "owner": source_class.get("owner", current_owner()),
        "teacher_id": teacher_id,
        "stage": normalize_stage(payload.get("stage")),
        "locked_student_count": source_class_student_count(source_class),
        "student_count_note": "",
        "student_followups": {},
        "note": "",
        "created_by": current_owner(),
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    store.setdefault("projects", []).append(project)
    save_store(store)
    return jsonify(build_payload()), 201


@renewal_bp.patch("/projects/<project_id>")
@login_required
def update_project(project_id):
    payload = request.get_json(silent=True) or {}
    store = load_store()
    project = next((item for item in store.get("projects", []) if item.get("id") == project_id), None)
    if project is None or not can_edit_project(project):
        return jsonify({"error": "续费项目不存在。"}), 404

    if "stage" in payload:
        project["stage"] = normalize_stage(payload.get("stage"))
    if "note" in payload:
        project["note"] = str(payload.get("note") or "").strip()[:500]
    if "student_count" in payload or "locked_student_count" in payload:
        next_count = payload.get("student_count", payload.get("locked_student_count"))
        project["locked_student_count"] = normalize_locked_student_count(next_count)
    if "student_count_note" in payload:
        project["student_count_note"] = str(payload.get("student_count_note") or "").strip()[:300]
    project["updated_at"] = now_iso()
    save_store(store)
    return jsonify(build_payload())


@renewal_bp.get("/projects/<project_id>")
@login_required
def get_project(project_id):
    store = load_store()
    project = find_project(store, project_id)
    if project is None or not can_read_project(project):
        return jsonify({"error": "续费项目不存在。"}), 404
    class_store = load_class_store()
    classes_by_id = {item.get("id"): item for item in class_store.get("classes", []) if item.get("id")}
    source_class = classes_by_id.get(project.get("class_id"))
    changed = prune_project_followups(project, source_class)
    changed = ensure_project_student_count_lock(project, source_class) or changed
    if changed:
        save_store(store)
    return jsonify({"project": serialize_project_detail(project, classes_by_id)})


@renewal_bp.patch("/projects/<project_id>/students/<student_id>")
@login_required
def update_student_enrollment(project_id, student_id):
    payload = request.get_json(silent=True) or {}
    store = load_store()
    project = find_project(store, project_id)
    if project is None or not can_edit_project(project):
        return jsonify({"error": "续费项目不存在。"}), 404

    class_store = load_class_store()
    classes_by_id = {item.get("id"): item for item in class_store.get("classes", []) if item.get("id")}
    source_class = classes_by_id.get(project.get("class_id"))
    if source_class is None:
        return jsonify({"error": "该续费班级已不在完课班级列表中。"}), 404
    student = next(
        (item for item in source_class.get("students", []) if str(item.get("id")) == str(student_id)),
        None,
    )
    if student is None:
        return jsonify({"error": "学员不存在。"}), 404

    record = student_followup_record(project, student_id)
    had_update = False
    class_had_update = False
    if "student_name" in payload:
        next_name = str(payload.get("student_name") or "").strip()[:80]
        if not next_name:
            return jsonify({"error": "请输入学员姓名。"}), 400
        if next_name != str(student.get("name") or "").strip():
            updated_at = now_iso()
            student["name"] = next_name
            student["updated_at"] = updated_at
            source_class["updated_at"] = updated_at
            class_had_update = True
    if "followup_status" in payload:
        record["status"] = normalize_followup_status(payload.get("followup_status"))
        had_update = True
    elif "enrolled" in payload:
        record["enrolled"] = bool(payload.get("enrolled"))
        had_update = True
    elif "current_blocker" in payload:
        record["current_blocker"] = normalize_blocker(payload.get("current_blocker"))
        had_update = True
    elif "weekly_followup" in payload and isinstance(payload.get("weekly_followup"), dict):
        had_update = append_weekly_followup(record, payload.get("weekly_followup"))
    elif "general_followup" in payload and isinstance(payload.get("general_followup"), dict):
        had_update = append_general_followup(record, payload.get("general_followup"))
    if "followup_note" in payload:
        had_update = append_followup_note(record, payload.get("followup_note")) or had_update
    if "followup_note_replace" in payload:
        had_update = replace_followup_note(record, payload.get("followup_note_replace")) or had_update
    if "followup_note_update" in payload:
        had_update = update_followup_note(record, payload.get("followup_note_update")) or had_update
    if "followup_note_delete" in payload:
        had_update = delete_followup_note(record, payload.get("followup_note_delete")) or had_update
    if "leader_note" in payload:
        if not can_manage_accounts():
            return jsonify({"error": "只有管理员可以填写盘单。"}), 403
        next_note = str(payload.get("leader_note") or "").strip()[:500]
        if next_note != record.get("leader_note"):
            record["leader_note"] = next_note
            record["leader_note_updated_at"] = now_iso() if next_note else ""
            record["leader_note_done"] = False
            record["leader_note_done_at"] = ""
            had_update = True
    if "leader_action_type" in payload:
        if not can_manage_accounts():
            return jsonify({"error": "只有管理员可以设置盘单类型。"}), 403
        next_action_type = normalize_leader_action_type(payload.get("leader_action_type"))
        if next_action_type != record.get("leader_action_type"):
            record["leader_action_type"] = next_action_type
            has_plan_content = bool(record.get("leader_note") or record.get("leader_talk_text") or next_action_type == "去电")
            record["leader_note_updated_at"] = now_iso() if has_plan_content else ""
            record["leader_note_done"] = False
            record["leader_note_done_at"] = ""
            had_update = True
    if any(key in payload for key in ("leader_talk_keyword", "leader_talk_type", "leader_talk_title", "leader_talk_text")):
        if not can_manage_accounts():
            return jsonify({"error": "只有管理员可以选择盘单话术。"}), 403
        next_keyword = str(payload.get("leader_talk_keyword", record.get("leader_talk_keyword")) or "").strip()[:120]
        next_type = str(payload.get("leader_talk_type", record.get("leader_talk_type")) or "").strip()[:40]
        next_title = str(payload.get("leader_talk_title", record.get("leader_talk_title")) or "").strip()[:180]
        next_text = str(payload.get("leader_talk_text", record.get("leader_talk_text")) or "").strip()[:5000]
        if (
            next_keyword != record.get("leader_talk_keyword")
            or next_type != record.get("leader_talk_type")
            or next_title != record.get("leader_talk_title")
            or next_text != record.get("leader_talk_text")
        ):
            record["leader_talk_keyword"] = next_keyword
            record["leader_talk_type"] = next_type
            record["leader_talk_title"] = next_title
            record["leader_talk_text"] = next_text
            record["leader_note_updated_at"] = now_iso() if (record.get("leader_note") or next_text) else ""
            record["leader_note_done"] = False
            record["leader_note_done_at"] = ""
            had_update = True
    if "leader_note_done" in payload:
        record["leader_note_done"] = bool(payload.get("leader_note_done"))
        record["leader_note_done_at"] = now_iso() if record["leader_note_done"] else ""
        had_update = True
    if had_update and "weekly_followup" not in payload and "general_followup" not in payload:
        record["followed_at"] = now_iso()
    if had_update:
        project["updated_at"] = now_iso()
        save_store(store)
    if class_had_update:
        save_class_store(class_store)
    return jsonify({
        "project": serialize_project_detail(project, classes_by_id),
        "board": build_payload(),
    })


@renewal_bp.post("/history-upload")
@login_required
def upload_renewal_history():
    if not can_manage_accounts():
        return jsonify({"error": "只有管理员可以上传续费历史数据。"}), 403

    file_storage = request.files.get("file")
    if not file_storage or not file_storage.filename:
        return jsonify({"error": "请先选择要上传的续费历史数据表。"}), 400

    try:
        records = parse_renewal_upload_file(file_storage)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400
    except Exception:
        current_app.logger.exception("renewal history upload failed")
        return jsonify({"error": "续费历史数据解析失败，请确认表格格式。"}), 400

    if not records:
        return jsonify({"error": "没有识别到可导入的数据，请确认表头包含班级名称和学员姓名/账号。"}), 400

    store = load_store()
    classes_by_id = class_lookup()
    imported_count = 0
    created_project_count = 0
    touched_project_ids = set()
    skipped_rows = []

    for record in records:
        teacher_id = teacher_id_from_upload(record.get("teacher"))
        source_class, class_error = find_upload_class(classes_by_id, record.get("class_name"), teacher_id)
        if source_class is None:
            skipped_rows.append({
                "row": skipped_upload_label(record),
                "reason": class_error,
                "class_name": record.get("class_name", ""),
                "student": record.get("student_name") or record.get("student_account") or "",
            })
            continue

        student, student_error = find_student_for_upload(
            source_class,
            record.get("student_account"),
            record.get("student_name"),
        )
        if student is None:
            skipped_rows.append({
                "row": skipped_upload_label(record),
                "reason": student_error,
                "class_name": source_class.get("name", ""),
                "student": record.get("student_name") or record.get("student_account") or "",
            })
            continue

        stage = normalize_upload_stage(record.get("stage"))
        project, created = ensure_upload_project(store, source_class, stage)
        if created:
            created_project_count += 1

        student_id = str(student.get("id") or "")
        followup = student_followup_record(project, student_id)
        had_update = False
        project_stage = normalize_stage(project.get("stage"))

        status = normalize_upload_status(record.get("followup_status"))
        if status:
            followup["status"] = normalize_followup_status(status)
            had_update = True

        enrolled = parse_upload_bool(record.get("enrolled"))
        if enrolled is not None:
            followup["enrolled"] = enrolled
            had_update = True

        blocker = normalize_upload_blocker(record.get("blocker"), store)
        if blocker:
            followup["current_blocker"] = blocker
            had_update = True

        methods = normalize_upload_methods(record.get("methods"))
        if methods:
            week = parse_upload_week(record.get("week"))
            followup_payload = {
                "date": record.get("date"),
                "methods": methods,
            }
            if project_stage in RENEWAL_FOUR_WEEK_STAGES and week:
                followup_payload["week"] = week
                had_update = append_weekly_followup(followup, followup_payload) or had_update
            else:
                had_update = append_general_followup(followup, followup_payload) or had_update
        elif record.get("date"):
            followup["followed_at"] = f"{normalize_followup_date(record.get('date'))}T00:00:00"
            had_update = True

        note = clean_upload_note(record.get("note"))
        if note:
            had_update = append_followup_note(followup, note) or had_update

        if had_update:
            imported_count += 1
            project["updated_at"] = now_iso()
            touched_project_ids.add(project.get("id", ""))

    if imported_count or created_project_count:
        save_store(store)

    return jsonify({
        "ok": True,
        "imported_count": imported_count,
        "created_project_count": created_project_count,
        "updated_project_count": len([item for item in touched_project_ids if item]),
        "skipped_count": len(skipped_rows),
        "skipped_rows": skipped_rows[:12],
        "board": build_payload(),
    })


@renewal_bp.delete("/projects/<project_id>")
@login_required
def delete_project(project_id):
    store = load_store()
    before = len(store.get("projects", []))
    store["projects"] = [
        project
        for project in store.get("projects", [])
        if not (project.get("id") == project_id and can_edit_project(project))
    ]
    if len(store["projects"]) == before:
        return jsonify({"error": "续费项目不存在。"}), 404
    save_store(store)
    return jsonify(build_payload())
